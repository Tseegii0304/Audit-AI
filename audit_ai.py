"""
Audit AI v10.0 — Аудитын эрсдэлийг бууруулах хиймэл оюун, машин сургалтад суурилсан загвар
Гүйлгээний баланс + Ерөнхий журнал + Машин сургалт + Тайлбарлагдах ХОУ (XAI)
Эх сурвалж: УЛАМБАЯРЫН ЦЭЦЭГЖАРГАЛ — Бизнесийн удирдлагын ухааны докторын зэрэг горилсон бүтээл, МУИС-БС, 2026

pip install -r requirements.txt
streamlit run audit_ai.py
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from sklearn.ensemble import IsolationForest, RandomForestClassifier, GradientBoostingClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.model_selection import StratifiedKFold, cross_val_predict
from sklearn.metrics import precision_score, recall_score, f1_score, roc_auc_score, roc_curve
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet
from scipy import stats
import warnings, io, re, gzip, zipfile
from pathlib import Path
from datetime import datetime
from collections import Counter
warnings.filterwarnings('ignore')
from tab_descriptions import TabDescriptions

# ═══ GOV AUDIT ТУСГАЙ ФУНКЦҮҮД (v10.0-д нэмсэн) ═══
def normalize_ej(df, year=2025):
    """ЕЖ-г стандарт формат руу хөрвүүлэх"""
    d = df.copy()
    col_map = {}
    for c in d.columns:
        cl = str(c).lower().strip()
        if cl in ('огноо','date','dte'): col_map[c] = 'огноо'
        elif cl in ('дансны дугаар','account','acc','данс'): col_map[c] = 'данс'
        elif cl in ('дебит','debit','dt'): col_map[c] = 'дебит'
        elif cl in ('кредит','credit','ct','cr'): col_map[c] = 'кредит'
        elif cl in ('үлдэгдэл','balance','bal'): col_map[c] = 'үлдэгдэл'
        elif cl in ('тайлбар','description','desc','утга'): col_map[c] = 'тайлбар'
        elif cl in ('харилцагч','partner','counterparty'): col_map[c] = 'харилцагч'
    if col_map: d.rename(columns=col_map, inplace=True)
    for nc in ['дебит','кредит','үлдэгдэл']:
        if nc in d.columns: d[nc] = pd.to_numeric(d[nc], errors='coerce').fillna(0)
    d['year'] = year
    return d

def explain_row(r):
    reasons = []
    if r.get('is_round',0)==1: reasons.append("Тэгш дүн (ISA 240)")
    if r.get('benford_dev',0)>0.05: reasons.append("Бенфордын зөрүү (ISA 240)")
    if r.get('is_dup',0)==1: reasons.append("Давхардсан гүйлгээ (ISA 500)")
    if r.get('desc_mismatch',0)==1: reasons.append("Тайлбар зөрүүтэй (ISA 315)")
    if r.get('is_weekend',0)==1: reasons.append("Амралтын өдөр (ISA 240)")
    return " | ".join(reasons) if reasons else "Хэвийн"

from branch_ml_module import (
    BRANCH_REGISTRY, detect_branch, get_branch_summary,
    create_pseudo_labels, run_train_test_ml,
    run_branch_comparison, compare_unsupervised_supervised,
    run_cross_branch_patterns, run_learning_curve,
    run_hyperparameter_search, run_stability_analysis,
    run_mcnemar_test, run_contamination_sensitivity,
    generate_dissertation_tables,
    compute_detection_risk, ISA_FEATURE_MAP, get_isa_feature_report,
    run_benford_analysis, BENFORD_EXPECTED, compute_risk_score
)
try:
    import shap
except Exception:
    shap = None

try:
    from statsmodels.tsa.seasonal import seasonal_decompose
except Exception:
    seasonal_decompose = None

td = TabDescriptions()
st.set_page_config(page_title="Audit AI v10.0 — Аудитын хиймэл оюуны систем", page_icon="🛡️", layout="wide")

# Gov audit session state initialization
for _gk in ['g_expense_class','g_budget','g_cash','g_ej','g_laws','g_ej_done','g_ej_fi','g_ej_ml','g_ej_models','g_ej_xai','g_ez','g_policy']:
    if _gk not in st.session_state: st.session_state[_gk] = None

# ═══════════════════════════════════════════════════════════
# 🎨 ДИЗАЙН СИСТЕМ — Custom CSS Theme
# ═══════════════════════════════════════════════════════════
THEME_CSS = """
<style>
/* ── Midnight Audit Design System v11.0 ── */
@import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@400;500;600;700&family=Outfit:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');

:root {
    --primary: #162842;
    --primary-light: #1E3654;
    --primary-dark: #0B1423;
    --accent: #D4A017;
    --accent-light: #E8B831;
    --accent-dark: #B8860B;
    --success: #5B8A72;
    --success-light: #E8F5E9;
    --warning: #D4A017;
    --warning-light: #FCF0CB;
    --danger: #C7556E;
    --danger-light: #FCE4EC;
    --info: #4A8EC2;
    --info-light: #E3F2FD;
    --bg-main: #FDFCF8;
    --bg-card: #FFFFFF;
    --text-primary: #0B1423;
    --text-secondary: #475569;
    --text-muted: #94A3B8;
    --border: #E5DFD0;
    --border-light: #F2EEE4;
    --shadow-sm: 0 1px 4px rgba(6,11,20,0.06);
    --shadow-md: 0 4px 16px rgba(6,11,20,0.10);
    --shadow-lg: 0 10px 32px rgba(6,11,20,0.14);
    --shadow-gold: 0 4px 16px rgba(212,160,23,0.12);
    --radius-sm: 8px;
    --radius-md: 12px;
    --radius-lg: 16px;
    --font-display: 'Cormorant Garamond', Georgia, serif;
    --font-body: 'Outfit', 'Segoe UI', sans-serif;
    --font-mono: 'IBM Plex Mono', 'Fira Code', monospace;
}

/* ── Global Typography ── */
.stApp, .stApp * { font-family: var(--font-body) !important; }
h1, h2, h3 { font-family: var(--font-display) !important; color: var(--primary-dark) !important; font-weight: 600 !important; letter-spacing: -0.3px; }
h1 { font-size: 32px !important; font-weight: 700 !important; }

/* ── Main App Background ── */
.stApp { background: var(--bg-main) !important; }
.main .block-container {
    padding-top: 1rem !important; padding-left: 2.5rem !important;
    padding-right: 2.5rem !important; max-width: 100% !important;
}
@media (min-width: 1200px) { .main .block-container { padding-left: 3rem !important; padding-right: 3rem !important; } }
@media (max-width: 768px) { .main .block-container { padding-left: 1rem !important; padding-right: 1rem !important; }
    div[data-testid="stMetric"] [data-testid="stMetricValue"] { font-size: 20px !important; } }

/* ── Sidebar — Dark sapphire gradient ── */
section[data-testid="stSidebar"] {
    background: linear-gradient(175deg, #060B14 0%, #0B1423 40%, #162842 100%) !important;
    border-right: 1px solid rgba(212,160,23,0.08) !important;
    box-shadow: var(--shadow-lg);
}
section[data-testid="stSidebar"] * { color: #FFFFFF !important; }
section[data-testid="stSidebar"] .stMarkdown p,
section[data-testid="stSidebar"] .stMarkdown span,
section[data-testid="stSidebar"] .stCaption,
section[data-testid="stSidebar"] small { color: rgba(255,255,255,0.55) !important; }
section[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.06) !important; }
section[data-testid="stSidebar"] .stRadio label,
section[data-testid="stSidebar"] .stRadio p,
section[data-testid="stSidebar"] .stRadio span,
section[data-testid="stSidebar"] .stRadio div,
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p,
section[data-testid="stSidebar"] [role="radiogroup"] label,
section[data-testid="stSidebar"] [role="radiogroup"] label p,
section[data-testid="stSidebar"] [role="radiogroup"] label span,
section[data-testid="stSidebar"] [role="radiogroup"] label div { color: rgba(255,255,255,0.65) !important; }
section[data-testid="stSidebar"] [role="radiogroup"] label:hover { color: rgba(255,255,255,0.9) !important; }
section[data-testid="stSidebar"] [data-testid="stAlert"] {
    background: rgba(212,160,23,0.08) !important;
    border: 1px solid rgba(212,160,23,0.15) !important;
    border-radius: var(--radius-sm) !important;
}
section[data-testid="stSidebar"] [data-testid="stAlert"] p,
section[data-testid="stSidebar"] [data-testid="stAlert"] span { color: #F0CC5C !important; font-size: 13px !important; }

/* ── Metric Cards — Gold accent top bar ── */
div[data-testid="stMetric"] {
    background: var(--bg-card); border: 1px solid var(--border); border-radius: var(--radius-md);
    padding: 18px 22px; box-shadow: var(--shadow-sm); transition: all 0.3s ease;
    position: relative; overflow: hidden;
}
div[data-testid="stMetric"]::before {
    content: ''; position: absolute; top: 0; left: 0; right: 0; height: 2.5px;
    background: linear-gradient(90deg, var(--accent), var(--accent-light)); opacity: 0; transition: opacity 0.3s;
}
div[data-testid="stMetric"]:hover { box-shadow: var(--shadow-md); transform: translateY(-2px); }
div[data-testid="stMetric"]:hover::before { opacity: 1; }
div[data-testid="stMetric"] label {
    color: var(--text-secondary) !important; font-size: 11px !important;
    font-weight: 600 !important; text-transform: uppercase; letter-spacing: 0.5px;
}
div[data-testid="stMetric"] [data-testid="stMetricValue"] {
    font-family: var(--font-display) !important; color: var(--primary-dark) !important;
    font-weight: 700 !important; font-size: 30px !important;
}
div[data-testid="stMetric"] [data-testid="stMetricDelta"] { font-size: 12px !important; }

/* ── Tabs — Refined pill style ── */
.stTabs [data-baseweb="tab-list"] {
    gap: 2px; background: var(--border-light); border-radius: var(--radius-md); padding: 3px;
}
.stTabs [data-baseweb="tab"] {
    border-radius: var(--radius-sm) !important; padding: 8px 20px !important;
    font-weight: 500 !important; color: var(--text-secondary) !important;
    background: transparent !important; border: none !important; transition: all 0.2s ease;
}
.stTabs [aria-selected="true"] {
    background: var(--bg-card) !important; color: var(--primary-dark) !important;
    box-shadow: var(--shadow-sm) !important; font-weight: 600 !important;
}

/* ── Buttons — Gold primary ── */
.stButton > button[kind="primary"],
.stButton > button[data-testid="stBaseButton-primary"] {
    background: linear-gradient(135deg, #D4A017 0%, #E8B831 100%) !important;
    color: #0B1423 !important; border: none !important;
    border-radius: var(--radius-sm) !important; font-weight: 700 !important;
    padding: 10px 24px !important; box-shadow: 0 3px 12px rgba(212,160,23,0.25) !important;
    transition: all 0.2s ease !important; font-family: var(--font-body) !important;
}
.stButton > button[kind="primary"]:hover,
.stButton > button[data-testid="stBaseButton-primary"]:hover {
    box-shadow: 0 6px 20px rgba(212,160,23,0.35) !important; transform: translateY(-1px) !important;
}
.stButton > button[kind="secondary"],
.stButton > button[data-testid="stBaseButton-secondary"] {
    border: 1.5px solid var(--border) !important; border-radius: var(--radius-sm) !important;
    background: var(--bg-card) !important; color: var(--text-primary) !important;
    font-weight: 500 !important; transition: all 0.2s ease !important;
}
.stButton > button[kind="secondary"]:hover,
.stButton > button[data-testid="stBaseButton-secondary"]:hover {
    border-color: var(--accent) !important; color: var(--primary-dark) !important;
    box-shadow: var(--shadow-gold) !important;
}

/* ── Download Buttons — Sage accent ── */
.stDownloadButton > button {
    border: 1.5px solid #5B8A72 !important; color: #5B8A72 !important;
    border-radius: var(--radius-sm) !important; background: #E8F5E9 !important; font-weight: 600 !important;
}
.stDownloadButton > button:hover { background: #5B8A72 !important; color: white !important; }

/* ── DataFrames ── */
.stDataFrame { border: 1px solid var(--border) !important; border-radius: var(--radius-md) !important; overflow: hidden; box-shadow: var(--shadow-sm); }

/* ── Expanders ── */
.streamlit-expanderHeader {
    background: var(--border-light) !important; border-radius: var(--radius-sm) !important;
    font-weight: 500 !important; color: var(--text-primary) !important; border: 1px solid var(--border) !important;
}

/* ── Info/Warning/Success Boxes ── */
div[data-testid="stAlert"] { border-radius: var(--radius-sm) !important; border-width: 1px !important; }

/* ── File Uploader — Gold hover ── */
section[data-testid="stFileUploader"] {
    border: 2px dashed var(--border) !important; border-radius: var(--radius-md) !important;
    padding: 10px !important; background: var(--border-light) !important; transition: all 0.3s;
}
section[data-testid="stFileUploader"]:hover { border-color: var(--accent) !important; background: rgba(212,160,23,0.03) !important; }

/* ── Slider & Progress — Gold ── */
.stSlider > div > div > div { color: var(--accent-dark) !important; }
.stProgress > div > div > div > div { background: linear-gradient(90deg, var(--accent-dark) 0%, var(--accent) 100%) !important; }

/* ── Selectbox ── */
div[data-baseweb="select"] { border-radius: var(--radius-sm) !important; }

/* ── Horizontal Rule ── */
hr { border-top: 1px solid var(--border) !important; margin: 1.5rem 0 !important; }

/* ── Custom Card Class — with gold hover glow ── */
.audit-card {
    background: var(--bg-card); border: 1px solid var(--border); border-radius: var(--radius-lg);
    padding: 24px; margin: 12px 0; box-shadow: var(--shadow-sm); transition: all 0.3s;
}
.audit-card:hover { box-shadow: var(--shadow-gold); border-color: rgba(212,160,23,0.2); }

/* ── Risk Badge — Refined palette ── */
.risk-badge { display: inline-block; padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: 600; letter-spacing: 0.3px; }
.risk-high { background: #FCE4EC; color: #C7556E; }
.risk-medium { background: #FCF0CB; color: #B8860B; }
.risk-low { background: #E8F5E9; color: #5B8A72; }

/* ── ISA Reference Tag — Amber ── */
.isa-tag {
    display: inline-block; background: rgba(212,160,23,0.1); color: #B8860B;
    padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; margin: 0 2px;
}

/* ── Hero Banner (new) ── */
.hero-banner {
    background: linear-gradient(135deg, #060B14, #162842); border-radius: 16px;
    padding: 28px 32px; color: white; margin-bottom: 20px; position: relative; overflow: hidden;
}
.hero-banner::before {
    content: ''; position: absolute; top: -50%; right: -30%; width: 60%; height: 200%;
    background: radial-gradient(ellipse, rgba(212,160,23,0.08), transparent 70%); pointer-events: none;
}
.hero-banner h2 { font-family: var(--font-display) !important; font-size: 48px !important; font-weight: 700 !important; color: #E8B831 !important; }
.hero-banner p { font-size: 14px; color: rgba(255,255,255,0.5); }

/* ── ISA Compliance Badge (topbar) ── */
.isa-badge {
    display: inline-block; padding: 4px 14px;
    background: linear-gradient(135deg, #D4A017, #E8B831);
    color: #0B1423; border-radius: 20px; font-size: 10px; font-weight: 700; letter-spacing: 0.5px;
}
</style>
"""
st.markdown(THEME_CSS, unsafe_allow_html=True)

# ── Толгой хэсэг ──
st.markdown("""
<div style="background: linear-gradient(135deg, #0F4C81 0%, #1976D2 50%, #0D47A1 100%);
     padding: 28px 32px; border-radius: 16px; margin-bottom: 24px;
     box-shadow: 0 8px 32px rgba(15,76,129,0.2);">
    <div style="display:flex; align-items:center; justify-content:center; gap:16px;">
        <div style="background:rgba(255,255,255,0.15); border-radius:12px; padding:12px 16px;">
            <span style="font-size:32px;">🔍</span>
        </div>
        <div>
            <h1 style="color:#FFFFFF !important; margin:0; font-size:28px; font-weight:700; letter-spacing:-0.5px;">
                Audit AI v10.0</h1>
            <p style="color:rgba(255,255,255,0.8); margin:4px 0 0; font-size:14px; font-weight:400;">
                Дашбоард • Шинжилгээ • Машин сургалт • Тайлан — Аудиторт зориулсан бүрэн систем</p>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

# ── Plotly загварын тохиргоо (бүх графикт нэгдсэн загвар) ──
PLOTLY_COLORS = ['#D4A017', '#4A8EC2', '#5B8A72', '#C7556E', '#162842', '#E8B831', '#6AAAD6', '#73A68A']
PLOTLY_TEMPLATE = dict(
    layout=dict(
        font=dict(family='Inter, sans-serif', color='#1E293B'),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(248,250,252,1)',
        title=dict(font=dict(size=15, color='#0F4C81')),
        legend=dict(bgcolor='rgba(255,255,255,0.8)', bordercolor='#E2E8F0', borderwidth=1),
        colorway=PLOTLY_COLORS,
        xaxis=dict(gridcolor='#E2E8F0', linecolor='#CBD5E1'),
        yaxis=dict(gridcolor='#E2E8F0', linecolor='#CBD5E1'),
        margin=dict(t=50, b=30, l=40, r=20),
    )
)

def apply_chart_style(fig):
    """Бүх Plotly графикт нэгдсэн загвар хэрэглэнэ."""
    fig.update_layout(
        font=dict(family='Inter, sans-serif', color='#1E293B'),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(248,250,252,1)',
        title_font=dict(size=15, color='#0F4C81'),
        legend=dict(bgcolor='rgba(255,255,255,0.9)', bordercolor='#E2E8F0', borderwidth=1,
                    font=dict(size=11)),
        xaxis=dict(gridcolor='#E2E8F0', linecolor='#CBD5E1', zeroline=False),
        yaxis=dict(gridcolor='#E2E8F0', linecolor='#CBD5E1', zeroline=False),
    )
    return fig

# ── Set Plotly defaults globally ──
import plotly.io as pio
pio.templates['audit'] = go.layout.Template(
    layout=go.Layout(
        font=dict(family='Inter, sans-serif', color='#1E293B', size=12),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(248,250,252,1)',
        title=dict(font=dict(size=15, color='#0F4C81')),
        colorway=PLOTLY_COLORS,
        xaxis=dict(gridcolor='#E2E8F0', linecolor='#CBD5E1', zeroline=False),
        yaxis=dict(gridcolor='#E2E8F0', linecolor='#CBD5E1', zeroline=False),
        legend=dict(bgcolor='rgba(255,255,255,0.9)', bordercolor='#E2E8F0', borderwidth=1),
        margin=dict(t=50, b=30, l=40, r=20),
    )
)
pio.templates.default = 'audit'


# ── Session state defaults ──
SESSION_DEFAULTS = {
    'tb_analysis_done': False,
    'journal_ai_done': False,
    'branch_done': False,
    'tb_detected_rows': [],
    'journal_detected_rows': [],
    'tb_upload_cache': {},
    'journal_upload_cache': {},
    'materiality_result': None,
    'materiality_total': 0,
    'branch_comparison': None,
    'branch_summary': None,
    'branch_detect_info': [],
}
for _k, _v in SESSION_DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

def uploaded_files_to_cache(files):
    cache = {}
    for f in files or []:
        try:
            raw = f.getvalue()
        except Exception:
            raw = f.read()
            f.seek(0)
        cache[f.name] = raw
    return cache

def cache_to_file_objects(cache):
    objs = []
    for name, raw in (cache or {}).items():
        bio = io.BytesIO(raw)
        bio.name = name
        objs.append(bio)
    return objs

with st.sidebar:
    st.markdown("""
    <div style="text-align:center; padding:12px 0 20px;">
        <div style="width:48px;height:48px;background:linear-gradient(135deg,#D4A017,#E8B831);border-radius:14px;display:inline-flex;align-items:center;justify-content:center;margin-bottom:10px;box-shadow:0 6px 20px rgba(212,160,23,0.25);">
            <span style="font-size:22px;">🛡️</span>
        </div>
        <p style="color:#FFFFFF; font-size:22px; font-weight:600; margin:0; font-family:'Cormorant Garamond',Georgia,serif; letter-spacing:0.3px;">Audit AI</p>
        <p style="color:rgba(255,255,255,0.35); font-size:10px; margin:0; letter-spacing:1.5px; text-transform:uppercase;">v10.0 · Audit AI</p>
    </div>
    """, unsafe_allow_html=True)
    page = st.radio("Цэс:", [
        "1️⃣ Өгөгдөл оруулах, бэлтгэх",
        "📊 Дашбоард",
        "2️⃣ Гүйлгээний балансын шинжилгээ",
        "3️⃣ Ерөнхий журналын шинжилгээ",
        "4️⃣ Материаллаг байдлын тооцоо",
        "5️⃣ Салбарын шинжилгээ",
        "6️⃣ Сургалт/Шалгалт (машин сургалт)",
        "7️⃣ Нарийвчилсан машин сургалт",
        "8️⃣ Диссертацийн гаралт",
        "🏛️ Мөнгөн гүйлгээний журнал",
        "📊 ЭЗ ангилал нийцэл",
        "⚖️ Хуулийн зөрчил шалгалт",
        "💰 Зардлын ангилал",
        "🔍 Харилцагч ISA 550",
        "📋 Эрсдэлийн нэгтгэл",
    ])
    # Quick status
    st.markdown("---")
    st.markdown("""
    <p style="color:#E8B831; font-size:9.5px; font-weight:700; text-transform:uppercase; letter-spacing:2px; margin-bottom:8px; opacity:0.55;">
        📁 ӨГӨГДЛИЙН ТӨЛӨВ
    </p>
    """, unsafe_allow_html=True)
    _tb_n = len(st.session_state.get('prepared_tb_cache', {}))
    _led_n = len(st.session_state.get('prepared_ledger_cache', {}))
    _p1_n = len(st.session_state.get('prepared_part1_cache', {}))
    st.markdown(f"""
    <div style="display:flex; gap:8px; margin-bottom:12px;">
        <div style="flex:1; background:rgba(212,160,23,0.08); border:1px solid rgba(212,160,23,0.12); border-radius:8px; padding:8px; text-align:center;">
            <div style="color:#F0CC5C; font-size:18px; font-weight:700; font-family:'Cormorant Garamond',Georgia,serif;">{_tb_n}</div>
            <div style="color:rgba(255,255,255,0.4); font-size:10px;">ГБ</div>
        </div>
        <div style="flex:1; background:rgba(212,160,23,0.08); border:1px solid rgba(212,160,23,0.12); border-radius:8px; padding:8px; text-align:center;">
            <div style="color:#F0CC5C; font-size:18px; font-weight:700; font-family:'Cormorant Garamond',Georgia,serif;">{_led_n}</div>
            <div style="color:rgba(255,255,255,0.4); font-size:10px;">ЕЖ</div>
        </div>
        <div style="flex:1; background:rgba(212,160,23,0.08); border:1px solid rgba(212,160,23,0.12); border-radius:8px; padding:8px; text-align:center;">
            <div style="color:#F0CC5C; font-size:18px; font-weight:700; font-family:'Cormorant Garamond',Georgia,serif;">{_p1_n}</div>
            <div style="color:rgba(255,255,255,0.4); font-size:10px;">Нэгтгэл</div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    if st.session_state.get('tb_analysis_done'):
        st.success("✅ ГБ шинжилгээ бэлэн", icon="✅")
    if st.session_state.get('journal_ai_done'):
        st.success("✅ ЕЖ шинжилгээ бэлэн", icon="✅")
    st.markdown("---")
    st.markdown("""
    <p style="color:rgba(255,255,255,0.4); font-size:10px; text-align:center; margin-top:8px;">
        Аудитын ХОУ диссертаци<br>© 2026
    </p>
    """, unsafe_allow_html=True)

ACCT_RE_B = re.compile(r'Данс:\s*\[([^\]]+)\]\s*(.*)')
ACCT_RE_P = re.compile(r'Данс:\s*(\d{3}-\d{2}-\d{2}-\d{3})\s+(.*)')

def parse_account(text):
    m = ACCT_RE_B.match(text)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    m = ACCT_RE_P.match(text)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return None, None

def safe_float(v):
    if v is None or v == '':
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0

def process_raw_tb(file_obj):
    import openpyxl
    base_cols = ['account_code','account_name','opening_debit','opening_credit','turnover_debit','turnover_credit','closing_debit','closing_credit']
    try:
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            cell0 = str(row[0]).strip()
            cell1 = str(row[1]).strip() if len(row) > 1 and row[1] else ''

            # ═══ Format 1: "101-01-02-001" даштай код (Col0=seq, Col1=code, Col2=name) ═══
            if re.match(r'\d{3}-', cell1):
                try: int(float(cell0))
                except: continue
                rows.append({
                    'account_code': cell1,
                    'account_name': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                    'opening_debit': safe_float(row[3]) if len(row) > 3 else 0.0,
                    'opening_credit': safe_float(row[4]) if len(row) > 4 else 0.0,
                    'turnover_debit': safe_float(row[5]) if len(row) > 5 else 0.0,
                    'turnover_credit': safe_float(row[6]) if len(row) > 6 else 0.0,
                    'closing_debit': safe_float(row[7]) if len(row) > 7 else 0.0,
                    'closing_credit': safe_float(row[8]) if len(row) > 8 else 0.0,
                })
                continue

            # ═══ Format 2: Төрийн сан — "120009 - Нэр" (Col0=code+name, Col1=code) ═══
            # Col1 нь цэвэр тоон код (5-6 оронтой), Col0 нь "код - нэр"
            if re.match(r'^\d{5,6}$', cell1):
                acct_code = cell1
                # Нэрийг Col0-оос гаргах
                if ' - ' in cell0:
                    acct_name = cell0.split(' - ', 1)[1].strip()
                elif ' – ' in cell0:
                    acct_name = cell0.split(' – ', 1)[1].strip()
                else:
                    acct_name = cell0.replace(cell1, '').strip(' -–')
                # Col2-3=Эхний үлдэгдэл, Col4-5=Тухайн сар, Col6-7=Өссөн дүн, Col8-9=Эцсийн үлдэгдэл
                rows.append({
                    'account_code': acct_code,
                    'account_name': acct_name,
                    'opening_debit': safe_float(row[2]) if len(row) > 2 else 0.0,
                    'opening_credit': safe_float(row[3]) if len(row) > 3 else 0.0,
                    'turnover_debit': safe_float(row[4]) if len(row) > 4 else 0.0,
                    'turnover_credit': safe_float(row[5]) if len(row) > 5 else 0.0,
                    'closing_debit': safe_float(row[8]) if len(row) > 8 else 0.0,
                    'closing_credit': safe_float(row[9]) if len(row) > 9 else 0.0,
                })
                continue

            # ═══ Format 3: Цэвэр тоон Col0 (6+ оронтой код шууд) ═══
            if re.match(r'^\d{6,}$', cell0):
                rows.append({
                    'account_code': cell0,
                    'account_name': str(row[1]).strip() if len(row) > 1 and row[1] else '',
                    'opening_debit': safe_float(row[2]) if len(row) > 2 else 0.0,
                    'opening_credit': safe_float(row[3]) if len(row) > 3 else 0.0,
                    'turnover_debit': safe_float(row[4]) if len(row) > 4 else 0.0,
                    'turnover_credit': safe_float(row[5]) if len(row) > 5 else 0.0,
                    'closing_debit': safe_float(row[6]) if len(row) > 6 else 0.0,
                    'closing_credit': safe_float(row[7]) if len(row) > 7 else 0.0,
                })
                continue
        wb.close()
    except Exception:
        rows = []

    if not rows:
        empty_df = pd.DataFrame(columns=base_cols + ['opening_balance_signed','turnover_net_signed','closing_balance_signed','net_change_signed'])
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as w:
            empty_df[base_cols].to_excel(w, sheet_name='01_TB_CLEAN', index=False)
            empty_df.to_excel(w, sheet_name='02_ACCOUNT_SUMMARY', index=False)
        buf.seek(0)
        return buf, empty_df

    df = pd.DataFrame(rows, columns=base_cols)
    for c in base_cols[2:]:
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0.0)

    df['opening_balance_signed'] = df['opening_debit'] - df['opening_credit']
    df['turnover_net_signed'] = df['turnover_debit'] - df['turnover_credit']
    df['closing_balance_signed'] = df['closing_debit'] - df['closing_credit']
    df['net_change_signed'] = df['closing_balance_signed'] - df['opening_balance_signed']
    tb_sum = df[['account_code','account_name','opening_debit','opening_credit','opening_balance_signed',
                  'turnover_debit','turnover_credit','turnover_net_signed',
                  'closing_debit','closing_credit','closing_balance_signed','net_change_signed']].copy()
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        df[base_cols].to_excel(w, sheet_name='01_TB_CLEAN', index=False)
        tb_sum.to_excel(w, sheet_name='02_ACCOUNT_SUMMARY', index=False)
    buf.seek(0)
    return buf, tb_sum

COL_PATTERNS = {
    'account_code': ['дансны код','данс код','account code','account no','account number','acc code','код'],
    'account_name': ['дансны нэр','данс нэр','account name','acc name','нэр'],
    'transaction_date': ['огноо','date','transaction date','txn date'],
    'debit_mnt': ['дебит','debit','dt','дт','debit amount'],
    'credit_mnt': ['кредит','credit','ct','кт','credit amount'],
    'balance_mnt': ['үлдэгдэл','balance','bal','ending balance'],
    'counterparty_name': ['харилцагч','counterparty','partner','vendor','customer'],
    'transaction_description': ['тайлбар','гүйлгээний утга','утга','description','memo','narration'],
    'journal_no': ['журнал','journal','journal no'],
    'document_no': ['баримт','document','doc no'],
}
def _match_col(h, field):
    h2 = str(h).lower().strip()
    return any(p in h2 for p in COL_PATTERNS.get(field, []))
def _auto_map(headers):
    m, used = {}, set()
    for f in ['account_code','debit_mnt','credit_mnt','transaction_date','account_name','counterparty_name','transaction_description','balance_mnt','journal_no','document_no']:
        for i, h in enumerate(headers):
            if i in used: continue
            if _match_col(h, f): m[f]=i; used.add(i); break
    return m

def _find_header_row(all_rows, max_scan=12):
    """Гарчигын мөрийг автоматаар хайна."""
    best_i, best_s = 0, 0
    for i, row in enumerate(all_rows[:max_scan]):
        vals = [str(c).strip().lower() for c in row if c is not None]
        score = 0
        for v in vals:
            if 'огноо' in v or 'date' in v: score += 1
            if 'дебет' in v or 'debit' in v: score += 1
            if 'кредит' in v or 'credit' in v: score += 1
            if 'мөнгөн дүн' in v or 'amount' in v: score += 1
            if 'гүйлгээний утга' in v or 'description' in v: score += 1
            if 'код' in v or 'code' in v or 'account' in v: score += 1
            if 'баримт' in v or 'document' in v: score += 1
            if 'журнал' in v or 'journal' in v: score += 1
        if score > best_s:
            best_s, best_i = score, i
    return best_i, best_s

def process_edt(file_obj, report_year):
    """Янз бүрийн ЕЖ / ерөнхий журнал форматыг бүх sheet-ээр шалгаж уншина.
    3 parser: standard (Данс:[...]), dual-entry (дебет/кредит данс), rowwise (мөр бүрд данс+дүн)
    """
    import openpyxl
    EDT_COLUMNS = ['report_year','account_code','account_name','transaction_no','transaction_date',
                   'journal_no','document_no','counterparty_name','counterparty_id',
                   'transaction_description','debit_mnt','credit_mnt','balance_mnt','month']

    def _to_date(v):
        if v is None: return ''
        if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
        s = str(v).strip()
        for fmt in ('%Y-%m-%d','%Y/%m/%d','%Y.%m.%d','%d.%m.%Y','%y.%m.%d','%y-%m-%d'):
            try: return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
            except: pass
        m = re.match(r'^(\d{2})[./-](\d{2})[./-](\d{2})$', s)
        if m:
            yy, mm, dd = m.groups()
            return f'20{yy}-{mm}-{dd}'
        return s[:10]

    def _pick(headers, candidates):
        headers_l = [str(h).strip().lower() if h is not None else '' for h in headers]
        for cand in candidates:
            for i, h in enumerate(headers_l):
                if cand in h: return i
        return None

    # ═══ Parser 1: Стандарт ЕЖ (Данс: [...]) ═══
    def _parse_standard_sheet(ws):
        rows_out, cur_code, cur_name = [], None, None
        for row in ws.iter_rows(values_only=True):
            c0 = row[0] if len(row) > 0 else None
            if c0 is None: continue
            s = str(c0).strip()
            if s.startswith('Данс:'):
                code, name = parse_account(s)
                if code: cur_code, cur_name = code, name
                continue
            if any(s.startswith(x) for x in ['Компани:','ЕРӨНХИЙ','Тайлант','Үүсгэсэн','Журнал:','№','Эцсийн','Дт -','Нийт','Эхний','Нээгээд']) or s in ('Валютаар','Төгрөгөөр',''):
                continue
            try: tx_no = int(float(c0))
            except: continue
            if cur_code is None: continue
            tx_date = _to_date(row[1] if len(row) > 1 else '')
            rows_out.append({'report_year':str(report_year),'account_code':cur_code,'account_name':cur_name,
                'transaction_no':str(tx_no),'transaction_date':tx_date,
                'journal_no':str(row[5]).strip() if len(row)>5 and row[5] else '',
                'document_no':str(row[6]).strip() if len(row)>6 and row[6] else '',
                'counterparty_name':str(row[3]).strip() if len(row)>3 and row[3] else '',
                'counterparty_id':str(row[4]).strip() if len(row)>4 and row[4] else '',
                'transaction_description':str(row[7]).strip() if len(row)>7 and row[7] else '',
                'debit_mnt':safe_float(row[9]) if len(row)>9 else 0.0,
                'credit_mnt':safe_float(row[11]) if len(row)>11 else 0.0,
                'balance_mnt':safe_float(row[13]) if len(row)>13 else 0.0,
                'month':tx_date[:7] if len(tx_date)>=7 else ''})
        return pd.DataFrame(rows_out, columns=EDT_COLUMNS), len(rows_out)

    # ═══ Parser 2: Dual-entry журнал (дебет данс + кредит данс + дүн) ═══
    def _parse_dual_entry_sheet(ws):
        all_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            all_rows.append(list(row))
            if i >= 1500: break
        if not all_rows: return pd.DataFrame(columns=EDT_COLUMNS), 0
        hdr_i, hdr_score = _find_header_row(all_rows)
        if hdr_score < 3: return pd.DataFrame(columns=EDT_COLUMNS), 0
        headers = [str(c).strip() if c is not None else f'col_{j}' for j, c in enumerate(all_rows[hdr_i])]
        debit_idx = _pick(headers, ['дебет', 'debit'])
        credit_idx = _pick(headers, ['кредит', 'credit'])
        amount_idx = _pick(headers, ['мөнгөн дүн', 'amount', 'дүн'])
        date_idx = _pick(headers, ['огноо', 'date'])
        doc_idx = _pick(headers, ['баримт №', 'баримт', 'document', 'doc'])
        cp_idx = _pick(headers, ['байгууллагын нэр', 'харилцагч', 'counterparty', 'customer', 'vendor'])
        desc_idx = _pick(headers, ['гүйлгээний утга', 'тайлбар', 'description', 'memo'])
        j_idx = _pick(headers, ['журналын төрөл', 'журнал', 'journal'])
        if debit_idx is None or credit_idx is None or amount_idx is None:
            return pd.DataFrame(columns=EDT_COLUMNS), 0
        rows_out = []
        tx_counter = 0
        for row in all_rows[hdr_i+1:]:
            if not row or all(c is None or str(c).strip()=='' for c in row): continue
            debit_acct = str(row[debit_idx]).strip() if debit_idx < len(row) and row[debit_idx] is not None else ''
            credit_acct = str(row[credit_idx]).strip() if credit_idx < len(row) and row[credit_idx] is not None else ''
            amount = safe_float(row[amount_idx]) if amount_idx < len(row) else 0.0
            if not re.search(r'\d', debit_acct or '') or not re.search(r'\d', credit_acct or '') or amount == 0: continue
            tx_date = _to_date(row[date_idx]) if date_idx is not None and date_idx < len(row) else ''
            doc_no = str(row[doc_idx]).strip() if doc_idx is not None and doc_idx < len(row) and row[doc_idx] is not None else ''
            cp_name = str(row[cp_idx]).strip() if cp_idx is not None and cp_idx < len(row) and row[cp_idx] is not None else ''
            desc = str(row[desc_idx]).strip() if desc_idx is not None and desc_idx < len(row) and row[desc_idx] is not None else ''
            journal_no = str(row[j_idx]).strip() if j_idx is not None and j_idx < len(row) and row[j_idx] is not None else ''
            tx_counter += 1
            common = {'report_year':str(report_year),'transaction_no':str(tx_counter),'transaction_date':tx_date,
                      'journal_no':journal_no,'document_no':doc_no,'counterparty_name':cp_name,'counterparty_id':'',
                      'transaction_description':desc,'balance_mnt':0.0,'month':tx_date[:7] if len(tx_date)>=7 else ''}
            rows_out.append({**common, 'account_code':debit_acct, 'account_name':'', 'debit_mnt':amount, 'credit_mnt':0.0})
            rows_out.append({**common, 'account_code':credit_acct, 'account_name':'', 'debit_mnt':0.0, 'credit_mnt':amount})
        return pd.DataFrame(rows_out, columns=EDT_COLUMNS), len(rows_out)

    # ═══ Parser 3: Rowwise журнал (мөр бүрд данс код + дебит + кредит) ═══
    def _parse_rowwise_sheet(ws):
        all_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            all_rows.append(list(row))
            if i >= 1500: break
        if not all_rows: return pd.DataFrame(columns=EDT_COLUMNS), 0
        hdr_i, hdr_score = _find_header_row(all_rows)
        if hdr_score < 3: return pd.DataFrame(columns=EDT_COLUMNS), 0
        headers = [str(c).strip() if c is not None else f'col_{j}' for j, c in enumerate(all_rows[hdr_i])]
        date_idx = _pick(headers, ['огноо', 'date'])
        doc_idx = _pick(headers, ['баримтын дугаар', 'баримт №', 'баримт', 'document'])
        code_idx = _pick(headers, ['код', 'account', 'данс'])
        name_idx = None
        if code_idx is not None and code_idx + 1 < len(headers):
            name_idx = code_idx + 1
        debit_amt_idx = _pick(headers, ['sum of дебет', 'debit amount', 'дебет', 'дт', 'debit'])
        credit_amt_idx = _pick(headers, ['sum of кредит', 'credit amount', 'кредит', 'кт', 'credit'])
        desc_idx = _pick(headers, ['гүйлгээний утга', 'тайлбар', 'description'])
        journal_idx = _pick(headers, ['журнал', 'journal'])
        cp_idx = _pick(headers, ['харилцагч', 'байгууллагын нэр', 'vendor', 'customer'])
        if code_idx is None or (debit_amt_idx is None and credit_amt_idx is None):
            return pd.DataFrame(columns=EDT_COLUMNS), 0
        rows_out = []
        for idx, row in enumerate(all_rows[hdr_i+1:], start=1):
            if not row or all(c is None or str(c).strip()=='' for c in row): continue
            acct = str(row[code_idx]).strip() if code_idx < len(row) and row[code_idx] is not None else ''
            if not re.search(r'\d', acct or ''): continue
            db = safe_float(row[debit_amt_idx]) if debit_amt_idx is not None and debit_amt_idx < len(row) else 0.0
            cr = safe_float(row[credit_amt_idx]) if credit_amt_idx is not None and credit_amt_idx < len(row) else 0.0
            if db == 0 and cr == 0: continue
            tx_date = _to_date(row[date_idx]) if date_idx is not None and date_idx < len(row) else ''
            doc_no = str(row[doc_idx]).strip() if doc_idx is not None and doc_idx < len(row) and row[doc_idx] is not None else ''
            acct_name = str(row[name_idx]).strip() if name_idx is not None and name_idx < len(row) and row[name_idx] is not None else ''
            cp_name = str(row[cp_idx]).strip() if cp_idx is not None and cp_idx < len(row) and row[cp_idx] is not None else ''
            desc = str(row[desc_idx]).strip() if desc_idx is not None and desc_idx < len(row) and row[desc_idx] is not None else ''
            journal_no = str(row[journal_idx]).strip() if journal_idx is not None and journal_idx < len(row) and row[journal_idx] is not None else ''
            rows_out.append({'report_year':str(report_year),'account_code':acct,'account_name':acct_name,'transaction_no':str(idx),
                'transaction_date':tx_date,'journal_no':journal_no,'document_no':doc_no,'counterparty_name':cp_name,
                'counterparty_id':'','transaction_description':desc,'debit_mnt':db,'credit_mnt':cr,'balance_mnt':0.0,
                'month':tx_date[:7] if len(tx_date)>=7 else ''})
        return pd.DataFrame(rows_out, columns=EDT_COLUMNS), len(rows_out)

    # ═══ Parser 4: Монголын ерөнхий журнал (Д/д | Огноо | Дугаар | Утга | Данс | Дебет | Кредит) ═══
    def _parse_mongolian_journal(ws):
        """Монголын стандарт ерөнхий журнал формат:
        Row ~8: Д/д | Баримтын | | Гүйлгээний утга | Харьцсан данс | Дүн |
        Row ~9: | Огноо | Дугаар | | | Дебет | Кредит
        Data: seq | date | doc | desc | account | debit | credit
        """
        all_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            all_rows.append(list(row))
            if i >= 3000: break
        if len(all_rows) < 10:
            return pd.DataFrame(columns=EDT_COLUMNS), 0

        # Гарчиг олох: "Д/д" эсвэл "Дебет" + "Кредит" агуулсан мөрийг хайна
        data_start = None
        for i in range(min(20, len(all_rows))):
            row_text = ' '.join(str(c).strip().lower() for c in all_rows[i] if c is not None)
            if 'д/д' in row_text or ('дебет' in row_text and 'кредит' in row_text):
                # Дараагийн мөр нь "Огноо" "Дугаар" агуулж магадгүй (2 мөрт гарчиг)
                if i + 1 < len(all_rows):
                    next_text = ' '.join(str(c).strip().lower() for c in all_rows[i+1] if c is not None)
                    if 'огноо' in next_text or 'дугаар' in next_text:
                        data_start = i + 2
                    else:
                        data_start = i + 1
                else:
                    data_start = i + 1
                break

        if data_start is None:
            return pd.DataFrame(columns=EDT_COLUMNS), 0

        # Дебет, кредит баганын индексийг олох
        # Гарчигын мөрүүдээс "Дебет", "Кредит" хайна
        debit_col, credit_col = None, None
        for check_row in range(max(0, data_start - 3), data_start):
            for j, cell in enumerate(all_rows[check_row]):
                if cell is None: continue
                cl = str(cell).strip().lower()
                if cl in ('дебет', 'дебит', 'debit', 'дүн') and debit_col is None:
                    debit_col = j
                elif cl in ('кредит', 'кредит', 'credit') and credit_col is None:
                    credit_col = j

        # Хэрэв олдохгүй бол 7 баганатай бол F=5, G=6 гэж таамаглана
        if debit_col is None or credit_col is None:
            max_cols = max(len(r) for r in all_rows[data_start:data_start+5]) if all_rows[data_start:data_start+5] else 0
            if max_cols >= 7:
                debit_col = debit_col or 5
                credit_col = credit_col or 6
            else:
                return pd.DataFrame(columns=EDT_COLUMNS), 0

        rows_out = []
        for row in all_rows[data_start:]:
            if not row or all(c is None or str(c).strip() == '' for c in row): continue
            # A=seq, B=date, C=doc, D=desc, E=account, F=debit, G=credit
            c0 = row[0] if len(row) > 0 else None
            if c0 is None: continue
            # Д/д нь тоо байх ёстой
            try:
                int(float(c0))
            except:
                # "Нийт", "Дүн" гэх мэт мөрийг алгасна
                s0 = str(c0).strip()
                if any(s0.startswith(x) for x in ['Нийт','Дүн','Журнал','Ерөнхий','Бүгд']):
                    continue
                continue

            tx_date = _to_date(row[1] if len(row) > 1 else '')
            doc_no = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
            desc = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
            acct = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
            db = safe_float(row[debit_col]) if debit_col < len(row) else 0.0
            cr = safe_float(row[credit_col]) if credit_col < len(row) else 0.0

            if db == 0 and cr == 0: continue
            if not acct or acct in ('None','nan',''): acct = '000'

            rows_out.append({
                'report_year': str(report_year), 'account_code': acct, 'account_name': '',
                'transaction_no': str(len(rows_out) + 1), 'transaction_date': tx_date,
                'journal_no': '', 'document_no': doc_no,
                'counterparty_name': '', 'counterparty_id': '',
                'transaction_description': desc,
                'debit_mnt': db, 'credit_mnt': cr, 'balance_mnt': 0.0,
                'month': tx_date[:7] if len(tx_date) >= 7 else ''
            })
        return pd.DataFrame(rows_out, columns=EDT_COLUMNS), len(rows_out)

    # ═══ Бүх sheet, бүх parser-ийг оролдоно ═══
    try:
        file_obj.seek(0)
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        best_df = pd.DataFrame(columns=EDT_COLUMNS)
        best_cnt = 0
        for sname in wb.sheetnames:
            ws = wb[sname]
            for parser in (_parse_mongolian_journal, _parse_standard_sheet, _parse_dual_entry_sheet, _parse_rowwise_sheet):
                try:
                    df_try, cnt_try = parser(ws)
                except Exception:
                    continue
                if cnt_try > best_cnt:
                    best_df, best_cnt = df_try, cnt_try
        wb.close()
        return best_df, best_cnt
    except Exception:
        return pd.DataFrame(columns=EDT_COLUMNS), 0

def generate_part1(df_led, year):
    df = df_led.copy()
    yr = str(year)
    df['debit_mnt'] = pd.to_numeric(df['debit_mnt'], errors='coerce').fillna(0)
    df['credit_mnt'] = pd.to_numeric(df['credit_mnt'], errors='coerce').fillna(0)
    df['balance_mnt'] = pd.to_numeric(df['balance_mnt'], errors='coerce').fillna(0)
    monthly = df.groupby(['month', 'account_code']).agg(
        total_debit_mnt=('debit_mnt', 'sum'),
        total_credit_mnt=('credit_mnt', 'sum'),
        ending_balance_mnt=('balance_mnt', 'last'),
        transaction_count=('debit_mnt', 'count')
    ).reset_index()
    monthly.insert(0, 'report_year', yr)
    anames = df.groupby('account_code')['account_name'].first()
    acct = df.groupby('account_code').agg(
        total_debit_mnt=('debit_mnt', 'sum'),
        total_credit_mnt=('credit_mnt', 'sum'),
        closing_balance_mnt=('balance_mnt', 'last')
    ).reset_index()
    acct['account_name'] = acct['account_code'].map(anames)
    acct.insert(0, 'report_year', yr)
    rm = df.groupby(['month', 'account_code', 'counterparty_name']).agg(
        transaction_count=('debit_mnt', 'count'),
        total_debit=('debit_mnt', 'sum'),
        total_credit=('credit_mnt', 'sum'),
    ).reset_index()
    rm['total_amount_mnt'] = rm['total_debit'].abs() + rm['total_credit'].abs()
    rm.insert(0, 'report_year', yr)
    p75a = rm['total_amount_mnt'].quantile(0.75)
    p75c = rm['transaction_count'].quantile(0.75)
    rm['risk_flag_large_txn'] = (rm['total_amount_mnt'] > p75a).astype(int)
    rm['risk_flag_high_frequency'] = (rm['transaction_count'] > p75c).astype(int)
    rm['risk_score'] = rm['risk_flag_large_txn'] + rm['risk_flag_high_frequency']
    rm['risk_level'] = pd.cut(
        rm['risk_score'],
        bins=[-0.1, 0.5, 1.5, 99],
        labels=['Бага', 'Дунд', 'Өндөр']
    ).astype(str)
    rm['account_category'] = rm['account_code'].str[:1].map(
        {'1': 'Хөрөнгө', '2': 'Өр', '3': 'Эздийн өмч', '4': 'Зардал', '5': 'Орлого', '6': 'Орлого', '7': 'Зардал'}
    ).fillna('')
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        monthly.to_excel(w, sheet_name='02_MONTHLY_SUMMARY', index=False)
        acct.to_excel(w, sheet_name='03_ACCOUNT_SUMMARY', index=False)
        rm.to_excel(w, sheet_name='04_RISK_MATRIX', index=False)
    buf.seek(0)
    n_risk = len(rm[rm['risk_score'] > 0])
    return buf, monthly, acct, rm, n_risk

def read_ledger(f):
    raw = f.read()
    f.seek(0)
    if raw[:2] == b'\x1f\x8b':
        return pd.read_csv(io.StringIO(gzip.decompress(raw).decode('utf-8')), dtype={'account_code': str})
    return pd.read_csv(io.BytesIO(raw), dtype={'account_code': str})

def get_year(name):
    for y in range(2020, 2030):
        if str(y) in name:
            return y
    return 2025

def load_tb(files):
    frames = []
    stats = {}
    must_cols = ['account_code','account_name','opening_debit','opening_credit','opening_balance_signed',
                 'turnover_debit','turnover_credit','turnover_net_signed',
                 'closing_debit','closing_credit','closing_balance_signed','net_change_signed']
    for f in files:
        year = get_year(f.name)
        bid, blabel = detect_branch(getattr(f, 'name', ''))
        try:
            df = pd.read_excel(f, sheet_name='02_ACCOUNT_SUMMARY')
        except Exception:
            try:
                f.seek(0)
                df = pd.read_excel(f)
            except Exception:
                df = pd.DataFrame()
        if df.empty:
            continue
        for c in must_cols:
            if c not in df.columns:
                if c in ['account_code','account_name']:
                    df[c] = ''
                else:
                    df[c] = 0.0
        df['year'] = year
        df['branch_id'] = bid
        df['branch_label'] = blabel
        for c in ['turnover_debit', 'turnover_credit', 'closing_debit', 'closing_credit', 'opening_debit', 'opening_credit', 'net_change_signed']:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
        stats[year] = {'accounts': len(df), 'turnover_d': df['turnover_debit'].sum(), 'turnover_c': df['turnover_credit'].sum(), 'branch': blabel}
        frames.append(df[must_cols + ['year', 'branch_id', 'branch_label']])
    if not frames:
        return pd.DataFrame(columns=must_cols + ['year', 'branch_id', 'branch_label']), {}
    return pd.concat(frames, ignore_index=True), stats


def load_ledger_stats(files, sample_per_year=20000, chunksize=100000):
    """Ledger файлуудыг chunk-ээр уншиж stats + sample DataFrame буцаана.
    Өндөр хэмжээтэй ledger дээр Streamlit Cloud OOM болохоос сэргийлнэ.
    """
    stats = {}
    sampled_frames = []
    needed_cols = [
        'report_year','account_code','account_name','transaction_no','transaction_date',
        'journal_no','document_no','counterparty_name','counterparty_id',
        'transaction_description','debit_mnt','credit_mnt','balance_mnt','month'
    ]

    def _iter_chunks(fobj):
        fobj.seek(0)
        raw = fobj.read()
        fobj.seek(0)
        if raw[:2] == b'\x1f\x8b':
            bio = io.BytesIO(gzip.decompress(raw))
            return pd.read_csv(bio, dtype={'account_code': str}, chunksize=chunksize)
        return pd.read_csv(io.BytesIO(raw), dtype={'account_code': str}, chunksize=chunksize)

    for f in files:
        year = get_year(f.name)
        total_rows = 0
        acct_set = set()
        month_set = set()
        monthly_parts = []
        year_samples = []

        try:
            for chunk in _iter_chunks(f):
                total_rows += len(chunk)

                for c in needed_cols:
                    if c not in chunk.columns:
                        chunk[c] = '' if c in (
                            'report_year','account_code','account_name','transaction_no','transaction_date',
                            'journal_no','document_no','counterparty_name','counterparty_id',
                            'transaction_description','month'
                        ) else 0

                chunk['account_code'] = chunk['account_code'].astype(str)
                chunk['debit_mnt'] = pd.to_numeric(chunk['debit_mnt'], errors='coerce').fillna(0)
                chunk['credit_mnt'] = pd.to_numeric(chunk['credit_mnt'], errors='coerce').fillna(0)
                chunk['report_year'] = str(year)

                acct_set.update(chunk['account_code'].dropna().astype(str).unique().tolist())
                month_set.update(chunk['month'].dropna().astype(str).unique().tolist())

                mo = chunk.groupby('month').agg(
                    rows=('debit_mnt', 'count'),
                    debit=('debit_mnt', 'sum'),
                    credit=('credit_mnt', 'sum')
                ).reset_index()
                monthly_parts.append(mo)

                # sample cap
                current_n = sum(len(x) for x in year_samples)
                remain = max(sample_per_year - current_n, 0)
                if remain > 0:
                    take_n = min(len(chunk), max(1000, remain))
                    year_samples.append(chunk.sample(n=min(take_n, len(chunk)), random_state=42)[needed_cols].copy())

        except Exception:
            # fallback жижиг файл дээр full read
            f.seek(0)
            df = read_ledger(f)
            total_rows = len(df)
            for c in needed_cols:
                if c not in df.columns:
                    df[c] = '' if c in (
                        'report_year','account_code','account_name','transaction_no','transaction_date',
                        'journal_no','document_no','counterparty_name','counterparty_id',
                        'transaction_description','month'
                    ) else 0
            df['account_code'] = df['account_code'].astype(str)
            df['debit_mnt'] = pd.to_numeric(df['debit_mnt'], errors='coerce').fillna(0)
            df['credit_mnt'] = pd.to_numeric(df['credit_mnt'], errors='coerce').fillna(0)
            df['report_year'] = str(year)
            acct_set.update(df['account_code'].dropna().astype(str).unique().tolist())
            month_set.update(df['month'].dropna().astype(str).unique().tolist())
            monthly_parts.append(df.groupby('month').agg(rows=('debit_mnt', 'count'), debit=('debit_mnt', 'sum'), credit=('credit_mnt', 'sum')).reset_index())
            year_samples.append(df.sample(n=min(sample_per_year, len(df)), random_state=42)[needed_cols].copy())

        mo = pd.concat(monthly_parts, ignore_index=True).groupby('month').agg(
            rows=('rows', 'sum'),
            debit=('debit', 'sum'),
            credit=('credit', 'sum')
        ).sort_index() if monthly_parts else pd.DataFrame(columns=['rows','debit','credit'])

        stats[year] = {
            'rows': int(total_rows),
            'accounts': int(len(acct_set)),
            'months': int(len(month_set)),
            'monthly': mo
        }

        if year_samples:
            year_sample = pd.concat(year_samples, ignore_index=True).head(sample_per_year)
            year_sample['report_year'] = str(year)
            sampled_frames.append(year_sample)

    full_df = pd.concat(sampled_frames, ignore_index=True) if sampled_frames else pd.DataFrame(columns=needed_cols)
    return stats, full_df

def load_part1(files):
    all_rm = []
    all_mo = []
    for f in files:
        year = get_year(f.name)
        try:
            rm = pd.read_excel(f, sheet_name='04_RISK_MATRIX')
            rm['year'] = year
            all_rm.append(rm)
        except Exception:
            pass
        try:
            mo = pd.read_excel(f, sheet_name='02_MONTHLY_SUMMARY')
            mo['year'] = year
            all_mo.append(mo)
        except Exception:
            pass
    rm_all = pd.concat(all_rm, ignore_index=True) if all_rm else pd.DataFrame()
    mo_all = pd.concat(all_mo, ignore_index=True) if all_mo else pd.DataFrame()
    return rm_all, mo_all



def _fast_read_journal_input(typ, f, year, max_rows=60000, chunksize=100000, progress_cb=None):
    """Ерөнхий журналын шинжилгээнд хурдан унших helper.
    - ledger/csv.gz бол chunk sample ашиглана
    - edt/xlsx бол parse хийсний дараа max_rows-аар sample авна
    """
    name = Path(getattr(f, 'name', f'journal_{year}')).name
    if typ == 'ledger':
        try:
            f.seek(0)
            stats_j, sample_df = load_ledger_stats([f], sample_per_year=max_rows, chunksize=chunksize)
            rows = 0
            accounts = 0
            if stats_j:
                meta = list(stats_j.values())[0]
                rows = int(meta.get('rows', 0))
                accounts = int(meta.get('accounts', 0))
            return sample_df.copy(), {'rows': rows, 'accounts': accounts, 'sample_rows': len(sample_df), 'source': 'chunk_sample'}
        except Exception:
            f.seek(0)
            led_df = read_ledger(f)
            if len(led_df) > max_rows:
                if progress_cb:
                    progress_cb('Төлөөлөх sample сонгож байна', '⏳', f'{len(led_df):,} мөрөөс {max_rows:,} мөр үлдээнэ')
                led_df = led_df.sample(max_rows, random_state=42)
            if progress_cb:
                progress_cb('Ledger бүрэн уншилт дууслаа', '✅', f'{len(led_df):,} мөр бэлэн боллоо')
            return led_df.copy(), {'rows': len(led_df), 'accounts': led_df.get('account_code', pd.Series(dtype=str)).astype(str).nunique(), 'sample_rows': len(led_df), 'source': 'full_read'}
    elif typ == 'edt':
        if progress_cb:
            progress_cb('ЕЖ parser ажиллуулж байна', '⏳', 'Excel sheet-үүдийг шалгаж journal мөрүүдийг стандарт багана руу буулгаж байна')
        f.seek(0)
        edt_df, cnt = process_edt(f, year)
        if not edt_df.empty and len(edt_df) > max_rows:
            if progress_cb:
                progress_cb('ЕЖ sample сонгож байна', '⏳', f'{len(edt_df):,} мөрөөс {max_rows:,} мөр үлдээнэ')
            edt_df = edt_df.sample(max_rows, random_state=42)
        if progress_cb:
            progress_cb('ЕЖ parser дууслаа', '✅', f'Нийт {int(cnt):,} мөрөөс {len(edt_df):,} мөр бэлэн боллоо')
        return edt_df.copy(), {'rows': int(cnt), 'accounts': edt_df.get('account_code', pd.Series(dtype=str)).astype(str).nunique() if not edt_df.empty else 0, 'sample_rows': len(edt_df), 'source': 'parsed_sample'}
    return pd.DataFrame(), {'rows': 0, 'accounts': 0, 'sample_rows': 0, 'source': 'none'}
def _render_reading_steps(step_rows, expanded=True):
    if not step_rows:
        return
    view = pd.DataFrame(step_rows)
    cols = [c for c in ['№','Файл','Төрөл','Он','Алхам','Төлөв','Нийт мөр','Sample мөр','Данс','Эх үүсвэр'] if c in view.columns]
    with st.expander("👁️ Уншиж байгаа процедур / алхмууд", expanded=expanded):
        st.dataframe(view[cols], use_container_width=True, hide_index=True)


def clean_for_risk(df):
    """NaN / хоосон утгыг цэвэрлэж, эрсдэлийн шинжилгээнд бэлтгэнэ."""
    d = df.copy()
    text_cols = ['account_code','account_name','counterparty_name','transaction_description','journal_no','document_no']
    for c in text_cols:
        if c in d.columns:
            d[c] = d[c].astype(str).replace(['nan','None','NaN','<NA>'],'').fillna('')
            d[c] = d[c].replace(r'^\s*$', '', regex=True)
    for c in ['debit_mnt','credit_mnt','balance_mnt']:
        if c in d.columns:
            d[c] = pd.to_numeric(d[c], errors='coerce').fillna(0.0)
    if 'account_code' in d.columns:
        d = d[d['account_code'].astype(str).str.strip() != '']
    return d.reset_index(drop=True)


def engineer_txn_features(d):
    """Гүйлгээ бүрээс шинж чанар үүсгэнэ. Дутуу багана байвал 0 утга ашиглана."""
    d = d.copy()
    # Баганууд байгаа эсэхийг шалгаж, дутууг нэмэх
    for c in ['debit_mnt','credit_mnt','account_code','account_name','counterparty_name','transaction_description','transaction_date']:
        if c not in d.columns:
            d[c] = '' if c in ('account_code','account_name','counterparty_name','transaction_description','transaction_date') else 0
    d['debit_mnt'] = pd.to_numeric(d['debit_mnt'], errors='coerce').fillna(0)
    d['credit_mnt'] = pd.to_numeric(d['credit_mnt'], errors='coerce').fillna(0)
    d['account_code'] = d['account_code'].astype(str).fillna('000')
    d['account_name'] = d['account_name'].astype(str).fillna('')
    d['counterparty_name'] = d['counterparty_name'].astype(str).fillna('')
    d['transaction_description'] = d['transaction_description'].astype(str).fillna('')
    d['transaction_date'] = d['transaction_date'].astype(str).fillna('')

    d['amount'] = d['debit_mnt'].abs() + d['credit_mnt'].abs()
    d['log_amount'] = np.log1p(d['amount'])
    d['is_debit'] = (d['debit_mnt'] > 0).astype(int)

    # Дансны ангилал
    try:
        le2 = LabelEncoder()
        d['acct_cat_num'] = le2.fit_transform(d['account_code'].str[:3])
    except:
        d['acct_cat_num'] = 0

    # Бенфорд
    digits = d['amount'].apply(lambda x: int(str(int(abs(x)))[0]) if abs(x) >= 1 else 0)
    d['benford_digit'] = digits
    benford_exp = {1:0.301,2:0.176,3:0.125,4:0.097,5:0.079,6:0.067,7:0.058,8:0.051,9:0.046}
    af = d[d['benford_digit']>0]['benford_digit'].value_counts(normalize=True)
    d['benford_dev'] = d['benford_digit'].map(lambda x: abs(af.get(x,0)-benford_exp.get(x,0)) if x>0 else 0)

    # Тэгш тоо
    d['is_round'] = (((d['amount']>=1e6)&(d['amount']%1e6==0)).astype(int) + ((d['amount']>=1e3)&(d['amount']%1e3==0)).astype(int))

    # Данс доторх z-score
    try:
        as2 = d.groupby('account_code')['amount'].agg(['mean','std']).fillna(0)
        as2.columns = ['acct_mean','acct_std']
        d = d.merge(as2, on='account_code', how='left')
        d['amt_zscore'] = np.where(d['acct_std']>0, (d['amount']-d['acct_mean'])/d['acct_std'], 0)
        d['amt_zscore'] = d['amt_zscore'].clip(-10,10).fillna(0)
    except:
        d['acct_mean'] = 0; d['acct_std'] = 0; d['amt_zscore'] = 0

    # Ховор харилцагч
    try:
        cp_f = d['counterparty_name'].value_counts()
        d['cp_rare'] = (d['counterparty_name'].map(cp_f).fillna(0) <= 3).astype(int)
    except:
        d['cp_rare'] = 0

    # Ховор данс-харилцагч хос
    try:
        d['pair'] = d['account_code'] + '|' + d['counterparty_name']
        pf = d['pair'].value_counts()
        d['pair_rare'] = (d['pair'].map(pf).fillna(0) <= 2).astype(int)
    except:
        d['pair_rare'] = 0

    # Тайлбар
    d['desc_empty'] = (d['transaction_description'].str.len() == 0).astype(int)

    # Давхардал
    try:
        d['dup_key'] = d['account_code'] + '|' + d['amount'].astype(str) + '|' + d['transaction_date']
        dk = d['dup_key'].value_counts()
        d['is_dup'] = (d['dup_key'].map(dk).fillna(1) > 1).astype(int)
    except:
        d['is_dup'] = 0

    # Цаг
    d['day'] = pd.to_numeric(d['transaction_date'].str[8:10], errors='coerce').fillna(15)
    d['month_num'] = pd.to_numeric(d['transaction_date'].str[5:7], errors='coerce').fillna(6)
    d['is_month_end'] = (d['day'] >= 28).astype(int)
    d['is_year_end'] = (d['month_num'] == 12).astype(int)

    # ═══ ТАЙЛБАР ↔ ДАНСНЫ НЭР ТУЛГАЛТ ═══
    d['desc_mismatch'] = 0
    d['name_no_overlap'] = 0
    d['dir_mismatch'] = 0

    # Том өгөгдөл дээр текст тулгалт маш удаан → 12,000+ мөр бол алгасна
    _DO_TEXT_MATCH = len(d) <= 12000
    if _DO_TEXT_MATCH:
        try:
            stop_w = {'дансны','данс','нийт','бусад','зардал','орлого','төлбөр','хөрөнгө','тооцоо','бүртгэл','дүн','төгрөг','сая','мянга','журнал','гүйлгээ','баримт'}
            # Данс бүрийн ердийн тайлбарын үгс
            acct_words = {}
            for code in d['account_code'].unique():
                all_desc = ' '.join(d.loc[d['account_code']==code, 'transaction_description'].str.lower())
                wc = Counter(re.findall(r'[а-яөүёА-ЯӨҮЁ\w]{3,}', all_desc))
                acct_words[code] = set(w for w,c in wc.items() if c >= 3 and len(w) >= 3)

            def _check_mismatch(code, tx_desc):
                tx = str(tx_desc).lower() if tx_desc else ''
                if not tx or code not in acct_words or not acct_words[code]: return 0
                tx_words = set(re.findall(r'[а-яөүёА-ЯӨҮЁ\w]{3,}', tx))
                return 0 if len(tx_words & acct_words[code]) > 0 else 1
            d['desc_mismatch'] = [_check_mismatch(c, t) for c, t in zip(d['account_code'], d['transaction_description'])]

            def _extract_kw(text):
                if not text: return set()
                return set(w for w in re.findall(r'[а-яөүёА-ЯӨҮЁ\w]{3,}', str(text).lower()) if w not in stop_w and len(w) >= 3)
            def _check_overlap(aname, tdesc):
                nk = _extract_kw(aname)
                dk2 = _extract_kw(tdesc)
                if not nk or not dk2: return 0
                return 0 if len(nk & dk2) > 0 else 1
            d['name_no_overlap'] = [_check_overlap(a, t) for a, t in zip(d['account_name'], d['transaction_description'])]
        except:
            pass

    # Дансны чиглэл зөрчил
    try:
        af2 = d['account_code'].str[0]
        d.loc[(af2=='1')&(d['credit_mnt']>0)&(d['debit_mnt']==0), 'dir_mismatch'] = 1
        d.loc[(af2=='2')&(d['debit_mnt']>0)&(d['credit_mnt']==0), 'dir_mismatch'] = 1
        d.loc[(af2=='5')&(d['debit_mnt']>0)&(d['credit_mnt']==0), 'dir_mismatch'] = 1
        d.loc[(af2.isin(['6','7','8']))&(d['credit_mnt']>0)&(d['debit_mnt']==0), 'dir_mismatch'] = 1
    except:
        pass

    return d

def run_txn_anomaly(df, cont=0.05):
    """Гүйлгээний аномали илрүүлэлт."""
    feats = ['log_amount','acct_cat_num','benford_dev','is_round','amt_zscore','cp_rare','pair_rare',
             'desc_empty','is_month_end','is_year_end','is_dup','is_debit','desc_mismatch','name_no_overlap','dir_mismatch']
    # Бүх feature багана байгаа эсэхийг шалгах
    for f in feats:
        if f not in df.columns:
            df[f] = 0
    X = df[feats].fillna(0).replace([np.inf,-np.inf], 0).astype(float)
    iso = IsolationForest(contamination=cont, random_state=42, n_estimators=200, n_jobs=1)
    df['txn_anomaly'] = (iso.fit_predict(X)==-1).astype(int)
    df['txn_score'] = -iso.score_samples(X)
    try:
        z = np.abs(StandardScaler().fit_transform(X))
        df['txn_zscore_flag'] = (z.max(axis=1)>2.5).astype(int)
    except:
        df['txn_zscore_flag'] = 0
    # Эрсдэлийн жинлэсэн оноо (ISA стандарттай нийцүүлсэн)
    df['txn_risk'] = (
        df['txn_anomaly'] * 3 +         # IF аномали (ISA 240)
        df['txn_zscore_flag'] * 2 +       # Z-score хэт хазайлт
        df['is_dup'] * 3 +               # Давхардсан гүйлгээ (ISA 240) — жин нэмсэн
        df['cp_rare'] * 1 +              # Ховор харилцагч (ISA 550)
        df['pair_rare'] * 1 +            # Ховор данс×харилцагч хос (ISA 550)
        (df['amt_zscore'].abs() > 3).astype(int) * 2 +  # Дундажаас хэт зөрсөн (ISA 520)
        df['desc_empty'] * 2 +           # Тайлбаргүй гүйлгээ (ISA 500) — жин нэмсэн
        df['desc_mismatch'] * 2 +        # Тайлбар↔данс зөрчил (ISA 500)
        df['name_no_overlap'] * 1 +      # Нэр давхцахгүй (ISA 500)
        df['dir_mismatch'] * 3 +         # Чиглэлийн зөрчил (ISA 240) — жин нэмсэн
        df.get('is_round', pd.Series(0, index=df.index)).astype(int) * 1  # Тэгш тоо
    )
    df['txn_risk_level'] = pd.cut(df['txn_risk'], bins=[-1,3,7,12,100],
        labels=['🟢 Бага','🟡 Дунд','🟠 Өндөр','🔴 Маш өндөр'])
    return df, feats


def run_txn_ml_ensemble(df, contamination=0.05, n_clusters=8):
    """Ерөнхий журналын ML ensemble: боломжит бүх хувилбарын аномали оноо."""
    d = clean_for_risk(df)
    if d is None or d.empty:
        return pd.DataFrame(), [], pd.DataFrame(), pd.DataFrame()
    d = engineer_txn_features(d)
    if len(d) < 5:
        d = d.copy()
        d['ml_iso_flag'] = 0
        d['ml_lof_flag'] = 0
        d['ml_svm_flag'] = 0
        d['ml_kmeans_flag'] = 0
        d['ml_zscore_flag'] = 0
        d['ml_vote_count'] = 0
        d['ml_anomaly_flag'] = 0
        d['ml_risk_level'] = '🟢 Бага'
        d['xai_top_feature'] = ''
        model_summary = pd.DataFrame([{'Алгоритм':'Too few rows','Илрүүлсэн аномали':0,'Хувь':0.0}])
        xai_importance = pd.DataFrame({'feature':[], 'importance':[]})
        return d, [], model_summary, xai_importance
    feat_cols = ['log_amount','acct_cat_num','benford_dev','is_round','amt_zscore','cp_rare','pair_rare',
                 'desc_empty','is_month_end','is_year_end','is_dup','is_debit','desc_mismatch','name_no_overlap','dir_mismatch']
    for c in feat_cols:
        if c not in d.columns:
            d[c] = 0
    X = d[feat_cols].fillna(0).replace([np.inf,-np.inf],0).astype(float)
    scaler = StandardScaler()
    Xs = scaler.fit_transform(X) if len(X) > 1 else X.values

    # Isolation Forest
    iso = IsolationForest(contamination=min(max(contamination,0.01),0.40), random_state=42, n_estimators=150)
    iso_pred = (iso.fit_predict(X) == -1).astype(int)
    iso_score = -iso.score_samples(X)

    # Percentile-based anomaly (LOF орлуулсан — хурдан)
    from scipy.stats import zscore as scipy_zscore
    pct_scores = np.abs(Xs).sum(axis=1) if len(Xs) > 0 else np.zeros(len(d))
    pct_cut = np.percentile(pct_scores, max(80, int((1-contamination)*100))) if len(pct_scores) > 5 else pct_scores.mean()
    lof_pred = (pct_scores >= pct_cut).astype(int)
    lof_score = pct_scores

    # Mahalanobis-like distance (SVM орлуулсан — хурдан)
    try:
        cov = np.cov(Xs.T)
        cov_inv = np.linalg.pinv(cov)
        mean = Xs.mean(axis=0)
        diff = Xs - mean
        maha = np.sqrt(np.sum(diff @ cov_inv * diff, axis=1))
        maha_cut = np.percentile(maha, max(80, int((1-contamination)*100)))
        svm_pred = (maha >= maha_cut).astype(int)
        svm_score = maha
    except:
        svm_pred = np.zeros(len(d), dtype=int)
        svm_score = np.zeros(len(d))

    # KMeans distance anomaly
    k = max(2, min(n_clusters, len(d)-1 if len(d) > 2 else 2))
    km = KMeans(n_clusters=k, random_state=42, n_init=10)
    km.fit(Xs)
    km_dist = km.transform(Xs).min(axis=1)
    km_cut = np.percentile(km_dist, max(80, int((1-contamination)*100))) if len(km_dist) > 5 else km_dist.mean()
    km_pred = (km_dist >= km_cut).astype(int)

    zmax = np.abs(Xs).max(axis=1) if len(d) > 0 else np.array([])
    z_pred = (zmax > 2.8).astype(int) if len(d) > 0 else np.array([])

    result = d.copy()
    result['ml_iso_flag'] = iso_pred
    result['ml_lof_flag'] = lof_pred
    result['ml_svm_flag'] = svm_pred
    result['ml_kmeans_flag'] = km_pred
    result['ml_zscore_flag'] = z_pred
    result['ml_iso_score'] = iso_score
    result['ml_lof_score'] = lof_score
    result['ml_svm_score'] = svm_score
    result['ml_kmeans_score'] = km_dist
    result['ml_vote_count'] = result[['ml_iso_flag','ml_lof_flag','ml_svm_flag','ml_kmeans_flag','ml_zscore_flag']].sum(axis=1)
    result['ml_anomaly_flag'] = (result['ml_vote_count'] >= 2).astype(int)
    result['ml_risk_level'] = pd.cut(result['ml_vote_count'], bins=[-0.1,1.5,2.5,5], labels=['🟢 Бага','🟠 Өндөр','🔴 Маш өндөр']).astype(str)

    xai_importance = pd.DataFrame({
        'feature': feat_cols,
        'importance': np.abs(np.corrcoef(np.column_stack([Xs, result['ml_anomaly_flag'].values]).T)[-1,:-1]) if len(d) > 3 else np.zeros(len(feat_cols))
    })
    xai_importance['importance'] = xai_importance['importance'].replace([np.inf,-np.inf],0).fillna(0)
    xai_importance = xai_importance.sort_values('importance', ascending=False)

    if shap is not None and len(d) > 20 and result['ml_anomaly_flag'].nunique() > 1:
        try:
            y_ml = result['ml_anomaly_flag'].astype(int).values
            rf_xai = RandomForestClassifier(n_estimators=120, random_state=42, class_weight='balanced')
            rf_xai.fit(X, y_ml)
            explainer = shap.TreeExplainer(rf_xai)
            shap_vals = explainer.shap_values(X)
            if isinstance(shap_vals, list):
                sv = shap_vals[-1]
            else:
                sv = shap_vals
            shap_abs = np.abs(sv).mean(axis=0)
            xai_importance = pd.DataFrame({'feature': feat_cols, 'importance': shap_abs}).sort_values('importance', ascending=False)
            result['xai_top_feature'] = np.array(feat_cols)[np.argmax(np.abs(sv), axis=1)]
        except Exception:
            result['xai_top_feature'] = xai_importance.iloc[0]['feature'] if not xai_importance.empty else ''
    else:
        if not xai_importance.empty:
            topf = xai_importance.iloc[0]['feature']
            result['xai_top_feature'] = np.where(result[topf].fillna(0) != 0, topf, '')
        else:
            result['xai_top_feature'] = ''

    model_summary = pd.DataFrame([
        {'Алгоритм':'Isolation Forest','Илрүүлсэн аномали':int(result['ml_iso_flag'].sum()),'Хувь':round(float(result['ml_iso_flag'].mean()*100),2)},
        {'Алгоритм':'Percentile-based','Илрүүлсэн аномали':int(result['ml_lof_flag'].sum()),'Хувь':round(float(result['ml_lof_flag'].mean()*100),2)},
        {'Алгоритм':'Mahalanobis distance','Илрүүлсэн аномали':int(result['ml_svm_flag'].sum()),'Хувь':round(float(result['ml_svm_flag'].mean()*100),2)},
        {'Алгоритм':'KMeans distance','Илрүүлсэн аномали':int(result['ml_kmeans_flag'].sum()),'Хувь':round(float(result['ml_kmeans_flag'].mean()*100),2)},
        {'Алгоритм':'Z-score','Илрүүлсэн аномали':int(result['ml_zscore_flag'].sum()),'Хувь':round(float(result['ml_zscore_flag'].mean()*100),2)},
        {'Алгоритм':'Ensemble ≥2 votes','Илрүүлсэн аномали':int(result['ml_anomaly_flag'].sum()),'Хувь':round(float(result['ml_anomaly_flag'].mean()*100),2)},
    ])
    return result, feat_cols, model_summary, xai_importance


def render_xai_summary(xai_importance, top_n=10):
    if xai_importance is None or len(xai_importance) == 0:
        st.info('XAI тайлбар гаргах хангалттай өгөгдөл алга.')
        return
    st.markdown('#### 🔎 XAI — Эрсдэлийг хамгийн их тайлбарлаж буй шинжүүд')
    show = xai_importance.head(top_n).copy()
    show['importance'] = pd.to_numeric(show['importance'], errors='coerce').fillna(0)
    st.dataframe(show, use_container_width=True, hide_index=True)
    fig_xai = px.bar(show.sort_values('importance', ascending=True), x='importance', y='feature', orientation='h', title='XAI нөлөөллийн эрэмбэ')
    st.plotly_chart(fig_xai, use_container_width=True)

def run_ml(tb_all, cont, n_est):
    df = tb_all.copy()
    if df.empty or len(df) < 5:
        return pd.DataFrame(), np.array([]), np.array([]), [], {}, '', pd.DataFrame(), np.array([])
    needed = ['account_code','turnover_debit','turnover_credit','closing_debit','closing_credit','opening_debit','net_change_signed','year']
    for c in needed:
        if c not in df.columns:
            df[c] = 0 if c != 'account_code' else ''
    df['cat_code'] = df['account_code'].astype(str).str[:3]
    le = LabelEncoder()
    df['cat_num'] = le.fit_transform(df['cat_code'].fillna(''))
    df['log_turn_d'] = np.log1p(pd.to_numeric(df['turnover_debit'], errors='coerce').fillna(0).abs())
    df['log_turn_c'] = np.log1p(pd.to_numeric(df['turnover_credit'], errors='coerce').fillna(0).abs())
    df['log_close_d'] = np.log1p(pd.to_numeric(df['closing_debit'], errors='coerce').fillna(0).abs())
    df['log_close_c'] = np.log1p(pd.to_numeric(df['closing_credit'], errors='coerce').fillna(0).abs())
    df['turn_ratio'] = (pd.to_numeric(df['turnover_debit'], errors='coerce').fillna(0) / pd.to_numeric(df['turnover_credit'], errors='coerce').replace(0, np.nan)).fillna(0).replace([np.inf, -np.inf], 0)
    if 'net_change_signed' in df.columns:
        df['log_abs_change'] = np.log1p(pd.to_numeric(df['net_change_signed'], errors='coerce').fillna(0).abs())
    else:
        df['log_abs_change'] = np.log1p((pd.to_numeric(df['closing_debit'], errors='coerce').fillna(0) - pd.to_numeric(df['opening_debit'], errors='coerce').fillna(0)).abs())
    # ISA 520: Аналитик горим — өсөлтийн хурд нэмэх
    opening_bal = pd.to_numeric(df.get('opening_debit', 0), errors='coerce').fillna(0).abs() + pd.to_numeric(df.get('opening_credit', 0), errors='coerce').fillna(0).abs()
    closing_bal = pd.to_numeric(df.get('closing_debit', 0), errors='coerce').fillna(0).abs() + pd.to_numeric(df.get('closing_credit', 0), errors='coerce').fillna(0).abs()
    df['growth_rate'] = np.where(opening_bal > 0, (closing_bal - opening_bal) / opening_bal, 0)
    df['growth_rate'] = df['growth_rate'].clip(-10, 10).fillna(0)
    feats = ['cat_num', 'log_turn_d', 'log_turn_c', 'log_close_d', 'log_close_c', 'turn_ratio', 'log_abs_change', 'growth_rate', 'year']
    X = df[feats].fillna(0).replace([np.inf, -np.inf], 0)
    iso = IsolationForest(contamination=min(max(cont, 0.01), 0.4), random_state=42, n_estimators=200)
    df['iso_anomaly'] = (iso.fit_predict(X) == -1).astype(int)
    sc = StandardScaler()
    df['zscore_anomaly'] = (np.abs(sc.fit_transform(X)).max(axis=1) > 2.0).astype(int)
    p95 = df['turn_ratio'].quantile(0.95)
    df['turn_anomaly'] = ((df['turn_ratio'] > p95) | (df['turn_ratio'] < -p95)).astype(int)
    df['ensemble_anomaly'] = ((df['iso_anomaly'] == 1) | ((df['zscore_anomaly'] == 1) & (df['turn_anomaly'] == 1))).astype(int)
    y = df['ensemble_anomaly'].values
    if len(np.unique(y)) < 2 or len(df) < 10:
        fi = pd.DataFrame({'feature': feats, 'importance': [0.0]*len(feats)})
        res = {'Random Forest': {'pred': y, 'prob': np.zeros(len(y)), 'precision': 0.0, 'recall': 0.0, 'f1': 0.0, 'auc': 0.0}}
        return df, X, y, feats, res, 'Random Forest', fi, np.zeros(len(df), dtype=int)
    n_splits = min(5, int(np.bincount(y).min())) if np.bincount(y).min() > 1 else 2
    cv = StratifiedKFold(n_splits=max(2, n_splits), shuffle=True, random_state=42)
    models = {
        'Random Forest': RandomForestClassifier(n_estimators=n_est, max_depth=10, random_state=42, class_weight='balanced'),
        'Gradient Boosting': GradientBoostingClassifier(n_estimators=150, learning_rate=0.1, random_state=42),
        'Logistic Regression': LogisticRegression(max_iter=1000, random_state=42, class_weight='balanced'),
    }
    res = {}
    for nm, mdl in models.items():
        try:
            yp = cross_val_predict(mdl, X, y, cv=cv, method='predict')
            ypr = cross_val_predict(mdl, X, y, cv=cv, method='predict_proba')[:, 1]
            res[nm] = {'pred': yp, 'prob': ypr, 'precision': precision_score(y, yp, zero_division=0), 'recall': recall_score(y, yp, zero_division=0), 'f1': f1_score(y, yp, zero_division=0), 'auc': roc_auc_score(y, ypr)}
        except Exception:
            yp = np.zeros(len(y), dtype=int)
            ypr = np.zeros(len(y), dtype=float)
            res[nm] = {'pred': yp, 'prob': ypr, 'precision': 0.0, 'recall': 0.0, 'f1': 0.0, 'auc': 0.0}
    best = max(res, key=lambda k: res[k]['f1']) if res else ''
    rf = models['Random Forest']
    try:
        rf.fit(X, y)
        fi = pd.DataFrame({'feature': feats, 'importance': rf.feature_importances_}).sort_values('importance', ascending=False)
    except Exception:
        fi = pd.DataFrame({'feature': feats, 'importance': [0.0]*len(feats)})
    nt = len(df)
    ns = max(1, int(nt * 0.20))
    at = pd.to_numeric(df['turnover_debit'], errors='coerce').fillna(0).abs() + pd.to_numeric(df['turnover_credit'], errors='coerce').fillna(0).abs()
    wt = (at / at.sum()).fillna(1 / nt) if at.sum() != 0 else pd.Series(np.repeat(1/nt, nt))
    np.random.seed(42)
    ms = np.zeros(nt, dtype=int)
    ms[np.random.choice(nt, size=min(ns, nt), replace=False, p=wt.values)] = 1
    ym = (ms & y).astype(int)
    return df, X, y, feats, res, best, fi, ym

# ═══════════════════════════════════════
# 🏷️ ДАНСНЫ АНГИЛАЛ — ХАСАХ БҮЛГҮҮД
# ═══════════════════════════════════════
# 6 бүлэг: шимтгэл, хаалтын бичилт, идэвхгүй, тогтмол зардал, коммунал, үндсэн орлого

EXCL_RULES = {
    'nan_data': {
        'label': '⚠️ NaN / дутуу өгөгдөл',
        'help': 'account_code, account_name, description, counterparty дутуу мөрүүд',
        'default': True,
        'account_prefixes': [],
        'name_keywords': ['nan','none','unknown'],
        'desc_keywords': ['nan','none','unknown'],
    },
    'шимтгэл': {
        'label': '🏦 Шимтгэл, хураамж, татвар',
        'help': 'Банкны шимтгэл, ХХОАт, НДШ, НӨАт, хүү зэрэг давтамжтай, бага дүнтэй бичилтүүд',
        'default': True,
        'account_prefixes': ['7027','7028','7029','354','356','3541','3542','3543','3544'],
        'name_keywords': [
            'шимтгэл','хураамж','банкны шимтгэл','үйлчилгээний хураамж',
            'комисс','commission','fee','bank charge','service charge',
            'тэмдэгтийн хураамж','нийгмийн даатгал','ндш','ххоат','нөат','vat',
            'хүү','interest','алданги','торгууль','penalty',
        ],
        'desc_keywords': [
            'шимтгэл','хураамж','commission','fee','interest','хүү','алданги',
        ],
    },
    'хаалтын_бичилт': {
        'label': '📕 Хаалтын бичилт, залруулга',
        'help': 'Жилийн эцсийн хаалт, залруулга, буцаалт, сторно, нээлтийн бичилтүүд',
        'default': True,
        'account_prefixes': [],
        'name_keywords': [],
        'desc_keywords': [
            'хаалт','хаах','closing','close','year end','year-end',
            'жилийн эцсийн','хаалтын бичилт','тайлант үеийн хаалт',
            'залруулга','adjustment','adjusting','аудитын залруулга',
            'буцаалт','reversal','сторно','storno',
            'нээлтийн бичилт','opening entry','нээлт',
            'хуримтлагдсан элэгдэл','элэгдэл тооцох','depreciation',
        ],
    },
    'идэвхгүй': {
        'label': '⏸️ Идэвхгүй данс (эргэлтгүй)',
        'help': 'Тухайн жилд ямар ч эргэлтгүй (дебит=0, кредит=0) данснууд',
        'default': True,
        'account_prefixes': [],
        'name_keywords': [],
        'desc_keywords': [],
    },
    'тогтмол_зардал': {
        'label': '📋 Тогтмол зардал (цалин, түрээс г.м.)',
        'help': 'Цалин, НДШ, түрээс, даатгал, элэгдэл зэрэг сар бүр давтагддаг зардлууд',
        'default': False,
        'account_prefixes': ['701','702','703','704','706','710','711','712','713','714','7011','7012','7013','7014','7021','7022','7023'],
        'name_keywords': [
            'цалин','хөдөлмөрийн хөлс','salary','wage','цалингийн',
            'түрээс','rent','lease','түрээсийн',
            'даатгал','insurance','даатгалын',
            'элэгдэл','depreciation','хорогдол','amortization',
            'нөөц','provision','нөөцийн',
            'тэтгэмж','тэтгэвэр','pension',
            'урамшуулал','bonus',
        ],
        'desc_keywords': [
            'цалин','salary','түрээс','rent','даатгал','insurance',
            'элэгдэл','depreciation','нөөц','provision',
        ],
    },
    'коммунал': {
        'label': '💡 Коммунал (тог, ус, дулаан, холбоо)',
        'help': 'Цахилгаан, ус, дулаан, интернет, утас, шуудан зэрэг коммунал зардлууд',
        'default': False,
        'account_prefixes': ['7024','7025','7026'],
        'name_keywords': [
            'цахилгаан','электр','electricity','power',
            'ус','усны','water',
            'дулаан','дулааны','heating','heat',
            'тог','тогны',
            'холбоо','холбооны','утас','утасны','telephone','phone','telecom',
            'интернет','internet','сүлжээ','network',
            'шуудан','шуудангийн','postal',
            'коммунал','utility','utilities',
        ],
        'desc_keywords': [
            'цахилгаан','electricity','ус','water','дулаан','heating',
            'тог','утас','phone','интернет','internet','коммунал','utility',
        ],
    },
    'үндсэн_орлого': {
        'label': '💰 Үндсэн үйл ажиллагааны орлого',
        'help': 'Борлуулалтын орлого, үйлчилгээний орлого — бизнесийн үндсэн урсгал',
        'default': False,
        'account_prefixes': ['511','512','521','522','531','532','601','602','611','612'],
        'name_keywords': [
            'борлуулалтын орлого','борлуулалт','sales revenue','revenue',
            'үйлчилгээний орлого','service revenue','service income',
            'үндсэн үйл ажиллагааны орлого','operating revenue',
            'бараа борлуулсны орлого','бүтээгдэхүүн борлуулалт',
            'ажил үйлчилгээний орлого',
        ],
        'desc_keywords': [
            'борлуулалт','sales','орлого','revenue','income',
        ],
    },
}

def classify_exclusions(df, level='account'):
    """Данс/гүйлгээг 6 хасах ангилалд хуваана.
    Returns: df with 'exclusion_tag' column
    Tags: 'шимтгэл','хаалтын_бичилт','идэвхгүй','тогтмол_зардал','коммунал','үндсэн_орлого','' (хасахгүй)
    """
    d = df.copy()
    d['exclusion_tag'] = ''
    code_str = d['account_code'].astype(str) if 'account_code' in d.columns else pd.Series('', index=d.index)
    name_lower = d['account_name'].astype(str).str.lower() if 'account_name' in d.columns else pd.Series('', index=d.index)

    if level == 'transaction':
        desc_lower = d['transaction_description'].astype(str).str.lower() if 'transaction_description' in d.columns else pd.Series('', index=d.index)
        combined = name_lower + ' ' + desc_lower
    else:
        combined = name_lower

    # ── NaN / дутуу өгөгдөл ──
    if level == 'transaction':
        has_missing = (code_str.str.strip() == '') | (combined.str.strip() == '')
        if 'counterparty_name' in d.columns:
            has_missing = has_missing | (d['counterparty_name'].astype(str).str.strip() == '')
        d.loc[has_missing, 'exclusion_tag'] = 'nan_data'
    else:
        has_missing = (code_str.str.strip() == '') | (name_lower.str.strip() == '')
        d.loc[has_missing, 'exclusion_tag'] = 'nan_data'

    # ── Дүрмүүдийг дарааллаар хэрэглэх (эхнийх нь давуу) ──
    for tag, rule in EXCL_RULES.items():
        if tag in ('идэвхгүй','nan_data'):
            continue  # Тусгай шалгалтууд
        untagged = d['exclusion_tag'] == ''
        # Дансны код prefix
        for prefix in rule.get('account_prefixes', []):
            mask = code_str.str.startswith(prefix) & untagged
            d.loc[mask, 'exclusion_tag'] = tag
        # Нэр/тайлбар keyword
        kws = rule.get('desc_keywords', []) if level == 'transaction' else rule.get('name_keywords', [])
        for kw in kws:
            mask = combined.str.contains(kw, na=False, regex=False) & (d['exclusion_tag'] == '')
            d.loc[mask, 'exclusion_tag'] = tag

    # ── Идэвхгүй данс (эргэлт = 0) — зөвхөн дансны түвшинд ──
    if level == 'account':
        for c in ['turnover_debit', 'turnover_credit']:
            if c not in d.columns: d[c] = 0
        turn_total = pd.to_numeric(d['turnover_debit'], errors='coerce').fillna(0).abs() + \
                     pd.to_numeric(d['turnover_credit'], errors='coerce').fillna(0).abs()
        d.loc[(turn_total == 0) & (d['exclusion_tag'] == ''), 'exclusion_tag'] = 'идэвхгүй'

    return d


# ═══════════════════════════════════════
# 🧠 УХААЛАГ ФАЙЛ ТАНИХ СИСТЕМ
# ═══════════════════════════════════════
def detect_file_type(f):
    """Файлын төрлийг автоматаар таних. Returns: (type, year)
    Types: 'raw_tb', 'edt', 'cash', 'expense_class', 'tb_std', 'ledger', 'part1', 'unknown'
    """
    name = f.name.lower()
    fname_orig = f.name
    year = get_year(f.name)

    # CSV/GZ → Ledger
    if name.endswith('.csv') or name.endswith('.gz') or name.endswith('.csv.gz'):
        return 'ledger', year

    # XLSX → need to check
    if not any(name.endswith(ext) for ext in ('.xlsx', '.xls', '.xlsm', '.xlsb')):
        return 'unknown', year

    # ── Файлын нэрээр хурдан таних ──
    name_check = fname_orig.lower().replace('_', ' ').replace('-', ' ')
    # ХАРИЛЦАХЫН ХУУЛГА / Банкны хуулга (ЕЖ-ээс ӨМНӨ шалгах!)
    cash_keywords = ['харилцахын хуулга', 'харилцах хуулга', 'харилцах үндсэн', 'харилцах нэмэлт',
                     'харилцахын', 'bank statement', 'cash journal', 'мөнгөн гүйлгээ',
                     'мж', 'төрийн сангийн харилцах']
    for kw in cash_keywords:
        if kw in name_check:
            return 'cash', year
    # Зардлын ангилал / Expense classification
    expense_keywords = ['зардлын ангилал', 'expense class', '190-р тушаал', '190 тушаал',
                        'эз ангилал', 'эдийн засгийн ангилал', 'pdf to excel', 'pdf_to_excel',
                        'зардал ангилал', 'сангийн сайд']
    for kw in expense_keywords:
        if kw in name_check:
            return 'expense_class', year
    # ЕЖ / Ерөнхий журнал / Journal
    edt_keywords = ['ерөнхий журнал', 'ерөнхий дэвтэр', 'едт', 'edt', 'general ledger', 'general journal',
                    'еренхий журнал', 'journal entry', 'journal entries']
    for kw in edt_keywords:
        if kw in name_check:
            return 'edt', year
    # ГҮЙЛГЭЭ_БАЛАНС / Trial Balance / Journal TB
    tb_keywords = ['гүйлгээ баланс', 'гүйлгээ_баланс', 'гуйлгээ баланс', 'trial balance',
                   'гүйлгэ баланс', 'гуйлгэ баланс']
    for kw in tb_keywords:
        if kw in name_check:
            return 'raw_tb', year
    # TB_standardized
    if 'tb_standardized' in name_check or 'tb standardized' in name_check:
        return 'tb_std', year
    # Part1
    if 'part1' in name_check or 'part 1' in name_check:
        return 'part1', year
    # Ledger
    if 'ledger' in name_check or 'prototype_ledger' in name_check:
        return 'ledger', year

    # ── Sheet бүтцээр таних ──
    import openpyxl
    try:
        raw = f.read()
        f.seek(0)
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        sheets = wb.sheetnames

        # TB_standardized: has '02_ACCOUNT_SUMMARY' sheet
        if '02_ACCOUNT_SUMMARY' in sheets:
            if '04_RISK_MATRIX' in sheets:
                wb.close()
                return 'part1', year
            wb.close()
            return 'tb_std', year

        # Part1: has '04_RISK_MATRIX' sheet
        if '04_RISK_MATRIX' in sheets:
            wb.close()
            return 'part1', year

        # ── Агуулгаар таних (200 мөр хүртэл шалгана) ──
        ws = wb[sheets[0]]
        sample_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            sample_rows.append(row)
            if i >= 200:
                break
        wb.close()

        # ХАРИЛЦАХ: "ЖУРНАЛ" header + 312xxx account in first data rows
        _has_journal_header = False
        _has_312_account = False
        _has_eronhij = False
        for row in sample_rows[:10]:
            for cell in row[:3]:
                if cell is not None:
                    s = str(cell).strip().upper()
                    if s == 'ЖУРНАЛ':
                        _has_journal_header = True
                    if 'ЕРӨНХИЙ' in s:
                        _has_eronhij = True
            # Check for 312xxx (Харилцах данс) in Үндсэн данс column
            if len(row) >= 4 and row[3] is not None:
                acct = str(row[3]).strip()
                if acct.startswith('312') or acct.startswith('311'):
                    _has_312_account = True

        # ХАРИЛЦАХ: has ЖУРНАЛ header + 312 account + no ЕРӨНХИЙ
        if _has_journal_header and _has_312_account and not _has_eronhij:
            return 'cash', year

        # ЕЖ: contains "Данс:" or "Компани:" or "ЕРӨНХИЙ" or "Журнал:" pattern
        if _has_journal_header and not _has_312_account:
            return 'edt', year

        for row in sample_rows:
            if row[0] is not None:
                s = str(row[0]).strip()
                if s.startswith('Данс:') or s.startswith('Компани:') or s.startswith('ЕРӨНХИЙ') or s.startswith('Журнал:'):
                    return 'edt', year
            for cell in row[:5]:
                if cell is not None and 'Данс:' in str(cell):
                    return 'edt', year

        # ГҮЙЛГЭЭ_БАЛАНС: has account codes like 101-XX-XX-XXX in column B
        for row in sample_rows:
            if len(row) >= 2 and row[1] is not None:
                code = str(row[1]).strip()
                if re.match(r'\d{3}-\d{2}-\d{2}-\d{3}', code):
                    return 'raw_tb', year

        # Fallback: check if it looks like a balance sheet
        for row in sample_rows:
            if row[0] is not None:
                try:
                    int(float(row[0]))
                    if len(row) >= 8 and row[1] is not None and re.match(r'\d{3}-', str(row[1])):
                        return 'raw_tb', year
                except:
                    pass

        # Зардлын ангилал: 'Code'+'Description' OR 8xxxxx codes
        _has_code_desc = False
        _has_8xx_codes = False
        for row in sample_rows[:5]:
            row_strs = [str(c).strip().lower() for c in row if c is not None]
            if 'code' in row_strs and 'description' in row_strs:
                _has_code_desc = True
            if any('дансны код' in s or 'зардлын нэр' in s or 'ангилал' in s for s in row_strs):
                _has_code_desc = True
        for row in sample_rows[1:20]:
            if row[0] is not None:
                c = str(row[0]).strip()
                if re.match(r'^[89]\d{5}$', c) or re.match(r'^[89]\d{4}$', c) or re.match(r'^2[012]\d{4}$', c):
                    _has_8xx_codes = True
                    break
        if _has_code_desc or _has_8xx_codes:
            return 'expense_class', year

        return 'unknown', year
    except Exception:
        f.seek(0)
        return 'unknown', year


def parse_account_names(file_obj):
    """Дансны код + нэрийн лавлах файл уншина.
    Формат: A баганад дансны код (1, 31, 312, 3121, 31213, ...), B баганад нэр.
    Санхүүгийн байдлын тайлан (СТ-1А) эсвэл дансны жагсаалт файл дэмжинэ.
    """
    import openpyxl
    try:
        file_obj.seek(0)
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        code_map = {}
        for row in ws.iter_rows(values_only=True):
            c0 = str(row[0]).strip() if row[0] is not None else ''
            c1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            if not c0 or not c1:
                continue
            # Зөвхөн тоон код авах (1, 31, 312, 3121, 31213, ...)
            c0_clean = re.sub(r'[^0-9]', '', c0)
            if c0_clean and len(c0_clean) >= 1:
                code_map[c0_clean] = c1.strip()
        wb.close()
        return code_map
    except Exception:
        return {}

def merge_account_names(df, code_map):
    """Гүйлгээний DataFrame-д дансны нэрийг prefix matching-аар нэгтгэнэ.
    Жишээ: 312130201 → 31213 → 'Арилжааны банк дахь харилцах'
    """
    if not code_map or 'account_code' not in df.columns:
        return df
    d = df.copy()

    def _find_name(code):
        code_str = re.sub(r'[^0-9]', '', str(code))
        # Урт prefix-ээс богино руу хайна (хамгийн нарийвчлалтай нэрийг олно)
        for length in range(len(code_str), 0, -1):
            prefix = code_str[:length]
            if prefix in code_map:
                return code_map[prefix]
        return ''

    # Хоосон эсвэл байхгүй нэртэй мөрүүдэд нэр нэмэх
    if 'account_name' not in d.columns:
        d['account_name'] = ''
    mask = d['account_name'].fillna('').str.strip() == ''
    if mask.any():
        d.loc[mask, 'account_name'] = d.loc[mask, 'account_code'].apply(_find_name)

    return d

def detect_account_names_file(file_obj):
    """Дансны нэрийн лавлах файл мөн эсэхийг шалгана."""
    import openpyxl
    try:
        file_obj.seek(0)
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        score = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 20: break
            row_text = ' '.join(str(c).strip().lower() for c in row if c is not None)
            if 'дансны код' in row_text: score += 3
            if 'балансын үзүүлэлт' in row_text or 'дансны нэр' in row_text: score += 3
            if 'санхүүгийн байдл' in row_text or 'ст-1' in row_text: score += 2
            if 'эхний үлдэгдэл' in row_text or 'эцсийн үлдэгдэл' in row_text: score += 1
        wb.close()
        file_obj.seek(0)
        return score >= 3
    except:
        file_obj.seek(0)
        return False

def materiality_base_from_tb(tb_df):
    if tb_df is None or tb_df.empty:
        return 0.0
    candidates = []
    for c in ['turnover_debit','turnover_credit','closing_debit','closing_credit','opening_debit','opening_credit']:
        if c in tb_df.columns:
            candidates.append(pd.to_numeric(tb_df[c], errors='coerce').fillna(0).abs().sum())
    return float(max(candidates)) if candidates else 0.0

def build_materiality_by_account(tb_df, overall_materiality, performance_ratio=0.75, trivial_ratio=0.05):
    """ISA 320 + ISA 330: Данс тус бүрийн материаллаг байдал + аудитын горим.
    Эрсдэлийн коэффициент (ISA 320.A12): Дансны өөрчлөлт, ангиллаас хамааран залруулна.
    Аудитын горим (ISA 330): Эрсдэлийн түвшнээс хамааран горим санал болгоно.
    """
    if tb_df is None or tb_df.empty:
        return pd.DataFrame()
    d = tb_df.copy()
    for c in ['account_code','account_name','closing_debit','closing_credit',
              'turnover_debit','turnover_credit','opening_debit','opening_credit']:
        if c not in d.columns:
            d[c] = '' if c in ['account_code','account_name'] else 0.0
    for c in ['closing_debit','closing_credit','turnover_debit','turnover_credit','opening_debit','opening_credit']:
        d[c] = pd.to_numeric(d[c], errors='coerce').fillna(0)

    # ── Суурь дүн тооцох ──
    d['closing_abs'] = d['closing_debit'].abs() + d['closing_credit'].abs()
    d['turnover_abs'] = d['turnover_debit'].abs() + d['turnover_credit'].abs()
    d['суурь_дүн'] = np.where(d['closing_abs'] > 0, d['closing_abs'], d['turnover_abs'])

    # ── Дансны ангилал (ISA 315) ──
    d['ангилал'] = d['account_code'].astype(str).str[0].map(
        {'1':'Хөрөнгө','2':'Өр төлбөр','3':'Эздийн өмч',
         '4':'Зардал','5':'Орлого','6':'Орлого',
         '7':'Үйл ажиллагааны зардал','8':'Бусад зардал','9':'Нэгдсэн данс'}
    ).fillna('Бусад')

    # ── Аналитик горим: Өөрчлөлтийн хувь (ISA 520) ──
    opening = d['opening_debit'].abs() + d['opening_credit'].abs()
    d['өөрчлөлт_%'] = np.where(opening > 0, (d['суурь_дүн'] - opening) / opening * 100, 0).round(1)

    # ── Эрсдэлийн коэффициент (ISA 320.A12) ──
    def _risk_coeff(row):
        pct = abs(row.get('өөрчлөлт_%', 0))
        cat = str(row.get('account_code', ''))[:1]
        if pct > 50: return 0.50   # Маш өндөр өөрчлөлт → 2× бага босго
        if pct > 30 and cat == '1': return 0.60  # Хөрөнгийн өндөр өөрчлөлт
        if cat in ('5','6','7','8'): return 0.75  # Орлого/зардал субъектив
        if pct > 20: return 0.75
        if pct < 5: return 1.20   # Бага эрсдэл → илүү өндөр босго
        return 1.00
    d['эрсдэлийн_коэфф'] = d.apply(_risk_coeff, axis=1)

    # ── Материаллаг байдлын хуваарилалт ──
    total_base = max(d['суурь_дүн'].sum(), 1)
    d['жин_%'] = (d['суурь_дүн'] / total_base * 100).round(3)
    d['зөвшөөрөгдөх_алдаа'] = (d['жин_%'] / 100 * overall_materiality * d['эрсдэлийн_коэфф']).round(0)
    d['гүйцэтгэлийн_мат'] = (d['зөвшөөрөгдөх_алдаа'] * performance_ratio).round(0)
    d['анхаарах_доод'] = (d['зөвшөөрөгдөх_алдаа'] * trivial_ratio).round(0)

    # ── Босго давсан эсэх (ISA 320.A12) ──
    d['босго_давсан'] = np.where(d['суурь_дүн'] > d['зөвшөөрөгдөх_алдаа'], '⚠️ Тийм', '✅ Үгүй')

    # ── Эрсдэлийн түвшин ──
    risk_score = d['жин_%'] * (2 - d['эрсдэлийн_коэфф'])
    d['эрсдэлийн_түвшин'] = pd.cut(risk_score, bins=[-0.001, 1.0, 5.0, 100.0],
        labels=['Бага', 'Дунд', 'Өндөр']).astype(str)

    # ── ISA 330 аудитын горимын санал ──
    def _audit_proc(row):
        if row.get('босго_давсан') == '⚠️ Тийм':
            return 'Нарийвчилсан шалгалт + Баталгаажуулалт (ISA 505)'
        lv = row.get('эрсдэлийн_түвшин', 'Бага')
        if lv == 'Өндөр': return 'Нарийвчилсан шалгалт (ISA 330.18)'
        if lv == 'Дунд': return 'Шинжилгээний процедур (ISA 520) + Хязгаарлагдмал шалгалт'
        return 'Шинжилгээний процедур (ISA 520)'
    d['аудитын_горим'] = d.apply(_audit_proc, axis=1)

    out = d[['account_code','account_name','ангилал','суурь_дүн','turnover_abs',
        'өөрчлөлт_%','жин_%','эрсдэлийн_коэфф',
        'зөвшөөрөгдөх_алдаа','гүйцэтгэлийн_мат','анхаарах_доод',
        'босго_давсан','эрсдэлийн_түвшин','аудитын_горим']].copy()
    out.columns = ['Дансны код','Дансны нэр','Ангилал','Эцсийн үлдэгдэл','Нийт эргэлт',
        'Өөрчлөлт %','Жин %','Эрсдэлийн коэфф',
        'Зөвшөөрөгдөх алдаа ₮','Гүйцэтгэлийн мат ₮','Анхаарах доод ₮',
        'Босго давсан','Эрсдэлийн түвшин','Аудитын горим (ISA 330)']
    return out.sort_values('Зөвшөөрөгдөх алдаа ₮', ascending=False).reset_index(drop=True)



FILE_TYPE_LABELS = {
    'raw_tb': ('📗 Гүйлгээний баланс', 'Гүйлгээний балансын түүхий файл → стандартчилсан хүснэгт болгон хөрвүүлнэ'),
    'edt': ('📘 Ерөнхий журнал (ЕЖ)', 'Ерөнхий журналын гүйлгээ → стандарт баганатай гүйлгээний файл болгон хөрвүүлнэ'),
    'tb_std': ('📊 Стандарт гүйлгээний баланс', 'Стандартчилсан гүйлгээний баланс → шинжилгээнд бэлэн'),
    'ledger': ('📄 Гүйлгээний дэлгэрэнгүй', 'Гүйлгээний дэлгэрэнгүй файл → журналын шинжилгээнд бэлэн'),
    'part1': ('📈 Сарын нэгтгэл', 'Сарын нэгтгэл + Эрсдэлийн матриц → дансны шинжилгээнд бэлэн'),
    'cash': ('🏦 Харилцахын хуулга (МЖ)', 'Харилцах дансны мөнгөн гүйлгээний журнал → МЖ шинжилгээнд ашиглана'),
    'expense_class': ('💰 Зардлын ангилал', 'Сангийн сайдын 190-р тушаалын зардлын ангилал → зардлын нийцлийн шалгалтад ашиглана'),
    'unknown': ('❓ Тодорхойгүй', 'Файлын төрлийг таних боломжгүй'),
}

for _k, _v in {
    'prepared_tb_cache': {},
    'prepared_part1_cache': {},
    'prepared_ledger_cache': {},
    'prep_detected_rows': [],
    'prep_process_rows': [],
    'journal_process_rows': [],
    'tb_error': '',
    'journal_error': '',
}.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v



def _push_process_row(scope, file_name, stage, status='⏳', detail='', year='', ftype='', rows=None, sample_rows=None):
    key = f"{scope}_process_rows"
    rows_list = list(st.session_state.get(key, []))
    payload = {
        'Файл': file_name,
        'Төрөл': ftype,
        'Он': year,
        'Алхам': stage,
        'Төлөв': status,
        'Дэлгэрэнгүй': detail,
        'Нийт мөр': rows if rows is not None else '',
        'Sample мөр': sample_rows if sample_rows is not None else '',
    }
    rows_list.append(payload)
    st.session_state[key] = rows_list
    return rows_list


def _render_process_log(scope, title='⏱️ Явцын дэлгэрэнгүй', expanded=True):
    rows_list = st.session_state.get(f"{scope}_process_rows", [])
    if not rows_list:
        return
    view = pd.DataFrame(rows_list)
    cols = [c for c in ['Файл','Төрөл','Он','Алхам','Төлөв','Дэлгэрэнгүй','Нийт мөр','Sample мөр'] if c in view.columns]
    with st.expander(title, expanded=expanded):
        st.dataframe(view[cols], use_container_width=True, hide_index=True)

def _cache_add(cache_key, filename, raw_bytes):
    st.session_state[cache_key][filename] = raw_bytes


def _cache_files(cache_key):
    out = []
    for name, raw in st.session_state.get(cache_key, {}).items():
        bio = io.BytesIO(raw)
        bio.name = name
        out.append(bio)
    return out


def _df_to_csv_bytes(df):
    return df.to_csv(index=False).encode('utf-8-sig')


def _df_to_excel_bytes(df_map):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        for sname, df in df_map.items():
            df.to_excel(w, sheet_name=sname[:31], index=False)
    return buf.getvalue()


def _prepare_from_uploaded(uploaded, acct_name_map=None, progress_cb=None):
    detected_rows = []
    acct_name_map = acct_name_map or {}
    st.session_state['prep_process_rows'] = []
    for f in uploaded or []:
        try:
            file_name = getattr(f, 'name', 'unknown_file')
            if progress_cb:
                progress_cb(file_name, 'Файлын төрлийг таньж байна', '⏳', 'Формат, бүтэц, оныг тодорхойлж байна')
            ftype, year = detect_file_type(f)
            f.seek(0)
            label, desc = FILE_TYPE_LABELS.get(ftype, FILE_TYPE_LABELS['unknown'])
            detected_rows.append({'Файл': file_name, 'Төрөл': label, 'Он': year, 'Тайлбар': desc})
            if progress_cb:
                progress_cb(file_name, 'Файлын төрөл тогтоогдлоо', '✅', desc, year=year, ftype=label)
            if ftype == 'raw_tb':
                if progress_cb:
                    progress_cb(file_name, 'TB стандартчилж байна', '⏳', 'Гүйлгээ балансын түүхий файлыг нэгэн жигд бүтэц рүү хөрвүүлж байна', year=year, ftype=label)
                buf, tb_sum = process_raw_tb(f)
                if tb_sum is not None and not tb_sum.empty:
                    _cache_add('prepared_tb_cache', f'TB_standardized_{year}.xlsx', buf.getvalue())
                    if progress_cb:
                        progress_cb(file_name, 'TB кэш хадгаллаа', '✅', f'{len(tb_sum):,} мөртэй стандарт TB бэлдлээ', year=year, ftype=label, rows=len(tb_sum), sample_rows=len(tb_sum))
            elif ftype == 'tb_std':
                _cache_add('prepared_tb_cache', file_name, f.getvalue())
                if progress_cb:
                    progress_cb(file_name, 'Стандарт TB хадгаллаа', '✅', 'Шинжилгээнд шууд ашиглах TB файл кэшлэгдлээ', year=year, ftype=label)
            elif ftype == 'part1':
                _cache_add('prepared_part1_cache', file_name, f.getvalue())
                if progress_cb:
                    progress_cb(file_name, 'Part1 хадгаллаа', '✅', 'Сарын нэгтгэл болон эрсдэлийн матрицын файл кэшлэгдлээ', year=year, ftype=label)
            elif ftype == 'ledger':
                _cache_add('prepared_ledger_cache', file_name, f.getvalue())
                if progress_cb:
                    progress_cb(file_name, 'Ledger кэш хадгаллаа', '✅', 'Journal шинжилгээнд ашиглах CSV/ledger файл хадгалагдлаа', year=year, ftype=label)
            elif ftype == 'edt':
                if progress_cb:
                    progress_cb(file_name, 'ЕЖ-г ledger рүү хөрвүүлж байна', '⏳', 'Sheet parser ажиллуулж стандарт баганатай CSV үүсгэнэ', year=year, ftype=label)
                f.seek(0)
                edt_df, cnt = process_edt(f, year)
                if acct_name_map and not edt_df.empty:
                    if progress_cb:
                        progress_cb(file_name, 'Дансны нэр тулгаж байна', '⏳', 'Лавлах файлаас дансны нэрийг нэгтгэж байна', year=year, ftype=label, rows=cnt, sample_rows=len(edt_df))
                    edt_df = merge_account_names(edt_df, acct_name_map)
                if cnt > 0 and not edt_df.empty:
                    csv_name = f'ledger_from_EJ_{year}_{Path(file_name).stem}.csv'
                    _cache_add('prepared_ledger_cache', csv_name, _df_to_csv_bytes(edt_df))
                    # g_ej-д нэгтгэх (Зардлын ангилал, Харилцагч, Эрсдэлийн нэгтгэл цэсүүдэд хэрэгтэй)
                    existing_ej = st.session_state.get('g_ej')
                    if existing_ej is not None and isinstance(existing_ej, pd.DataFrame) and not existing_ej.empty:
                        st.session_state['g_ej'] = pd.concat([existing_ej, edt_df], ignore_index=True)
                    else:
                        st.session_state['g_ej'] = edt_df
                    if progress_cb:
                        progress_cb(file_name, 'ЕЖ хөрвүүлэлт дууслаа', '✅', f'Ledger CSV бэлдлээ: {csv_name}', year=year, ftype=label, rows=cnt, sample_rows=len(edt_df))
                elif progress_cb:
                    progress_cb(file_name, 'ЕЖ-с өгөгдөл уншигдсангүй', '⚠️', 'Parser тохирох мөр олсонгүй', year=year, ftype=label)
            elif ftype == 'expense_class':
                if progress_cb:
                    progress_cb(file_name, 'Зардлын ангилал уншиж байна', '⏳', 'Сангийн сайдын тушаалын зардлын код уншиж байна', year=year, ftype=label)
                f.seek(0)
                try:
                    ec_df = pd.read_excel(f)
                except:
                    f.seek(0)
                    ec_df = pd.read_csv(f)
                # Баганы нэрийг стандартчилах
                col_map = {}
                for c in ec_df.columns:
                    cl = str(c).lower().strip()
                    if cl in ('code', 'код', 'дансны код', 'зардлын код'): col_map[c] = 'Code'
                    elif cl in ('description', 'тайлбар', 'зардлын нэр', 'ангилал', 'нэр'): col_map[c] = 'Description'
                if col_map:
                    ec_df.rename(columns=col_map, inplace=True)
                st.session_state['g_expense_class'] = ec_df
                if progress_cb:
                    progress_cb(file_name, 'Зардлын ангилал хадгаллаа', '✅', f'{len(ec_df):,} мөр зардлын код уншлаа', year=year, ftype=label, rows=len(ec_df))
            elif ftype == 'cash':
                if progress_cb:
                    progress_cb(file_name, 'Харилцахын хуулга уншиж байна', '⏳', 'МЖ файлыг уншиж байна', year=year, ftype=label)
                f.seek(0)
                try:
                    cash_df = pd.read_excel(f)
                except:
                    f.seek(0)
                    cash_df = pd.read_csv(f)
                cash_df['_file'] = file_name
                cash_df['_year'] = year
                # Existing cash data-тай нэгтгэх
                existing = st.session_state.get('g_cash')
                if existing is not None and isinstance(existing, pd.DataFrame) and not existing.empty:
                    st.session_state['g_cash'] = pd.concat([existing, cash_df], ignore_index=True)
                else:
                    st.session_state['g_cash'] = cash_df
                # Харилцахын хуулгыг g_ej-д мөн нэмэх (Харилцагч ISA 550 шинжилгээнд)
                _cash_as_ej = cash_df.copy()
                _cash_col_remap = {}
                for _cc in _cash_as_ej.columns:
                    _ccl = str(_cc).lower()
                    if 'огноо' in _ccl or 'date' in _ccl: _cash_col_remap[_cc] = 'transaction_date'
                    elif 'үндсэн данс' in _ccl: _cash_col_remap[_cc] = 'account_code'
                    elif 'харьцсан данс' in _ccl or 'харилцагч' in _ccl: _cash_col_remap[_cc] = 'counterparty_name'
                    elif 'дебет' in _ccl or 'дебит' in _ccl: _cash_col_remap[_cc] = 'debit_mnt'
                    elif 'кредит' in _ccl or 'кредет' in _ccl: _cash_col_remap[_cc] = 'credit_mnt'
                    elif 'утга' in _ccl or 'тайлбар' in _ccl: _cash_col_remap[_cc] = 'transaction_description'
                if _cash_col_remap:
                    _cash_as_ej.rename(columns=_cash_col_remap, inplace=True)
                if 'debit_mnt' in _cash_as_ej.columns and 'credit_mnt' in _cash_as_ej.columns:
                    _cash_as_ej['amount'] = _cash_as_ej['debit_mnt'].fillna(0) + _cash_as_ej['credit_mnt'].fillna(0)
                # account_code-оос код гаргах
                if 'account_code' in _cash_as_ej.columns:
                    _cash_as_ej['account_code'] = _cash_as_ej['account_code'].astype(str).str.extract(r'(\d+)')[0]
                if 'counterparty_name' in _cash_as_ej.columns:
                    _cash_as_ej['counterparty_name'] = _cash_as_ej['counterparty_name'].astype(str).str.replace(r'^\d+\s*-\s*', '', regex=True)
                _existing_ej2 = st.session_state.get('g_ej')
                if _existing_ej2 is not None and isinstance(_existing_ej2, pd.DataFrame) and not _existing_ej2.empty:
                    st.session_state['g_ej'] = pd.concat([_existing_ej2, _cash_as_ej], ignore_index=True)
                else:
                    st.session_state['g_ej'] = _cash_as_ej
                if progress_cb:
                    progress_cb(file_name, 'Харилцахын хуулга хадгаллаа', '✅', f'{len(cash_df):,} мөр МЖ-д нэмэгдлээ', year=year, ftype=label, rows=len(cash_df))
            else:
                if progress_cb:
                    progress_cb(file_name, 'Алгаслаа', '⚠️', 'Тухайн файлыг автоматаар бэлтгэх дүрэм олдсонгүй', year=year, ftype=label)
        except Exception as e:
            detected_rows.append({'Файл': getattr(f, 'name', 'unknown_file'), 'Төрөл': '❌ Алдаа', 'Он': '', 'Тайлбар': str(e)})
            if progress_cb:
                progress_cb(getattr(f, 'name', 'unknown_file'), 'Бэлтгэх үед алдаа гарлаа', '❌', str(e))
    st.session_state['prep_detected_rows'] = detected_rows


def _build_part1_from_prepared_ledgers(progress_cb=None):
    ledger_files = _cache_files('prepared_ledger_cache')
    created = 0
    for lf in ledger_files:
        try:
            year = get_year(lf.name)
            if progress_cb:
                progress_cb(lf.name, 'Part1 үүсгэж байна', '⏳', 'Ledger файлаас сарын нэгтгэл ба эрсдэлийн матриц боловсруулж байна', year=year, ftype='📄 Гүйлгээний дэлгэрэнгүй')
            lf.seek(0)
            led_df = read_ledger(lf)
            if led_df.empty:
                if progress_cb:
                    progress_cb(lf.name, 'Part1 үүсгээгүй', '⚠️', 'Ledger файл хоосон байсан тул алгаслаа', year=year, ftype='📄 Гүйлгээний дэлгэрэнгүй')
                continue
            part1_buf, _, _, _, _ = generate_part1(led_df, year)
            out_name = f'Part1_generated_{year}_{Path(lf.name).stem[:40]}.xlsx'
            _cache_add('prepared_part1_cache', out_name, part1_buf.getvalue())
            created += 1
            if progress_cb:
                progress_cb(lf.name, 'Part1 үүсгэлт дууслаа', '✅', f'{out_name} файлыг хадгаллаа', year=year, ftype='📄 Гүйлгээний дэлгэрэнгүй', rows=len(led_df), sample_rows=len(led_df))
        except Exception as e:
            if progress_cb:
                progress_cb(getattr(lf, 'name', 'ledger_file'), 'Сарын нэгтгэл үүсгэх үед алдаа гарлаа', '❌', str(e))
            continue
    return created


def _render_downloads(title, cache_key, mime):
    files = st.session_state.get(cache_key, {})
    if files:
        st.markdown(title)
        for name, raw in files.items():
            st.download_button(f'📥 {name}', raw, file_name=name, mime=mime, key=f'dl_{cache_key}_{name}')


def _render_branch_year_filter(df, key_prefix):
    """Салбар + Он шүүлтүүр. Үргэлж 3 багана (React tree тогтвортой байлгахын тулд)."""
    filtered = df.copy()
    has_branch = 'branch_label' in df.columns and df['branch_label'].nunique() > 1
    yr_col = None
    for yc in ['year', 'report_year']:
        if yc in df.columns and df[yc].nunique() > 1:
            yr_col = yc
            break

    # Үргэлж 3 багана (React component tree тогтвортой)
    fc1, fc2, fc3 = st.columns([1, 1, 2])
    sel_branch = 'Бүгд'
    sel_year = 'Бүгд'

    with fc1:
        if has_branch:
            branches = ['Бүгд'] + sorted(df['branch_label'].dropna().unique().tolist())
            sel_branch = st.selectbox("🏢 Салбар", branches, key=f'{key_prefix}_branch')

    with fc2:
        if yr_col:
            years = ['Бүгд'] + sorted(df[yr_col].dropna().unique().astype(str).tolist())
            sel_year = st.selectbox("📅 Он", years, key=f'{key_prefix}_year')

    if sel_branch != 'Бүгд' and 'branch_label' in filtered.columns:
        filtered = filtered[filtered['branch_label'] == sel_branch]
    if sel_year != 'Бүгд' and yr_col:
        filtered = filtered[filtered[yr_col].astype(str) == sel_year]

    with fc3:
        if has_branch or yr_col:
            st.caption(f"Шүүлтүүр: {len(filtered):,} / {len(df):,}")

    return filtered

def _show_dataframe_download(df, filename, label='📥 CSV татах'):
    if df is not None and not df.empty:
        st.download_button(label, _df_to_csv_bytes(df), file_name=filename, mime='text/csv', key=f'dl_{filename}')


if page.startswith("1"):
    st.header("1️⃣ Өгөгдөл оруулах, бэлтгэх")
    st.caption('⚡ Хурдан бэлтгэл: файлын бүтцийг автоматаар таньж, гүйлгээний баланс болон ерөнхий журналыг стандарт формат руу хөрвүүлнэ.')
    st.markdown("Файлаа нэг удаа оруулаад дараагийн цэсүүд дээр дахин ашиглаж болно.")

    uploaded = st.file_uploader("📎 Бүх файлуудаа энд оруулна уу", type=['xlsx','xls','xlsm','xlsb','csv','tsv','gz'], accept_multiple_files=True, key='smart_prep_main')
    acct_name_file = st.file_uploader("📋 Дансны нэрийн лавлах файл (заавал биш)", type=['xlsx','xls','xlsm','xlsb','csv'], key='acct_names_prep_main')
    acct_name_map = parse_account_names(acct_name_file) if acct_name_file else {}
    
    st.markdown("---")
    st.markdown("**🏛️ Төрийн аудитын нэмэлт файлууд** *(заавал биш)*")
    gov_expense_file = st.file_uploader("💰 Зардлын ангилал (Сангийн сайдын 190-р тушаал)", type=['xlsx','xls','xlsm','xlsb','csv'], key='gov_expense')
    gov_budget_file = st.file_uploader("📊 Төсвийн гүйцэтгэлийн тайлан", type=['xlsx','xls','xlsm','xlsb','csv'], key='gov_budget')
    gov_cash_files = st.file_uploader("🏛️ Мөнгөн гүйлгээний журнал (хэдэн ч он, хэдэн ч файл)", type=['xlsx','xls','xlsm','xlsb','csv','tsv','gz'], accept_multiple_files=True, key='gov_cash')
    
    if gov_expense_file:
        try:
            _ef_name = gov_expense_file.name.lower()
            if _ef_name.endswith('.csv') or _ef_name.endswith('.tsv'):
                _sep = '\t' if _ef_name.endswith('.tsv') else ','
                _ec_df = pd.read_csv(gov_expense_file, sep=_sep)
            else:
                _ec_df = pd.read_excel(gov_expense_file)
            # Баганы нэр стандартчилах
            _ec_col_map = {}
            for c in _ec_df.columns:
                cl = str(c).lower().strip()
                if cl in ('code', 'код', 'дансны код', 'зардлын код'): _ec_col_map[c] = 'Code'
                elif cl in ('description', 'тайлбар', 'зардлын нэр', 'ангилал', 'нэр'): _ec_col_map[c] = 'Description'
            if _ec_col_map:
                _ec_df.rename(columns=_ec_col_map, inplace=True)
            st.session_state['g_expense_class'] = _ec_df
            st.success(f"✅ Зардлын ангилал: {len(_ec_df):,} мөр")
        except Exception as _ee:
            st.error(f"❌ Зардлын ангилал уншихад алдаа: {_ee}")
    if gov_budget_file:
        st.session_state['g_budget'] = pd.read_excel(gov_budget_file)
        st.success(f"✅ Төсвийн тайлан: {len(st.session_state['g_budget'])} мөр")
    if gov_cash_files:
        _cash_dfs = []
        for _cf in gov_cash_files:
            try:
                _cfn = _cf.name.lower()
                if _cfn.endswith('.csv') or _cfn.endswith('.tsv'):
                    _sep = '\t' if _cfn.endswith('.tsv') else ','
                    _cash_dfs.append(pd.read_csv(_cf, sep=_sep))
                elif _cfn.endswith('.gz'):
                    _cash_dfs.append(pd.read_csv(_cf, compression='gzip'))
                else:
                    _cash_dfs.append(pd.read_excel(_cf))
                st.success(f"✅ {_cf.name}: {len(_cash_dfs[-1]):,} мөр")
            except Exception as _ce:
                st.error(f"❌ {_cf.name}: {_ce}")
        if _cash_dfs:
            st.session_state['g_cash'] = pd.concat(_cash_dfs, ignore_index=True)
            st.success(f"✅ Мөнгөн гүйлгээ нийт: {len(st.session_state['g_cash']):,} мөр")


    cpa, cpb = st.columns([1,1])
    with cpa:
        if st.button('🛠️ Файлуудыг таньж бэлтгэх', type='primary', use_container_width=True):
            if uploaded:
                st.session_state['prep_process_rows'] = []
                prep_status = st.empty()
                prep_progress = st.progress(0, text='Өгөгдөл бэлтгэх ажил эхэлж байна...')
                prep_log_box = st.empty()
                total_steps = max(len(uploaded) + max(len(uploaded), 1), 1)
                step_counter = {'n': 0}

                def _prep_progress_cb(file_name, stage, status='⏳', detail='', year='', ftype='', rows=None, sample_rows=None):
                    step_counter['n'] += 1
                    _push_process_row('prep', file_name, stage, status=status, detail=detail, year=year, ftype=ftype, rows=rows, sample_rows=sample_rows)
                    prep_status.info(f'{status} {file_name} — {stage}')
                    prep_progress.progress(min(step_counter['n'] / total_steps, 1.0), text=f'Өгөгдөл бэлтгэж байна... {file_name} / {stage}')
                    prep_log_box.empty()
                    with prep_log_box.container():
                        _render_process_log('prep', '👁️ Өгөгдөл бэлтгэх явцын дэлгэрэнгүй', expanded=True)

                _prepare_from_uploaded(uploaded, acct_name_map, progress_cb=_prep_progress_cb)
                created = _build_part1_from_prepared_ledgers(progress_cb=_prep_progress_cb)
                prep_progress.progress(1.0, text='Өгөгдөл бэлтгэх ажил дууслаа.')
                prep_status.success(f'✅ Файлууд бэлтгэгдлээ. ЕЖ-ээс {created} сарын нэгтгэл файл үүсгэлээ.')
                with prep_log_box.container():
                    _render_process_log('prep', '👁️ Өгөгдөл бэлтгэх явцын дэлгэрэнгүй', expanded=True)
                st.success(f'✅ Файлууд бэлтгэгдлээ. ЕЖ-ээс {created} сарын нэгтгэл файл үүсгэлээ.')
            else:
                st.warning('Файл оруулна уу.')
    with cpb:
        if st.button('🧹 Бүх хадгалсан өгөгдлийг цэвэрлэх', use_container_width=True):
            for key in ['prepared_tb_cache','prepared_part1_cache','prepared_ledger_cache','prep_detected_rows',
                        'tb_analysis_done','journal_ai_done','branch_done','tb_all','tb_stats','rm_all','mo_all','tb_filtered',
                        'tb_ml_df','tb_feature_importance','journal_ml_result','journal_ml_show',
                        'journal_model_summary','journal_xai','journal_ledger_stats','tb_upload_cache','journal_upload_cache',
                        'prep_process_rows','journal_process_rows',
                        'materiality_result','materiality_total','branch_engineered','branch_dfs',
                        'branch_comparison','branch_summary','branch_detect_info',
                        'ml7_results','ml7_unsup']:
                if key in st.session_state:
                    v = st.session_state[key]
                    if isinstance(v, dict): st.session_state[key] = {}
                    elif isinstance(v, list): st.session_state[key] = []
                    elif isinstance(v, bool): st.session_state[key] = False
                    elif isinstance(v, (int, float)): st.session_state[key] = 0
                    elif v is None: st.session_state[key] = None
                    elif isinstance(v, pd.DataFrame): st.session_state[key] = pd.DataFrame()
                    else: st.session_state[key] = ''
            st.success('Session цэвэрлэгдлээ.')

    if st.session_state.get('prep_detected_rows'):
        st.dataframe(pd.DataFrame(st.session_state['prep_detected_rows']), use_container_width=True, hide_index=True)
    _render_process_log('prep', '👁️ Өгөгдөл бэлтгэх явцын дэлгэрэнгүй', expanded=False)

    ca, cb, cc = st.columns(3)
    with ca:
        st.metric('📗 Гүйлгээний баланс', len(st.session_state.get('prepared_tb_cache', {})))
    with cb:
        st.metric('📘 Ерөнхий журнал', len(st.session_state.get('prepared_ledger_cache', {})))
    with cc:
        st.metric('📈 Сарын нэгтгэл', len(st.session_state.get('prepared_part1_cache', {})))

    _render_downloads('### 📦 Бэлэн гүйлгээний баланс', 'prepared_tb_cache', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    _render_downloads('### 📦 Бэлэн ерөнхий журнал', 'prepared_ledger_cache', 'text/csv')
    _render_downloads('### 📦 Бэлэн сарын нэгтгэл', 'prepared_part1_cache', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ═══════════════════════════════════════════════════════════
# 📊 ДАШБОАРД — Нэг дэлгэцэнд бүх гол мэдээлэл
# ═══════════════════════════════════════════════════════════
elif page.startswith("📊"):
    st.header("📊 Аудитын дашбоард")

    has_tb = st.session_state.get('tb_analysis_done', False)
    has_journal = st.session_state.get('journal_ai_done', False)

    # Бэлтгэсэн файл байгаа ч шинжилгээ хийгдээгүй бол товч харуулах
    _has_prepared_tb = bool(st.session_state.get('prepared_tb_cache', {}))
    _has_prepared_ej = bool(st.session_state.get('prepared_ledger_cache', {}))
    _needs_analysis = (not has_tb and _has_prepared_tb) or (not has_journal and _has_prepared_ej)

    if _needs_analysis:
        st.info("📊 Бэлтгэсэн файлууд байна. Шинжилгээг эхлүүлнэ үү.")
        if st.button("🚀 Дашбоардын шинжилгээг эхлүүлэх", type='primary', use_container_width=True, key='dash_auto_run'):
            if not has_tb and _has_prepared_tb:
                with st.spinner("📊 Гүйлгээний баланс шинжилж байна..."):
                    try:
                        _tb_files = _cache_files('prepared_tb_cache')
                        _p1_files = _cache_files('prepared_part1_cache')
                        _tb_all, _tb_stats = load_tb(_tb_files) if _tb_files else (pd.DataFrame(), {})
                        _rm_all, _mo_all = load_part1(_p1_files) if _p1_files else (pd.DataFrame(), pd.DataFrame())
                        if not _tb_all.empty:
                            _ml_df, _X, _y, _feats, _res, _best, _fi, _ym = run_ml(_tb_all, 0.10, 200)
                            st.session_state['tb_analysis_done'] = True
                            st.session_state['tb_all'] = _tb_all
                            st.session_state['tb_stats'] = _tb_stats
                            st.session_state['rm_all'] = _rm_all
                            st.session_state['mo_all'] = _mo_all
                            st.session_state['tb_filtered'] = _tb_all
                            st.session_state['tb_ml_df'] = _ml_df
                            st.session_state['tb_feature_importance'] = _fi
                            has_tb = True
                    except Exception as e:
                        st.warning(f"ГБ шинжилгээний алдаа: {e}")
            if not has_journal and _has_prepared_ej:
                with st.spinner("📝 Ерөнхий журнал шинжилж байна..."):
                    try:
                        _led_files = _cache_files('prepared_ledger_cache')
                        _frames = []
                        for _lf in _led_files:
                            _sample_df, _meta = _fast_read_journal_input('ledger', _lf, get_year(_lf.name), max_rows=50000)
                            if _sample_df is not None and not _sample_df.empty:
                                _bid, _blbl = detect_branch(_lf.name)
                                _sample_df['branch_id'] = _bid
                                _sample_df['branch_label'] = _blbl
                                _frames.append(_sample_df)
                        if _frames:
                            _led_sample = pd.concat(_frames, ignore_index=True)
                            _led_sample = clean_for_risk(_led_sample)
                            _ml_result_j, _ml_feats_j, _model_summary_j, _xai_j = run_txn_ml_ensemble(_led_sample, contamination=0.05, n_clusters=8)
                            if not _ml_result_j.empty:
                                st.session_state['journal_ai_done'] = True
                                st.session_state['journal_ml_result'] = _ml_result_j
                                st.session_state['journal_ml_show'] = _ml_result_j
                                st.session_state['journal_model_summary'] = _model_summary_j
                                st.session_state['journal_xai'] = _xai_j
                                has_journal = True
                    except Exception as e:
                        st.warning(f"ЕЖ шинжилгээний алдаа: {e}")
            if has_tb or has_journal:
                try: st.rerun()
                except: pass

    # ── Өгөгдөл бэлэн эсэхийг дахин шалгах ──
    tb_all = st.session_state.get('tb_all', pd.DataFrame())
    tb_stats = st.session_state.get('tb_stats', {})
    ml_df = st.session_state.get('tb_ml_df', pd.DataFrame())
    rm_all = st.session_state.get('rm_all', pd.DataFrame())
    mo_all = st.session_state.get('mo_all', pd.DataFrame())
    journal_result = st.session_state.get('journal_ml_result', pd.DataFrame())
    journal_stats = st.session_state.get('journal_ledger_stats', {})

    # Бэлтгэсэн файл ч байхгүй, шинжилгээ ч хийгдээгүй
    _has_any_prepared = bool(st.session_state.get('prepared_tb_cache', {})) or bool(st.session_state.get('prepared_ledger_cache', {}))

    if not has_tb and not has_journal and not _has_any_prepared:
        st.warning("⚠️ Дашбоард харахын тулд эхлээд өгөгдөл бэлтгэж, шинжилгээ хийнэ үү.")
        st.markdown("""
        <div style="background: linear-gradient(135deg, #e3f2fd 0%, #f5f5f5 100%);
             padding: 24px; border-radius: 12px; border-left: 5px solid #1976D2; margin: 20px 0;">
            <h4 style="color: #1565C0; margin-top: 0;">🚀 Эхлэх алхмууд:</h4>
            <ol style="color: #333; font-size: 14px; line-height: 2.0;">
                <li><b>1️⃣ Өгөгдөл оруулах</b> — Excel файлуудаа оруулж бэлтгэнэ</li>
                <li><b>📊 Дашбоард</b> — Энэ хуудас руу буцахад автоматаар шинжилгээ хийгдэнэ</li>
            </ol>
            <p style="color: #555; font-size: 13px; margin-top: 10px;">
                💡 <i>1️⃣ цэсэнд файл бэлтгэсний дараа энэ хуудас руу буцахад шинжилгээ автоматаар ажиллана.</i>
            </p>
        </div>
        """, unsafe_allow_html=True)
    elif not has_tb and not has_journal:
        st.info("⏳ Бэлтгэсэн файлуудаас шинжилгээ хийх боломжгүй байна. 2️⃣ эсвэл 3️⃣ цэсэнд шинжилгээг гараар ажиллуулна уу.")
    else:
        # ════════════════════════════════════════════════════
        # САЛБАР + ОН ШҮҮЛТҮҮР (Дашбоардын дээд хэсэг)
        # ════════════════════════════════════════════════════
        _dash_branch = 'Бүгд'
        _all_branches = set()
        if has_tb and not ml_df.empty and 'branch_label' in ml_df.columns:
            _all_branches.update(ml_df['branch_label'].dropna().unique())
        if has_journal and not journal_result.empty and 'branch_label' in journal_result.columns:
            _all_branches.update(journal_result['branch_label'].dropna().unique())

        # Үргэлж 3 багана (React tree тогтвортой)
        _dc1, _dc2, _dc3 = st.columns([1, 1, 2])
        with _dc1:
            if len(_all_branches) > 1:
                _dash_branch = st.selectbox("🏢 Салбар", ['Бүгд'] + sorted(_all_branches), key='dash_branch')

        # Шүүлтүүр хэрэглэх
        _ml_df_f = ml_df.copy() if has_tb and not ml_df.empty else pd.DataFrame()
        _journal_f = journal_result.copy() if has_journal and not journal_result.empty else pd.DataFrame()
        if _dash_branch != 'Бүгд':
            if not _ml_df_f.empty and 'branch_label' in _ml_df_f.columns:
                _ml_df_f = _ml_df_f[_ml_df_f['branch_label'] == _dash_branch]
            if not _journal_f.empty and 'branch_label' in _journal_f.columns:
                _journal_f = _journal_f[_journal_f['branch_label'] == _dash_branch]

        # ════════════════════════════════════════════════════
        # KPI METRIC CARDS
        # ════════════════════════════════════════════════════
        st.markdown("### 📈 Гол үзүүлэлтүүд")

        # TB-ийн мэдээлэл цуглуулах
        total_accounts = 0
        total_turnover_d = 0
        total_turnover_c = 0
        anomaly_count = 0
        anomaly_pct = 0.0

        if has_tb and not _ml_df_f.empty:
            total_accounts = _ml_df_f['account_code'].nunique() if 'account_code' in _ml_df_f.columns else len(_ml_df_f)
            for c in ['turnover_debit', 'turnover_credit']:
                if c in _ml_df_f.columns:
                    if c == 'turnover_debit':
                        total_turnover_d = pd.to_numeric(_ml_df_f[c], errors='coerce').fillna(0).abs().sum()
                    else:
                        total_turnover_c = pd.to_numeric(_ml_df_f[c], errors='coerce').fillna(0).abs().sum()
            if 'ensemble_anomaly' in _ml_df_f.columns:
                anomaly_count = int(_ml_df_f['ensemble_anomaly'].sum())
                anomaly_pct = float(_ml_df_f['ensemble_anomaly'].mean() * 100)

        # ЕЖ мэдээлэл
        journal_rows = 0
        journal_anomaly = 0
        journal_anomaly_pct = 0.0
        if has_journal and not _journal_f.empty:
            journal_rows = len(_journal_f)
            if 'ml_anomaly_flag' in _journal_f.columns:
                journal_anomaly = int(_journal_f['ml_anomaly_flag'].sum())
                journal_anomaly_pct = float(_journal_f['ml_anomaly_flag'].mean() * 100) if journal_rows > 0 else 0

        # ── Row 1: Үндсэн тоонууд ──
        k1, k2, k3, k4 = st.columns(4)
        with k1:
            st.metric("📋 Нийт данс", f"{total_accounts:,}", help="Шинжилгээнд хамрагдсан дансны тоо")
        with k2:
            total_turnover = total_turnover_d + total_turnover_c
            if total_turnover >= 1e9:
                st.metric("💰 Нийт эргэлт", f"₮{total_turnover/1e9:.1f} тэрбум", help="Дебит + Кредит нийт эргэлт")
            elif total_turnover >= 1e6:
                st.metric("💰 Нийт эргэлт", f"₮{total_turnover/1e6:.1f} сая", help="Дебит + Кредит нийт эргэлт")
            else:
                st.metric("💰 Нийт эргэлт", f"₮{total_turnover:,.0f}", help="Дебит + Кредит нийт эргэлт")
        with k3:
            delta_color = "inverse" if anomaly_pct > 10 else "off"
            st.metric("⚠️ Аномали данс", f"{anomaly_count:,}", delta=f"{anomaly_pct:.1f}%", delta_color=delta_color,
                      help="Ensemble аргаар илрүүлсэн хэвийн бус данс")
        with k4:
            st.metric("📝 ЕЖ гүйлгээ", f"{journal_rows:,}", delta=f"Аномали: {journal_anomaly:,}" if has_journal else "Хийгдээгүй",
                      delta_color="inverse" if journal_anomaly > 0 else "off",
                      help="Ерөнхий журналын гүйлгээний тоо")

        # ── Row 2: Эрсдэлийн тойм ──
        if has_journal and journal_rows > 0:
            k5, k6 = st.columns(2)
            with k5:
                st.metric("🔴 ЕЖ Аномали", f"{journal_anomaly:,}", delta=f"{journal_anomaly_pct:.1f}%",
                          delta_color="inverse", help="Машин сургалтын нэгдсэн загварын илрүүлсэн аномали гүйлгээ")
            with k6:
                normal_pct = 100 - journal_anomaly_pct
                st.metric("🟢 Хэвийн гүйлгээ", f"{journal_rows - journal_anomaly:,}",
                          delta=f"{normal_pct:.1f}%", help="Аномали бус гүйлгээний тоо")

        st.markdown("---")

        # ════════════════════════════════════════════════════
        # ГРАФИКУУД
        # ════════════════════════════════════════════════════
        if has_tb and not _ml_df_f.empty:
            dash_tab1, dash_tab2, dash_tab3, dash_tab4 = st.tabs([
                "📊 Дансны бүтэц", "🔍 Эрсдэлийн тойм", "📈 Чиг хандлага", "📋 Топ эрсдэлтэй данс"
            ])

            # ── TAB 1: Дансны бүтэц ──
            with dash_tab1:
                col_a, col_b = st.columns(2)

                with col_a:
                    # Дансны ангилалын pie chart
                    cat_col = _ml_df_f['account_code'].astype(str).str[:1].map(
                        {'1':'Хөрөнгө','2':'Өр төлбөр','3':'Эздийн өмч',
                         '4':'Зардал','5':'Орлого','6':'Орлого',
                         '7':'Үйл ажиллагааны зардал','8':'Бусад зардал','9':'Нэгдсэн данс'}
                    ).fillna('Бусад')
                    cat_counts = cat_col.value_counts().reset_index()
                    cat_counts.columns = ['Ангилал', 'Тоо']
                    fig_pie = px.pie(cat_counts, values='Тоо', names='Ангилал',
                                    title='Дансны ангилалын хуваарилалт',
                                    color_discrete_sequence=PLOTLY_COLORS,
                                    hole=0.4)
                    fig_pie.update_traces(textposition='inside', textinfo='percent+label')
                    fig_pie.update_layout(height=380, margin=dict(t=40, b=20, l=20, r=20))
                    st.plotly_chart(fig_pie, use_container_width=True)

                with col_b:
                    # Эргэлтийн ангилалаар bar chart
                    if 'turnover_debit' in _ml_df_f.columns and 'turnover_credit' in _ml_df_f.columns:
                        cat_turn = _ml_df_f.copy()
                        cat_turn['ангилал'] = cat_col
                        cat_turn['turnover_debit'] = pd.to_numeric(cat_turn['turnover_debit'], errors='coerce').fillna(0).abs()
                        cat_turn['turnover_credit'] = pd.to_numeric(cat_turn['turnover_credit'], errors='coerce').fillna(0).abs()
                        cat_agg = cat_turn.groupby('ангилал').agg(
                            Дебит=('turnover_debit', 'sum'),
                            Кредит=('turnover_credit', 'sum')
                        ).reset_index()
                        cat_melt = cat_agg.melt(id_vars='ангилал', var_name='Төрөл', value_name='Дүн')
                        fig_bar = px.bar(cat_melt, x='ангилал', y='Дүн', color='Төрөл',
                                         barmode='group', title='Ангилал бүрийн эргэлт',
                                         color_discrete_map={'Дебит':'#0F4C81','Кредит':'#FF6B35'})
                        fig_bar.update_layout(height=380, margin=dict(t=40, b=20, l=20, r=20),
                                              xaxis_title='', yaxis_title='Дүн (₮)')
                        st.plotly_chart(fig_bar, use_container_width=True)

            # ── TAB 2: Эрсдэлийн тойм ──
            with dash_tab2:
                col_c, col_d = st.columns(2)

                with col_c:
                    if 'ensemble_anomaly' in _ml_df_f.columns:
                        anom_dist = _ml_df_f['ensemble_anomaly'].value_counts().reset_index()
                        anom_dist.columns = ['Төлөв', 'Тоо']
                        anom_dist['Төлөв'] = anom_dist['Төлөв'].map({0: '🟢 Хэвийн', 1: '🔴 Аномали'})
                        fig_anom = px.pie(anom_dist, values='Тоо', names='Төлөв',
                                         title='ГБ аномали хуваарилалт',
                                         color='Төлөв',
                                         color_discrete_map={'🟢 Хэвийн':'#00897B', '🔴 Аномали':'#D32F2F'},
                                         hole=0.5)
                        fig_anom.update_layout(height=380, margin=dict(t=40, b=20, l=20, r=20))
                        st.plotly_chart(fig_anom, use_container_width=True)

                with col_d:
                    if has_journal and not _journal_f.empty and 'ml_anomaly_flag' in _journal_f.columns:
                        j_dist = _journal_f['ml_anomaly_flag'].value_counts().reset_index()
                        j_dist.columns = ['Төлөв', 'Тоо']
                        j_dist['Төлөв'] = j_dist['Төлөв'].map({0: '🟢 Хэвийн', 1: '🔴 Аномали'})
                        fig_janom = px.pie(j_dist, values='Тоо', names='Төлөв',
                                          title='ЕЖ аномали хуваарилалт',
                                          color='Төлөв',
                                          color_discrete_map={'🟢 Хэвийн':'#00897B', '🔴 Аномали':'#D32F2F'},
                                          hole=0.5)
                        fig_janom.update_layout(height=380, margin=dict(t=40, b=20, l=20, r=20))
                        st.plotly_chart(fig_janom, use_container_width=True)
                    else:
                        st.info("ЕЖ шинжилгээ хийгдээгүй байна. 3️⃣ цэсэнд хийнэ үү.")

                # Эрсдэлийн scatter plot
                if 'ensemble_anomaly' in _ml_df_f.columns and 'turnover_debit' in _ml_df_f.columns:
                    scatter_df = _ml_df_f.copy()
                    scatter_df['Эргэлт'] = pd.to_numeric(scatter_df['turnover_debit'], errors='coerce').fillna(0).abs() + \
                                            pd.to_numeric(scatter_df.get('turnover_credit', 0), errors='coerce').fillna(0).abs()
                    scatter_df['Өөрчлөлт'] = pd.to_numeric(scatter_df.get('net_change_signed', 0), errors='coerce').fillna(0).abs()
                    scatter_df['Төлөв'] = scatter_df['ensemble_anomaly'].map({0: 'Хэвийн', 1: 'Аномали'})
                    scatter_df['Дансны код'] = scatter_df['account_code'].astype(str)
                    fig_scatter = px.scatter(scatter_df, x='Эргэлт', y='Өөрчлөлт',
                                            color='Төлөв', hover_data=['Дансны код'],
                                            title='Дансны эрсдэлийн map: Эргэлт vs Өөрчлөлт',
                                            color_discrete_map={'Хэвийн':'#4CAF50', 'Аномали':'#F44336'},
                                            opacity=0.6)
                    fig_scatter.update_layout(height=450, margin=dict(t=40, b=20))
                    st.plotly_chart(fig_scatter, use_container_width=True)

            # ── TAB 3: Чиг хандлага ──
            with dash_tab3:
                if not mo_all.empty and 'month' in mo_all.columns:
                    mo_chart = mo_all.copy()
                    mo_chart['month'] = mo_chart['month'].astype(str)
                    if 'total_debit_mnt' in mo_chart.columns and 'total_credit_mnt' in mo_chart.columns:
                        mo_agg = mo_chart.groupby('month').agg(
                            Дебит=('total_debit_mnt', 'sum'),
                            Кредит=('total_credit_mnt', 'sum'),
                            Гүйлгээ=('transaction_count', 'sum')
                        ).reset_index().sort_values('month')

                        fig_trend = make_subplots(specs=[[{"secondary_y": True}]])
                        fig_trend.add_trace(
                            go.Bar(x=mo_agg['month'], y=mo_agg['Дебит'], name='Дебит', marker_color='#0F4C81', opacity=0.7),
                            secondary_y=False
                        )
                        fig_trend.add_trace(
                            go.Bar(x=mo_agg['month'], y=mo_agg['Кредит'], name='Кредит', marker_color='#FF6B35', opacity=0.7),
                            secondary_y=False
                        )
                        fig_trend.add_trace(
                            go.Scatter(x=mo_agg['month'], y=mo_agg['Гүйлгээ'], name='Гүйлгээний тоо',
                                       mode='lines+markers', marker=dict(color='#0F4C81', size=6),
                                       line=dict(color='#6A1B9A', width=2)),
                            secondary_y=True
                        )
                        fig_trend.update_layout(
                            title='Сарын эргэлт ба гүйлгээний тоо',
                            barmode='group', height=450,
                            margin=dict(t=40, b=20),
                            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1)
                        )
                        fig_trend.update_yaxes(title_text='Дүн (₮)', secondary_y=False)
                        fig_trend.update_yaxes(title_text='Гүйлгээний тоо', secondary_y=True)
                        st.plotly_chart(fig_trend, use_container_width=True)
                    else:
                        st.info("Сарын нэгтгэлийн мэдээлэл байхгүй байна.")
                elif has_journal and not _journal_f.empty and 'transaction_date' in _journal_f.columns:
                    # ЕЖ-ийн гүйлгээний хугацааны тархалт
                    j_trend = _journal_f.copy()
                    j_trend['month'] = j_trend['transaction_date'].astype(str).str[:7]
                    j_trend = j_trend[j_trend['month'].str.len() >= 7]
                    if not j_trend.empty:
                        j_mo = j_trend.groupby('month').agg(
                            Гүйлгээ=('debit_mnt', 'count'),
                            Дебит=('debit_mnt', 'sum'),
                            Кредит=('credit_mnt', 'sum')
                        ).reset_index().sort_values('month')
                        fig_jtrend = px.bar(j_mo, x='month', y=['Дебит','Кредит'],
                                            title='ЕЖ сарын эргэлт', barmode='group',
                                            color_discrete_map={'Дебит':'#0F4C81','Кредит':'#FF6B35'})
                        fig_jtrend.update_layout(height=400)
                        st.plotly_chart(fig_jtrend, use_container_width=True)
                else:
                    st.info("Сарын чиг хандлагын мэдээлэл байхгүй. Сарын нэгтгэл бэлтгэнэ үү.")

            # ── TAB 4: Топ эрсдэлтэй данс ──
            with dash_tab4:
                if 'ensemble_anomaly' in _ml_df_f.columns:
                    anom_df = _ml_df_f[_ml_df_f['ensemble_anomaly'] == 1].copy()
                    if not anom_df.empty:
                        show_cols = ['account_code']
                        if 'account_name' in anom_df.columns: show_cols.append('account_name')
                        for c in ['turnover_debit','turnover_credit','closing_debit','closing_credit','net_change_signed']:
                            if c in anom_df.columns: show_cols.append(c)
                        if 'year' in anom_df.columns: show_cols.append('year')

                        # ISA лавлагаатай тайлбар нэмэх
                        reason_parts = []
                        for _, row in anom_df.iterrows():
                            reasons = []
                            if row.get('iso_anomaly', 0) == 1: reasons.append('IF аномали (ISA 240)')
                            if row.get('zscore_anomaly', 0) == 1: reasons.append('Z-score хазайлт (ISA 520)')
                            if row.get('turn_anomaly', 0) == 1: reasons.append('Эргэлтийн харьцаа (ISA 520)')
                            tr = row.get('turn_ratio', 0)
                            if isinstance(tr, (int, float)) and abs(tr) > 5: reasons.append(f'D/C={tr:.1f}')
                            gr = row.get('growth_rate', 0)
                            if isinstance(gr, (int, float)) and abs(gr) > 0.5: reasons.append(f'Өсөлт={gr:.0%}')
                            reason_parts.append('; '.join(reasons) if reasons else 'Нэгдсэн')
                        anom_df['Эрсдэлийн шалтгаан'] = reason_parts
                        show_cols.append('Эрсдэлийн шалтгаан')

                        st.markdown(f"#### 🔴 Аномали данснууд: **{len(anom_df):,}** данс")
                        st.markdown("""
                        <div style="background: linear-gradient(135deg, #FFF8E1 0%, #FFF3E0 100%); padding: 16px 20px; border-radius: 12px; border-left: 4px solid #FF6B35; box-shadow: 0 2px 8px rgba(255,107,53,0.08); margin-bottom: 15px;">
                            <b>💡 Аудиторт:</b> Доорх данснууд нь статистик загвараар хэвийн бус гэж тодорхойлогдсон.
                            Тус бүрийн <i>Эрсдэлийн шалтгаан</i> баганыг уншиж, ОУАС (ISA) стандарттай уялдуулан нарийвчилсан шалгалт хийнэ.
                        </div>
                        """, unsafe_allow_html=True)
                        st.dataframe(anom_df[show_cols].head(50), use_container_width=True, hide_index=True)

                        # Excel татах
                        buf_dash = io.BytesIO()
                        with pd.ExcelWriter(buf_dash, engine='openpyxl') as ew:
                            anom_df[show_cols].to_excel(ew, sheet_name='Аномали_данс', index=False)
                        buf_dash.seek(0)
                        st.download_button('📥 Аномали дансуудыг Excel-ээр татах', buf_dash.getvalue(),
                                           file_name='anomaly_accounts_dashboard.xlsx',
                                           mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                           key='dl_dash_anomaly')
                    else:
                        st.success("✅ Аномали данс илрээгүй.")
                else:
                    st.info("Гүйлгээний балансын шинжилгээ хийгдээгүй. 2️⃣ цэсэнд шинжилгээ хийнэ үү.")

        # ── ЕЖ хэсгийн нэмэлт мэдээлэл ──
        if has_journal and not _journal_f.empty:
            st.markdown("---")
            st.markdown("### 📝 Ерөнхий журналын тойм")

            jcol1, jcol2, jcol3 = st.columns(3)
            with jcol1:
                if 'ml_risk_level' in _journal_f.columns:
                    risk_dist = _journal_f['ml_risk_level'].value_counts().reset_index()
                    risk_dist.columns = ['Түвшин', 'Тоо']
                    fig_risk = px.bar(risk_dist, x='Түвшин', y='Тоо',
                                     title='Эрсдэлийн түвшний хуваарилалт',
                                     color='Түвшин',
                                     color_discrete_map={
                                         '🟢 Бага':'#4CAF50', '🟠 Өндөр':'#FF9800', '🔴 Маш өндөр':'#F44336'
                                     })
                    fig_risk.update_layout(height=300, showlegend=False)
                    st.plotly_chart(fig_risk, use_container_width=True)

            with jcol2:
                if 'account_code' in _journal_f.columns:
                    top_acct = _journal_f.groupby('account_code').size().nlargest(10).reset_index()
                    top_acct.columns = ['Данс', 'Гүйлгээ']
                    fig_top = px.bar(top_acct, x='Гүйлгээ', y='Данс', orientation='h',
                                     title='Топ-10 идэвхтэй данс', color_discrete_sequence=['#0F4C81'])
                    fig_top.update_layout(height=300, yaxis={'categoryorder':'total ascending'})
                    st.plotly_chart(fig_top, use_container_width=True)

            with jcol3:
                # XAI summary
                xai_df = st.session_state.get('journal_xai', pd.DataFrame())
                if not xai_df.empty and 'feature' in xai_df.columns and 'importance' in xai_df.columns:
                    top_feat = xai_df.head(8).copy()
                    isa_map_short = {
                        'log_amount': 'ISA 240', 'benford_dev': 'ISA 240', 'is_round': 'ISA 240',
                        'amt_zscore': 'ISA 520', 'cp_rare': 'ISA 550', 'pair_rare': 'ISA 550',
                        'desc_empty': 'ISA 500', 'is_dup': 'ISA 240', 'desc_mismatch': 'ISA 500',
                        'dir_mismatch': 'ISA 240', 'is_month_end': 'ISA 240', 'is_year_end': 'ISA 240',
                        'acct_cat_num': 'ISA 315', 'is_debit': 'ISA 520', 'name_no_overlap': 'ISA 500',
                    }
                    top_feat['ISA'] = top_feat['feature'].map(isa_map_short).fillna('')
                    top_feat['label'] = top_feat['feature'] + ' (' + top_feat['ISA'] + ')'
                    fig_xai = px.bar(top_feat, x='importance', y='label', orientation='h',
                                     title='XAI: Эрсдэлийн нөлөөлөгчид',
                                     color_discrete_sequence=['#0F4C81'])
                    fig_xai.update_layout(height=300, yaxis={'categoryorder':'total ascending'}, yaxis_title='')
                    st.plotly_chart(fig_xai, use_container_width=True)

        # ── Доод хэсэг: Хурдан үйлдлүүд ──
        st.markdown("---")
        st.markdown("### ⚡ Дараагийн алхмууд")
        qa, qb, qc, qd = st.columns(4)
        with qa:
            st.markdown("""
            <div style="background:linear-gradient(135deg,#E8F5E9,#C8E6C9);padding:16px;border-radius:12px;text-align:center;border:1px solid #A5D6A7;">
                <div style="font-size:20px;margin-bottom:6px;">📊</div>
                <b style="color:#2E7D32;">ГБ шинжилгээ</b><br><span style="font-size:11px;color:#558B2F;">Гүйлгээний балансын аномали</span>
            </div>""", unsafe_allow_html=True)
        with qb:
            st.markdown("""
            <div style="background:linear-gradient(135deg,#E3F2FD,#BBDEFB);padding:16px;border-radius:12px;text-align:center;border:1px solid #90CAF9;">
                <div style="font-size:20px;margin-bottom:6px;">📝</div>
                <b style="color:#0F4C81;">ЕЖ шинжилгээ</b><br><span style="font-size:11px;color:#1976D2;">Ерөнхий журнал ML</span>
            </div>""", unsafe_allow_html=True)
        with qc:
            st.markdown("""
            <div style="background:linear-gradient(135deg,#FFF3E0,#FFE0B2);padding:16px;border-radius:12px;text-align:center;border:1px solid #FFCC80;">
                <div style="font-size:20px;margin-bottom:6px;">📐</div>
                <b style="color:#E65100;">Материаллаг</b><br><span style="font-size:11px;color:#F57C00;">ISA 320 тооцоо</span>
            </div>""", unsafe_allow_html=True)
        with qd:
            st.markdown("""
            <div style="background:linear-gradient(135deg,#F3E5F5,#E1BEE7);padding:16px;border-radius:12px;text-align:center;border:1px solid #CE93D8;">
                <div style="font-size:20px;margin-bottom:6px;">🤖</div>
                <b style="color:#6A1B9A;">Машин сургалт</b><br><span style="font-size:11px;color:#8E24AA;">Сургалт/Шалгалт</span>
            </div>""", unsafe_allow_html=True)


elif page.startswith("2"):
    st.header("2️⃣ Гүйлгээний балансын шинжилгээ")
    st.markdown("Энэ хэсэг гүйлгээний баланс болон сарын нэгтгэл файлаар ажиллана. Ерөнхий журналын машин сургалтын шинжилгээг 3️⃣ цэсэнд хийнэ.")

    all_files = st.file_uploader("📎 Нэмэлт гүйлгээний баланс / сарын нэгтгэл оруулах", type=['xlsx','xls','xlsm','xlsb','csv'], accept_multiple_files=True, key='tb_page_upload')
    if all_files:
        if not st.session_state.get('tb_upload_cache'):
            st.session_state.tb_upload_cache = uploaded_files_to_cache(all_files)
    elif st.session_state.get('tb_upload_cache'):
        all_files = cache_to_file_objects(st.session_state.tb_upload_cache)
        st.info(f"💾 Өмнө оруулсан {len(all_files)} гүйлгээний баланс файл хадгалагдсан байна.")

    tb_files = _cache_files('prepared_tb_cache')
    p1_files = _cache_files('prepared_part1_cache')

    st.caption(f"Бэлэн гүйлгээний баланс: {len(tb_files)} файл • Бэлэн сарын нэгтгэл: {len(p1_files)} файл • Нэмэлт файл: {len(all_files or [])} файл")

    c1s, c2s = st.columns(2)
    with c1s:
        cont = st.slider("🎯 Хэвийн бус дансны хувь (Isolation Forest)", 0.05, 0.20, 0.10, 0.01, key='tb_cont_work')
    with c2s:
        nest = st.slider("🌲 Random Forest модны тоо", 50, 500, 200, 50, key='tb_nest_work')

    with st.expander("🏷️ Эрсдэлийн шинжилгээнээс хасах бүлгүүд", expanded=False):
        excl_settings = {}
        for tag, rule in EXCL_RULES.items():
            excl_settings[tag] = st.checkbox(rule['label'], value=rule.get('default', False), help=rule.get('help',''), key=f'tb_excl_{tag}')

    if st.button('🚀 TB шинжилгээ эхлүүлэх', type='primary', use_container_width=True, key='run_tb_analysis_main'):
        try:
            # Нэмэлт файлуудыг товч дарсан дараа боловсруулна (render loop-аас сэргийлнэ)
            _tb_input = list(tb_files)
            _p1_input = list(p1_files)
            for f in all_files or []:
                ftype, year = detect_file_type(f)
                f.seek(0)
                if ftype == 'raw_tb':
                    buf, tb_sum = process_raw_tb(f)
                    if tb_sum is not None and not tb_sum.empty:
                        bio = io.BytesIO(buf.getvalue()); bio.name = f'TB_standardized_{year}_{Path(f.name).stem}.xlsx'
                        _tb_input.append(bio)
                elif ftype == 'tb_std':
                    _tb_input.append(f)
                elif ftype == 'part1':
                    _p1_input.append(f)
            tb_all, tb_stats = load_tb(_tb_input) if _tb_input else (pd.DataFrame(), {})
            rm_all, mo_all = load_part1(_p1_input) if _p1_input else (pd.DataFrame(), pd.DataFrame())
            tb_show = tb_all.copy()
            if not tb_show.empty:
                tb_show = classify_exclusions(tb_show, level='account')
                active_tags = [k for k, v in excl_settings.items() if v]
                if active_tags:
                    tb_show = tb_show[~tb_show['exclusion_tag'].isin(active_tags)].copy()
                ml_df, X, y, feats, res, best, fi, ym = run_ml(tb_show, cont, nest)
            else:
                ml_df, fi = pd.DataFrame(), pd.DataFrame()
            st.session_state['tb_analysis_done'] = True
            st.session_state['tb_all'] = tb_all
            st.session_state['tb_stats'] = tb_stats
            st.session_state['rm_all'] = rm_all
            st.session_state['mo_all'] = mo_all
            st.session_state['tb_filtered'] = tb_show
            st.session_state['tb_ml_df'] = ml_df
            st.session_state['tb_feature_importance'] = fi
            st.session_state['tb_error'] = ''
            st.success('✅ Гүйлгээний балансын шинжилгээ дууслаа.')
        except Exception as e:
            st.session_state['tb_error'] = str(e)
            st.exception(e)

    if st.session_state.get('tb_error'):
        st.error(st.session_state['tb_error'])

    if st.session_state.get('tb_analysis_done', False):
        tb_stats = st.session_state.get('tb_stats', {})
        rm_all = st.session_state.get('rm_all', pd.DataFrame())
        mo_all = st.session_state.get('mo_all', pd.DataFrame())
        ml_df = st.session_state.get('tb_ml_df', pd.DataFrame())
        fi = st.session_state.get('tb_feature_importance', pd.DataFrame())

        if tb_stats:
            rows = [{'Он': yr, 'Данс': vals.get('accounts', 0), 'Дебит эргэлт': vals.get('turnover_d', 0), 'Кредит эргэлт': vals.get('turnover_c', 0)} for yr, vals in sorted(tb_stats.items())]
            st.markdown('### 📊 Гүйлгээний балансын нэгтгэл')
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        if not ml_df.empty:
            st.markdown('### 🤖 Гүйлгээний балансын аномали шинжилгээ')
            # Салбар + Он шүүлтүүр
            ml_show = _render_branch_year_filter(ml_df, 'tb')
            # Нийт ба эрсдэлтэй дансны тоо
            total_accounts = ml_show['account_code'].nunique() if 'account_code' in ml_show.columns else len(ml_show)
            anomaly_accounts = ml_show[ml_show.get('ensemble_anomaly', pd.Series(dtype=int)) == 1]['account_code'].nunique() if 'ensemble_anomaly' in ml_show.columns and 'account_code' in ml_show.columns else 0
            tc1, tc2 = st.columns(2)
            tc1.metric("📊 Нийт данс", f"{total_accounts:,}")
            tc2.metric("⚠️ Эрсдэлтэй данс", f"{anomaly_accounts:,}")

            # ── TB шинжилгээний дэлгэрэнгүй tab-ууд ──
            tb_viz_tabs = st.tabs(["📋 Данс жагсаалт", "⚖️ Баланс шалгалт", "📊 Визуализаци", "🔍 Эрсдэлийн тайлбар", "📥 Тайлан"])
            with tb_viz_tabs[0]:
                show_cols = [c for c in ['year','account_code','account_name','opening_debit','opening_credit','turnover_debit','turnover_credit','closing_debit','closing_credit','iso_anomaly','zscore_anomaly','turn_anomaly','ensemble_anomaly'] if c in ml_show.columns]
                st.dataframe(ml_show.sort_values(['ensemble_anomaly','year'] if 'year' in ml_show.columns else ['ensemble_anomaly'], ascending=[False, True] if 'year' in ml_show.columns else [False])[show_cols].head(500), use_container_width=True, hide_index=True)
                _show_dataframe_download(ml_show, 'tb_anomaly_results.csv')

            with tb_viz_tabs[1]:
                # ═══ БАЛАНС ШАЛГАЛТ ═══
                st.markdown("""
                <div style="background:linear-gradient(135deg,#E3F2FD,#E8EAF6);padding:16px 20px;border-radius:12px;border-left:4px solid #0F4C81;box-shadow:0 2px 8px rgba(15,76,129,0.08);margin-bottom:15px;">
                    <b>⚖️ Гүйлгээний балансын стандарт шалгалт (ISA 505, 520)</b><br>
                    <span style="font-size:13px;">Эхний үлдэгдэл + Эргэлт = Эцсийн үлдэгдэл тэнцэж байгаа эсэх, дансны үлдэгдлийн чиглэл зөв эсэхийг шалгана.</span>
                </div>""", unsafe_allow_html=True)

                bal = ml_show.copy()
                bal_checks = []
                for _, row in bal.iterrows():
                    checks = {}
                    checks['Он'] = row.get('year', '')
                    checks['Данс'] = row.get('account_code', '')
                    checks['Нэр'] = str(row.get('account_name', ''))[:40]
                    od = float(row.get('opening_debit', 0) or 0)
                    oc = float(row.get('opening_credit', 0) or 0)
                    td = float(row.get('turnover_debit', 0) or 0)
                    tc = float(row.get('turnover_credit', 0) or 0)
                    cd = float(row.get('closing_debit', 0) or 0)
                    cc = float(row.get('closing_credit', 0) or 0)

                    # 1. Баланс тэнцэл: Эхний + Эргэлт = Эцсийн
                    expected_close_d = od + td - tc
                    expected_close_c = oc + tc - td
                    actual_close = cd - cc
                    expected_close = od - oc + td - tc
                    diff = actual_close - expected_close
                    checks['Эхний (Д-К)'] = round(od - oc, 2)
                    checks['Эргэлт (Д-К)'] = round(td - tc, 2)
                    checks['Эцсийн (Д-К)'] = round(cd - cc, 2)
                    checks['Хүлээгдэж буй эцсийн'] = round(expected_close, 2)
                    checks['Зөрүү'] = round(diff, 2)
                    checks['Баланс тэнцсэн'] = '✅' if abs(diff) < 1.0 else '❌'

                    # 2. Дансны чиглэл зөв эсэх
                    code_first = str(row.get('account_code', ''))[:1]
                    dir_ok = '✅'
                    dir_note = ''
                    if code_first in ('1', '4', '7'):  # Хөрөнгө, Зардал → дебит үлдэгдэл
                        if cd - cc < -1000:
                            dir_ok = '⚠️'
                            dir_note = f'Хөрөнгө/Зардал данс кредит үлдэгдэлтэй ({cd-cc:,.0f}₮)'
                    elif code_first in ('2', '3', '5', '6'):  # Өр, Өмч, Орлого → кредит үлдэгдэл
                        if cd - cc > 1000:
                            dir_ok = '⚠️'
                            dir_note = f'Өр/Орлого данс дебит үлдэгдэлтэй ({cd-cc:,.0f}₮)'
                    checks['Чиглэл'] = dir_ok
                    checks['Тайлбар'] = dir_note

                    # 3. Тэг эргэлттэй ч үлдэгдэлтэй данс
                    if td == 0 and tc == 0 and (abs(cd) > 0 or abs(cc) > 0):
                        checks['Тайлбар'] = (checks['Тайлбар'] + ' | ' if checks['Тайлбар'] else '') + 'Эргэлтгүй ч үлдэгдэлтэй'

                    bal_checks.append(checks)

                bal_df = pd.DataFrame(bal_checks)

                # Нэгтгэл metrics
                total_accts = len(bal_df)
                balanced = int((bal_df['Баланс тэнцсэн'] == '✅').sum())
                unbalanced = total_accts - balanced
                dir_issues = int((bal_df['Чиглэл'] == '⚠️').sum())

                bm1, bm2, bm3, bm4 = st.columns(4)
                bm1.metric("📊 Нийт данс", f"{total_accts:,}")
                bm2.metric("✅ Тэнцсэн", f"{balanced:,}", delta=f"{balanced/max(total_accts,1)*100:.1f}%")
                bm3.metric("❌ Тэнцээгүй", f"{unbalanced:,}", delta=f"{unbalanced}" if unbalanced > 0 else "0", delta_color="inverse")
                bm4.metric("⚠️ Чиглэл зөрчил", f"{dir_issues:,}")

                # Асуудалтай дансуудыг эхэнд
                bal_show = bal_df.sort_values(['Баланс тэнцсэн', 'Чиглэл'], ascending=[True, True])
                st.dataframe(bal_show, use_container_width=True, hide_index=True)

                # ═══ САНХҮҮГИЙН ТАЙЛАНГИЙН ТОХИРУУЛГА ═══
                st.markdown("---")
                st.markdown("### 📝 Санхүүгийн тайлангийн тохируулга (залруулах бичилт)")
                st.markdown("""
                <div style="background:linear-gradient(135deg,#FFF8E1,#FFF3E0);padding:14px 18px;border-radius:12px;border-left:4px solid #FF6B35;margin-bottom:12px;">
                    <span style="font-size:13px;">⚠️ <b>ISA 450:</b> Аудитын явцад илрүүлсэн алдааг нэгтгэж, материаллаг байдлын босготой харьцуулна.
                    Доорх жагсаалт нь баланс тэнцээгүй болон чиглэл зөрсөн дансуудын залруулах бичилтийн санал юм.</span>
                </div>""", unsafe_allow_html=True)

                adj_rows = []
                for _, row in bal_df.iterrows():
                    if row['Баланс тэнцсэн'] == '❌' and abs(row['Зөрүү']) >= 1.0:
                        diff_val = row['Зөрүү']
                        adj_rows.append({
                            'Данс': row['Данс'],
                            'Нэр': row['Нэр'],
                            'Төрөл': 'Баланс зөрүү',
                            'Зөрүүний дүн': diff_val,
                            'Залруулах бичилт': f"Дт {row['Данс']} / Кт Тохируулга: {abs(diff_val):,.0f}₮" if diff_val > 0 else f"Дт Тохируулга / Кт {row['Данс']}: {abs(diff_val):,.0f}₮",
                            'ISA лавлагаа': 'ISA 450 (Аудитын алдааны үнэлгээ)',
                        })
                    if row['Чиглэл'] == '⚠️' and row['Тайлбар']:
                        adj_rows.append({
                            'Данс': row['Данс'],
                            'Нэр': row['Нэр'],
                            'Төрөл': 'Чиглэлийн зөрчил',
                            'Зөрүүний дүн': row['Эцсийн (Д-К)'],
                            'Залруулах бичилт': f"Дансны үлдэгдлийн чиглэл шалгах: {row['Тайлбар']}",
                            'ISA лавлагаа': 'ISA 240 (Залилан), ISA 505 (Баталгаажуулалт)',
                        })
                if adj_rows:
                    adj_df = pd.DataFrame(adj_rows)
                    st.success(f"📝 Нийт {len(adj_df)} тохируулгын санал илэрлээ")
                    st.dataframe(adj_df, use_container_width=True, hide_index=True)
                    _show_dataframe_download(adj_df, 'тохируулгын_санал.csv')
                else:
                    st.success("✅ Баланс тэнцэж, чиглэлийн зөрчил байхгүй байна. Тохируулга шаардлагагүй.")

                # Баланс шалгалтын Excel
                bal_buf = io.BytesIO()
                with pd.ExcelWriter(bal_buf, engine='openpyxl') as bw:
                    bal_df.to_excel(bw, sheet_name='Баланс_шалгалт', index=False)
                    if adj_rows:
                        pd.DataFrame(adj_rows).to_excel(bw, sheet_name='Тохируулгын_санал', index=False)
                bal_buf.seek(0)
                st.download_button('📥 Баланс шалгалт + Тохируулга (Excel)', bal_buf.getvalue(),
                                   file_name='balance_check_adjustments.xlsx',
                                   mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                   key='dl_bal_check')

            with tb_viz_tabs[4]:
                # Scatter plot: Эргэлт vs Өөрчлөлт
                vcA, vcB = st.columns(2)
                with vcA:
                    if 'turnover_debit' in ml_show.columns and 'ensemble_anomaly' in ml_show.columns:
                        scat = ml_show.copy()
                        scat['Эргэлт'] = pd.to_numeric(scat.get('turnover_debit',0), errors='coerce').fillna(0).abs() + pd.to_numeric(scat.get('turnover_credit',0), errors='coerce').fillna(0).abs()
                        scat['Өөрчлөлт'] = pd.to_numeric(scat.get('net_change_signed',0), errors='coerce').fillna(0).abs()
                        scat['Төлөв'] = scat['ensemble_anomaly'].map({0:'Хэвийн',1:'Аномали'})
                        fig_scat = px.scatter(scat, x='Эргэлт', y='Өөрчлөлт', color='Төлөв',
                                              hover_data=['account_code'], opacity=0.6,
                                              color_discrete_map={'Хэвийн':'#00897B','Аномали':'#D32F2F'},
                                              title='Эргэлт vs Өөрчлөлт (ISA 520 аналитик)')
                        fig_scat.update_layout(height=380)
                        st.plotly_chart(fig_scat, use_container_width=True)
                with vcB:
                    if 'turn_ratio' in ml_show.columns and 'growth_rate' in ml_show.columns:
                        scat2 = ml_show.copy()
                        scat2['Төлөв'] = scat2.get('ensemble_anomaly', pd.Series(0,index=scat2.index)).map({0:'Хэвийн',1:'Аномали'})
                        fig_scat2 = px.scatter(scat2, x='turn_ratio', y='growth_rate', color='Төлөв',
                                               hover_data=['account_code'], opacity=0.6,
                                               color_discrete_map={'Хэвийн':'#00897B','Аномали':'#D32F2F'},
                                               title='D/C харьцаа vs Өсөлтийн хурд (ISA 240)')
                        fig_scat2.update_layout(height=380, xaxis_title='Дебит/Кредит харьцаа', yaxis_title='Өсөлтийн хурд')
                        st.plotly_chart(fig_scat2, use_container_width=True)

                # Treemap: дансны бүтэц + эрсдэл
                if 'account_code' in ml_show.columns and 'turnover_debit' in ml_show.columns:
                    tm = ml_show.copy()
                    tm['ангилал'] = tm['account_code'].astype(str).str[:1].map(
                        {'1':'Хөрөнгө','2':'Өр','3':'Эздийн өмч','4':'Зардал','5':'Орлого','6':'Орлого','7':'Үйл аж. зардал','8':'Бусад'}
                    ).fillna('Бусад')
                    tm['эргэлт'] = pd.to_numeric(tm.get('turnover_debit',0),errors='coerce').fillna(0).abs() + pd.to_numeric(tm.get('turnover_credit',0),errors='coerce').fillna(0).abs()
                    tm['эрсдэл'] = tm.get('ensemble_anomaly',0).astype(str).map({'0':'Хэвийн','1':'Аномали'}).fillna('Хэвийн')
                    tm_agg = tm.groupby(['ангилал','эрсдэл']).agg(Эргэлт=('эргэлт','sum'), Данс=('account_code','nunique')).reset_index()
                    if not tm_agg.empty:
                        fig_tm = px.treemap(tm_agg, path=['ангилал','эрсдэл'], values='Эргэлт',
                                            color='эрсдэл', color_discrete_map={'Хэвийн':'#00897B','Аномали':'#D32F2F'},
                                            title='Дансны бүтэц ба эрсдэл (ISA 315)')
                        fig_tm.update_layout(height=420)
                        st.plotly_chart(fig_tm, use_container_width=True)

            with tb_viz_tabs[4]:
                st.markdown("""
                <div style="background: linear-gradient(135deg, #fff3e0 0%, #fffde7 100%);
                     padding: 20px; border-radius: 12px; border-left: 5px solid #E65100; margin-bottom: 20px;">
                    <h4 style="color: #E65100; margin-top: 0;">💡 Аудиторт зориулсан эрсдэлийн тайлбар</h4>
                    <p style="font-size: 14px; line-height: 1.7;">
                        Аномали илрүүлэлт нь 3 бие даасан алгоритмаар хийгдэнэ:<br>
                        <b>• Isolation Forest (ISA 240)</b> — Залилангийн эрсдэлийг тусгаарлалтын аргаар илрүүлнэ<br>
                        <b>• Z-score хазайлт (ISA 520)</b> — Дансны утга дунджаас хэт зөрсөн эсэхийг тоон аргаар тодорхойлно<br>
                        <b>• Эргэлтийн харьцаа (ISA 520)</b> — Дебит/Кредит харьцаа хэвийн бус байгаа эсэхийг шалгана<br>
                        Хоёр буюу түүнээс дээш алгоритм зөвшөөрвөл "аномали" гэж тэмдэглэнэ.
                    </p>
                </div>
                """, unsafe_allow_html=True)
                if 'ensemble_anomaly' in ml_show.columns:
                    anom_explain = ml_show[ml_show['ensemble_anomaly']==1].copy()
                    if not anom_explain.empty:
                        explain_parts = []
                        for _, row in anom_explain.iterrows():
                            reasons = []
                            if row.get('iso_anomaly',0)==1: reasons.append('🔴 IF аномали (ISA 240 — залилангийн эрсдэл)')
                            if row.get('zscore_anomaly',0)==1: reasons.append('🟠 Z-score хэт хазайлт (ISA 520 — аналитик горим)')
                            if row.get('turn_anomaly',0)==1: reasons.append('🟡 Эргэлтийн харьцаа (ISA 520)')
                            tr = row.get('turn_ratio',0)
                            gr = row.get('growth_rate',0)
                            extra = []
                            if isinstance(tr,(int,float)) and abs(tr)>5: extra.append(f'D/C харьцаа={tr:.1f}')
                            if isinstance(gr,(int,float)) and abs(gr)>0.5: extra.append(f'Өсөлт={gr:.0%}')
                            explain_parts.append({
                                'Данс': row.get('account_code',''),
                                'Нэр': row.get('account_name',''),
                                'Эрсдэлийн шалтгаан': ' | '.join(reasons),
                                'Нэмэлт': '; '.join(extra),
                                'Аудитын горим': 'Нарийвчилсан шалгалт (ISA 330.18)' if len(reasons)>=2 else 'Шинжилгээний процедур (ISA 520)'
                            })
                        explain_df = pd.DataFrame(explain_parts)
                        st.dataframe(explain_df, use_container_width=True, hide_index=True)

            with tb_viz_tabs[4]:
                st.markdown("### 📥 Аудитын тайлан татах")
                st.markdown("""
                <div style="background:linear-gradient(135deg,#E0F2F1,#E8F5E9);padding:16px 20px;border-radius:12px;border-left:4px solid #00897B;box-shadow:0 2px 8px rgba(0,137,123,0.08);margin-bottom:15px;">
                    <b>📋 Аудитын working paper формат:</b> Аномали данс, баланс шалгалт, эрсдэлийн тайлбар,
                    ISA лавлагаа, аудитын горимын санал, тохируулгын санал бүгд багтсан.
                </div>""", unsafe_allow_html=True)
                if 'ensemble_anomaly' in ml_show.columns:
                    report_buf = io.BytesIO()
                    with pd.ExcelWriter(report_buf, engine='openpyxl') as rw:
                        # Sheet 1: Бүх данс
                        rpt_cols = [c for c in ['year','account_code','account_name','opening_debit','opening_credit',
                                    'turnover_debit','turnover_credit','closing_debit','closing_credit',
                                    'net_change_signed','ensemble_anomaly','iso_anomaly',
                                    'zscore_anomaly','turn_anomaly','turn_ratio','growth_rate'] if c in ml_show.columns]
                        ml_show[rpt_cols].to_excel(rw, sheet_name='Бүх_данс', index=False)
                        # Sheet 2: Аномали данс + ISA тайлбар + алгоритмын код
                        anom_rpt = ml_show[ml_show['ensemble_anomaly']==1].copy()
                        if not anom_rpt.empty:
                            # Эрсдэлийн шалтгаан + алгоритмын код нэмэх
                            isa_reasons, algo_codes, audit_procedures = [], [], []
                            for _, row in anom_rpt.iterrows():
                                reasons, codes = [], []
                                if row.get('iso_anomaly',0)==1:
                                    reasons.append('ISA 240: Залилангийн эрсдэл')
                                    codes.append('IsolationForest(n=200, contamination=0.10).fit_predict(X)==-1')
                                if row.get('zscore_anomaly',0)==1:
                                    reasons.append('ISA 520: Аналитик горим (Z-score)')
                                    codes.append('StandardScaler().fit_transform(X) → max(|z|) > 2.0')
                                if row.get('turn_anomaly',0)==1:
                                    reasons.append('ISA 520: Эргэлтийн харьцаа')
                                    codes.append('turn_ratio = D/C → |ratio| > P95(all_ratios)')
                                tr = row.get('turn_ratio', 0)
                                gr = row.get('growth_rate', 0)
                                if isinstance(tr,(int,float)) and abs(tr)>5:
                                    reasons.append(f'ISA 240: D/C={tr:.1f} (хэвийн: 0.8-1.2)')
                                if isinstance(gr,(int,float)) and abs(gr)>0.5:
                                    reasons.append(f'ISA 520: Өсөлт={gr:.0%} (>50%)')
                                    codes.append(f'growth=(close-open)/open={gr:.2f}')
                                isa_reasons.append(' | '.join(reasons))
                                algo_codes.append('\n'.join(codes))
                                n_reasons = sum([row.get('iso_anomaly',0), row.get('zscore_anomaly',0), row.get('turn_anomaly',0)])
                                audit_procedures.append('Нарийвчилсан шалгалт (ISA 330.18)' if n_reasons>=2 else 'Шинжилгээний процедур (ISA 520)')
                            anom_rpt['ISA_эрсдэлийн_шалтгаан'] = isa_reasons
                            anom_rpt['Алгоритмын_код'] = algo_codes
                            anom_rpt['Аудитын_горим_санал'] = audit_procedures
                            out_cols = [c for c in ['year','account_code','account_name','turnover_debit','turnover_credit',
                                        'closing_debit','closing_credit','net_change_signed','turn_ratio','growth_rate',
                                        'iso_anomaly','zscore_anomaly','turn_anomaly','ensemble_anomaly',
                                        'ISA_эрсдэлийн_шалтгаан','Алгоритмын_код','Аудитын_горим_санал'] if c in anom_rpt.columns]
                            anom_rpt[out_cols].to_excel(rw, sheet_name='Аномали_данс', index=False)

                        # Sheet 3: Баланс шалгалт
                        if 'bal_df' in dir() and not bal_df.empty:
                            bal_df.to_excel(rw, sheet_name='Баланс_шалгалт', index=False)
                        # Sheet 4: Тохируулгын санал
                        if 'adj_rows' in dir() and adj_rows:
                            pd.DataFrame(adj_rows).to_excel(rw, sheet_name='Тохируулгын_санал', index=False)
                        # Sheet 5: Feature importance
                        if not fi.empty:
                            fi.to_excel(rw, sheet_name='Feature_importance', index=False)
                        # Sheet 6: Алгоритмын тайлбар
                        algo_ref = pd.DataFrame([
                            {'Алгоритм':'Isolation Forest','ISA':'ISA 240','Код':'IsolationForest(n_estimators=200, contamination=0.10, random_state=42)','Томьёо':'s(x,n) = 2^(-h(x)/c(n)), h(x)=path length','Тайлбар':'Бусад гүйлгээнээс ялгаатай хэв маяг илрүүлэх. Замын урт богино = аномали.'},
                            {'Алгоритм':'Z-score хазайлт','ISA':'ISA 520','Код':'StandardScaler().fit_transform(X) → np.abs(z).max(axis=1) > 2.0','Томьёо':'z = (x − μ) / σ, |z| > 2.0','Тайлбар':'Дансны утга дунджаас хэт зөрсөн эсэх. 2σ гадна = 95%+ ялгаатай.'},
                            {'Алгоритм':'Эргэлтийн харьцаа','ISA':'ISA 520','Код':'turn_ratio = turnover_debit / turnover_credit → |ratio| > P95','Томьёо':'D/C = Дебит / Кредит, P95 босго','Тайлбар':'Дебит/Кредит харьцаа 95-р хувиас дээш = хэвийн бус.'},
                            {'Алгоритм':'Ансамбль санал','ISA':'ISA 330','Код':'ensemble = (IF==1) | ((Z==1) & (Turn==1)), vote = IF+Z+Turn','Томьёо':'vote ≥ 2 → аномали','Тайлбар':'2+ алгоритм зөвшөөрвөл аномали. Олонхийн зарчим.'},
                            {'Алгоритм':'Өсөлтийн хурд','ISA':'ISA 520','Код':'growth = (closing_bal - opening_bal) / opening_bal','Томьёо':'|growth| > 50% → анхаарал','Тайлбар':'Он дамнасан өөрчлөлт хэт их бол бизнесийн эрсдэл.'},
                            {'Алгоритм':'Баланс тэнцэл','ISA':'ISA 505','Код':'expected = opening + debit - credit; diff = actual - expected','Томьёо':'Эхний + Эргэлт = Эцсийн','Тайлбар':'Баланс тэнцэхгүй бол бүртгэлийн алдаа.'},
                            {'Алгоритм':'Чиглэлийн зөрчил','ISA':'ISA 240','Код':'if code[0]=="1" and credit>0 and debit==0: dir_mismatch=1','Томьёо':'Хөрөнгө=Дебит, Өр=Кредит','Тайлбар':'Хөрөнгийн данс кредитлэгдсэн → залилангийн дохио.'},
                        ])
                        algo_ref.to_excel(rw, sheet_name='Алгоритмын_лавлагаа', index=False)

                    report_buf.seek(0)
                    st.download_button('📥 Аудитын бүрэн тайлан (Excel)', report_buf.getvalue(),
                                       file_name='audit_tb_full_report.xlsx',
                                       mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                       key='dl_tb_audit_report')
        if not fi.empty:
            st.markdown('### 🔎 Шинж чанарын ач холбогдол')
            st.dataframe(fi, use_container_width=True, hide_index=True)
        if not rm_all.empty:
            st.markdown('### 📋 Эрсдэлийн матриц (харилцагч×данс)')
            st.dataframe(rm_all.head(300), use_container_width=True, hide_index=True)
        if not mo_all.empty and 'month' in mo_all.columns and 'total_debit_mnt' in mo_all.columns:
            st.markdown('### 📈 Сарын хандлага')
            mo_plot = mo_all.copy()
            mo_plot['total_debit_mnt'] = pd.to_numeric(mo_plot['total_debit_mnt'], errors='coerce').fillna(0)
            # Он-оор шүүх
            if 'year' in mo_plot.columns or 'report_year' in mo_plot.columns:
                yr_col_mo = 'year' if 'year' in mo_plot.columns else 'report_year'
                yrs_mo = ['Бүгд'] + sorted(mo_plot[yr_col_mo].dropna().unique().astype(str).tolist())
                sel_yr_mo = st.selectbox("📅 Он сонгох", yrs_mo, key='tb_month_year')
                if sel_yr_mo != 'Бүгд':
                    mo_plot = mo_plot[mo_plot[yr_col_mo].astype(str) == sel_yr_mo]
                fig_mo = px.line(mo_plot, x='month', y='total_debit_mnt', color=yr_col_mo, title='Сарын дебит хөдөлгөөн (он тус бүрээр)')
            else:
                fig_mo = px.line(mo_plot, x='month', y='total_debit_mnt', title='Сарын дебит хөдөлгөөн')
            fig_mo.update_layout(yaxis_title='Дебит дүн (₮)', xaxis_title='Сар')
            st.plotly_chart(fig_mo, use_container_width=True)
    else:
        st.info('👆 Энд шууд файл оруулах эсвэл 1️⃣ цэсэнд бэлтгэсэн гүйлгээний баланс файлаа ашиглаад шинжилгээг эхлүүлнэ үү.')

elif page.startswith("3"):
    st.header("3️⃣ Ерөнхий журналын шинжилгээ")
    st.markdown("Ерөнхий журналын гүйлгээний түвшний машин сургалт + тайлбарлагдах ХОУ шинжилгээ.")

    ej_files = st.file_uploader("📎 Нэмэлт ЕЖ эсвэл ledger файл оруулах", type=['xlsx','xls','xlsm','xlsb','csv','tsv','gz'], accept_multiple_files=True, key='journal_page_upload')
    if ej_files:
        if not st.session_state.get('journal_upload_cache'):
            st.session_state.journal_upload_cache = uploaded_files_to_cache(ej_files)
    elif st.session_state.get('journal_upload_cache'):
        ej_files = cache_to_file_objects(st.session_state.journal_upload_cache)
        st.info(f"💾 Өмнө оруулсан {len(ej_files)} journal файл хадгалагдсан байна.")

    prepared_ledger_files = _cache_files('prepared_ledger_cache')
    st.caption(f"Бэлэн ерөнхий журнал: {len(prepared_ledger_files)} файл • Нэмэлт файл: {len(ej_files or [])} файл")

    cj1, cj2 = st.columns(2)
    with cj1:
        j_cont = st.slider('🎯 Аномалийн хувь', 0.01, 0.20, 0.05, 0.01, key='j_cont_work')
    with cj2:
        j_clusters = st.slider('🧩 KMeans кластерын тоо', 2, 20, 8, 1, key='j_clusters_work')

    with st.expander("🏷️ Эрсдэлийн шинжилгээнээс хасах бүлгүүд", expanded=False):
        excl_settings_j = {}
        for tag, rule in EXCL_RULES.items():
            excl_settings_j[tag] = st.checkbox(rule['label'], value=rule.get('default', False), help=rule.get('help',''), key=f'j_excl_{tag}')

    if st.button('🚀 Ерөнхий журналын шинжилгээ эхлүүлэх', type='primary', use_container_width=True, key='run_journal_analysis_main'):
        try:
            # Файлуудыг товч дарсан дараа боловсруулна (render loop-аас сэргийлнэ)
            journal_inputs = []
            for f in prepared_ledger_files:
                journal_inputs.append(('ledger', f, get_year(f.name)))
            for f in ej_files or []:
                ftype, year = detect_file_type(f)
                f.seek(0)
                if ftype in ('ledger', 'edt'):
                    journal_inputs.append((ftype, f, year))

            frames = []
            ledger_stats_j = {}
            step_rows = []
            st.session_state['journal_process_rows'] = []
            status_box = st.empty()
            progress_box = st.progress(0, text='ЕЖ шинжилгээ эхэлж байна...')
            table_box = st.empty()
            process_box = st.empty()

            total_inputs = max(len(journal_inputs), 1)
            for idx, (typ, f, year) in enumerate(journal_inputs, start=1):
                file_name = Path(getattr(f, 'name', f'journal_{year}')).name
                # Салбар таних
                _bid, _blabel = detect_branch(file_name)
                step_info = {
                    '№': idx,
                    'Файл': file_name,
                    'Салбар': _blabel,
                    'Төрөл': FILE_TYPE_LABELS.get(typ, FILE_TYPE_LABELS['unknown'])[0],
                    'Он': year,
                    'Алхам': 'Уншиж эхэлсэн',
                    'Төлөв': '⏳',
                    'Нийт мөр': 0,
                    'Sample мөр': 0,
                    'Данс': 0,
                    'Эх үүсвэр': ''
                }
                step_rows.append(step_info)
                status_box.info(f'⏳ {idx}/{total_inputs} файл уншиж байна: {file_name}')

                def _journal_progress_cb(stage, status='⏳', detail=''):
                    _push_process_row('journal', file_name, stage, status=status, detail=detail, year=year, ftype=FILE_TYPE_LABELS.get(typ, FILE_TYPE_LABELS['unknown'])[0])
                    process_box.empty()
                    with process_box.container():
                        _render_process_log('journal', '👁️ ЕЖ шинжилгээний явцын дэлгэрэнгүй', expanded=True)

                _journal_progress_cb('Файл уншилт эхэллээ', '⏳', 'Өгөгдлийн эх үүсвэрийг нээж sample бэлтгэх гэж байна')
                table_box.empty()
                with table_box.container():
                    _render_reading_steps(step_rows, expanded=True)

                sample_df, meta = _fast_read_journal_input(typ, f, year, max_rows=50000, chunksize=120000, progress_cb=_journal_progress_cb)
                rows = int(meta.get('rows', 0))
                accounts = int(meta.get('accounts', 0))
                sample_rows = int(meta.get('sample_rows', len(sample_df)))
                source = str(meta.get('source', ''))

                if sample_df is not None and not sample_df.empty:
                    sample_df['branch_id'] = _bid
                    sample_df['branch_label'] = _blabel
                    frames.append(sample_df)
                    ledger_stats_j[file_name] = {
                        'rows': rows if rows > 0 else len(sample_df),
                        'accounts': accounts,
                        'sample_rows': sample_rows,
                        'source': source
                    }
                    step_info.update({
                        'Алхам': 'Уншиж дууссан',
                        'Төлөв': '✅',
                        'Нийт мөр': rows if rows > 0 else len(sample_df),
                        'Sample мөр': sample_rows,
                        'Данс': accounts,
                        'Эх үүсвэр': source
                    })
                else:
                    step_info.update({
                        'Алхам': 'Өгөгдөл олдсонгүй',
                        'Төлөв': '⚠️',
                        'Эх үүсвэр': source
                    })

                progress_box.progress(min(idx / total_inputs, 1.0), text=f'Файл уншиж байна... {idx}/{total_inputs}')
                table_box.empty()
                with table_box.container():
                    _render_reading_steps(step_rows, expanded=True)
                with process_box.container():
                    _render_process_log('journal', '👁️ ЕЖ шинжилгээний явцын дэлгэрэнгүй', expanded=True)

            if not frames:
                raise ValueError('Journal/ledger өгөгдөл олдсонгүй.')

            _push_process_row('journal', 'Нэгдсэн dataset', 'Sample-уудыг нэгтгэж байна', status='⏳', detail='Уншсан файлуудын sample-уудыг нэг DataFrame болгон нэгтгэж байна', rows=sum(len(x) for x in frames), sample_rows=sum(len(x) for x in frames))
            with process_box.container():
                _render_process_log('journal', '👁️ ЕЖ шинжилгээний явцын дэлгэрэнгүй', expanded=True)
            status_box.info('🧠 Машин сургалтын шинжилгээ эхэлж байна...')
            progress_box.progress(0.82, text='Машин сургалтын шинжилгээ эхлүүлж байна...')
            ledger_sample_j = pd.concat(frames, ignore_index=True)
            _push_process_row('journal', 'Нэгдсэн dataset', 'Өгөгдлийг цэвэрлэж байна', status='⏳', detail='Risk analysis-д зориулж хоосон болон алдаатай утгыг цэвэрлэж байна', rows=len(ledger_sample_j), sample_rows=len(ledger_sample_j))
            ledger_sample_j = clean_for_risk(ledger_sample_j)
            with process_box.container():
                _render_process_log('journal', '👁️ ЕЖ шинжилгээний явцын дэлгэрэнгүй', expanded=True)
            progress_box.progress(0.90, text='Загвар ба тайлбарлагдах ХОУ тооцоолж байна...')
            _push_process_row('journal', 'ML engine', 'Аномали илрүүлэлт ажиллаж байна', status='⏳', detail=f'contamination={j_cont:.2f}, clusters={j_clusters}')
            with process_box.container():
                _render_process_log('journal', '👁️ ЕЖ шинжилгээний явцын дэлгэрэнгүй', expanded=True)
            ml_result_j, ml_feats_j, model_summary_j, xai_importance_j = run_txn_ml_ensemble(ledger_sample_j, contamination=j_cont, n_clusters=j_clusters)
            if ml_result_j.empty:
                raise ValueError('Машин сургалтын шинжилгээ хийхэд хангалттай мөр алга.')
            ml_result_j = classify_exclusions(ml_result_j, level='transaction')
            active_tags_j = [k for k, v in excl_settings_j.items() if v]
            ml_show_j = ml_result_j[~ml_result_j['exclusion_tag'].isin(active_tags_j)].copy() if active_tags_j else ml_result_j.copy()
            st.session_state['journal_ai_done'] = True
            st.session_state['journal_ml_result'] = ml_result_j
            st.session_state['journal_ml_show'] = ml_show_j
            st.session_state['journal_model_summary'] = model_summary_j
            st.session_state['journal_xai'] = xai_importance_j
            st.session_state['journal_ledger_stats'] = ledger_stats_j
            st.session_state['journal_read_steps'] = step_rows
            st.session_state['journal_error'] = ''
            _push_process_row('journal', 'ML engine', 'ЕЖ шинжилгээ дууслаа', status='✅', detail=f'Эцсийн үр дүн: {len(ml_result_j):,} мөр')
            progress_box.progress(1.0, text='ЕЖ шинжилгээ бүрэн дууслаа.')
            with process_box.container():
                _render_process_log('journal', '👁️ ЕЖ шинжилгээний явцын дэлгэрэнгүй', expanded=True)
            status_box.success('✅ Ерөнхий журналын шинжилгээ дууслаа.')
            st.success('✅ Ерөнхий журналын шинжилгээ дууслаа.')
        except Exception as e:
            st.session_state['journal_error'] = str(e)
            st.exception(e)

    if st.session_state.get('journal_error'):
        st.error(st.session_state['journal_error'])

    if st.session_state.get('journal_ai_done', False):
        ml_show_j = st.session_state.get('journal_ml_show', pd.DataFrame())
        model_summary_j = st.session_state.get('journal_model_summary', pd.DataFrame())
        xai_importance_j = st.session_state.get('journal_xai', pd.DataFrame())
        ledger_stats_j = st.session_state.get('journal_ledger_stats', {})

        if ledger_stats_j:
            st.markdown('### 📊 Ерөнхий журналын нэгтгэл')
            st.dataframe(pd.DataFrame([{'Файл': k, 'Мөр': v.get('rows', 0), 'Sample мөр': v.get('sample_rows', 0), 'Данс': v.get('accounts', 0), 'Эх үүсвэр': v.get('source', '')} for k, v in ledger_stats_j.items()]), use_container_width=True, hide_index=True)
        read_steps = st.session_state.get('journal_read_steps', [])
        _render_reading_steps(read_steps, expanded=False)
        if not model_summary_j.empty:
            st.markdown('### 🤖 Алгоритмын тойм')
            st.dataframe(model_summary_j, use_container_width=True, hide_index=True)
        if not ml_show_j.empty:
            st.markdown('### 🔍 Ерөнхий журналын шинжилгээний үр дүн')

            # ── Салбар + Он шүүлтүүр (бүх tab-д нөлөөлнө) ──
            # transaction_date-аас year баганыг нэмэх (байхгүй бол)
            if 'year' not in ml_show_j.columns and 'report_year' not in ml_show_j.columns:
                if 'transaction_date' in ml_show_j.columns:
                    try:
                        ml_show_j['year'] = pd.to_datetime(ml_show_j['transaction_date'], errors='coerce').dt.year
                    except: pass
            ml_show_filtered = _render_branch_year_filter(ml_show_j, 'ej')

            # ── Эрсдэлийн metric cards (шүүлтүүрийн дараа) ──
            j_total = len(ml_show_filtered)
            j_anom = int(ml_show_filtered.get('ml_anomaly_flag', pd.Series(0)).sum()) if 'ml_anomaly_flag' in ml_show_filtered.columns else 0
            j_high = len(ml_show_filtered[ml_show_filtered.get('ml_risk_level','')=='🔴 Маш өндөр']) if 'ml_risk_level' in ml_show_filtered.columns else 0
            jm1, jm2, jm3, jm4 = st.columns(4)
            jm1.metric("📝 Нийт гүйлгээ", f"{j_total:,}")
            jm2.metric("⚠️ Аномали", f"{j_anom:,}", delta=f"{j_anom/max(j_total,1)*100:.1f}%", delta_color="inverse")
            jm3.metric("🔴 Маш өндөр", f"{j_high:,}")
            jm4.metric("🟢 Хэвийн", f"{j_total-j_anom:,}")

            # ── Tab-тай дэлгэрэнгүй ──
            j_viz_tabs = st.tabs(["📋 Гүйлгээ жагсаалт", "📊 Визуализаци", "🔢 Бенфорд", "🧠 XAI тайлбар", "📥 Тайлан"])

            with j_viz_tabs[0]:
                show_cols = [c for c in ['transaction_date','account_code','account_name','counterparty_name','amount','ml_vote_count','ml_risk_level','branch_label','xai_top_feature','transaction_description'] if c in ml_show_filtered.columns]
                sort_col = 'ml_vote_count' if 'ml_vote_count' in ml_show_filtered.columns else show_cols[0]
                st.dataframe(ml_show_filtered.sort_values(sort_col, ascending=False)[show_cols].head(1000), use_container_width=True, hide_index=True)
                _show_dataframe_download(ml_show_filtered, 'journal_ml_results.csv')

            with j_viz_tabs[1]:
                vc1, vc2 = st.columns(2)
                with vc1:
                    # Эрсдэлийн түвшний pie chart
                    if 'ml_risk_level' in ml_show_filtered.columns:
                        risk_vc = ml_show_filtered['ml_risk_level'].value_counts().reset_index()
                        risk_vc.columns = ['Түвшин', 'Тоо']
                        fig_rp = px.pie(risk_vc, values='Тоо', names='Түвшин', title='Эрсдэлийн түвшний хуваарилалт',
                                        color='Түвшин', hole=0.4,
                                        color_discrete_map={'🟢 Бага':'#4CAF50','🟠 Өндөр':'#FF9800','🔴 Маш өндөр':'#F44336'})
                        fig_rp.update_layout(height=380)
                        st.plotly_chart(fig_rp, use_container_width=True)
                with vc2:
                    # Гүйлгээний дүнгийн тархалт
                    if 'amount' in ml_show_filtered.columns:
                        fig_hist = px.histogram(ml_show_filtered[ml_show_filtered['amount']>0], x='amount', nbins=50,
                                                title='Гүйлгээний дүнгийн тархалт', log_y=True,
                                                color_discrete_sequence=['#0F4C81'])
                        fig_hist.update_layout(height=380, xaxis_title='Дүн (₮)', yaxis_title='Тоо (log)')
                        st.plotly_chart(fig_hist, use_container_width=True)
                # Timeline
                if 'transaction_date' in ml_show_filtered.columns:
                    tl = ml_show_filtered.copy()
                    tl['сар'] = tl['transaction_date'].astype(str).str[:7]
                    tl = tl[tl['сар'].str.len()>=7]
                    if not tl.empty and 'ml_anomaly_flag' in tl.columns:
                        tl_agg = tl.groupby('сар').agg(Нийт=('amount','count'), Аномали=('ml_anomaly_flag','sum')).reset_index().sort_values('сар')
                        fig_tl = make_subplots(specs=[[{"secondary_y":True}]])
                        fig_tl.add_trace(go.Bar(x=tl_agg['сар'], y=tl_agg['Нийт'], name='Нийт гүйлгээ', marker_color='#0F4C81', opacity=0.6), secondary_y=False)
                        fig_tl.add_trace(go.Scatter(x=tl_agg['сар'], y=tl_agg['Аномали'], name='Аномали тоо', mode='lines+markers', marker=dict(color='#D32F2F',size=6), line=dict(color='#D32F2F',width=2)), secondary_y=True)
                        fig_tl.update_layout(title='Сарын гүйлгээ ба аномали (ISA 240 хугацааны шинжилгээ)', height=400, legend=dict(orientation='h',y=1.05))
                        fig_tl.update_yaxes(title_text='Нийт гүйлгээ', secondary_y=False)
                        fig_tl.update_yaxes(title_text='Аномали тоо', secondary_y=True)
                        st.plotly_chart(fig_tl, use_container_width=True)

            with j_viz_tabs[2]:
                st.markdown("""
                <div style="background:linear-gradient(135deg,#E3F2FD,#E8EAF6);padding:16px 20px;border-radius:12px;border-left:4px solid #0F4C81;box-shadow:0 2px 8px rgba(15,76,129,0.08);margin-bottom:15px;">
                    <b>🔢 Бенфордын хууль (ISA 240):</b> Санхүүгийн гүйлгээний эхний оронгийн цифр нь
                    Бенфордын хуулийг дагах ёстой. Хэрэв ялгаатай бол залилангийн эрсдэл нэмэгдэнэ.
                </div>""", unsafe_allow_html=True)
                if 'benford_digit' in ml_show_filtered.columns:
                    bf_actual = ml_show_filtered[ml_show_filtered['benford_digit']>0]['benford_digit'].value_counts(normalize=True).sort_index()
                    bf_expected = pd.Series({1:0.301,2:0.176,3:0.125,4:0.097,5:0.079,6:0.067,7:0.058,8:0.051,9:0.046})
                    bf_df = pd.DataFrame({'Цифр': range(1,10), 'Бодит':bf_actual.reindex(range(1,10),fill_value=0).values, 'Бенфорд':bf_expected.values})
                    bf_melt = bf_df.melt(id_vars='Цифр', var_name='Төрөл', value_name='Хувь')
                    fig_bf = px.bar(bf_melt, x='Цифр', y='Хувь', color='Төрөл', barmode='group',
                                    title='Бенфордын хуулийн шинжилгээ (ISA 240)',
                                    color_discrete_map={'Бодит':'#E65100','Бенфорд':'#1565C0'})
                    fig_bf.update_layout(height=400, xaxis=dict(dtick=1))
                    st.plotly_chart(fig_bf, use_container_width=True)
                    # Хэлбийлтийн тест
                    bf_diff = (bf_actual.reindex(range(1,10),fill_value=0) - bf_expected).abs()
                    max_dev_digit = bf_diff.idxmax() if not bf_diff.empty else 0
                    max_dev_val = bf_diff.max() if not bf_diff.empty else 0
                    if max_dev_val > 0.03:
                        st.warning(f"⚠️ Цифр {max_dev_digit} дээр хамгийн их зөрүү: {max_dev_val:.3f} (ISA 240 анхааруулга)")
                    else:
                        st.success("✅ Бенфордын хуулиас мэдэгдэхүйц зөрүү илрээгүй")

            with j_viz_tabs[3]:
                if not xai_importance_j.empty:
                    st.markdown('### 🧠 XAI — Тайлбарлагдах хиймэл оюун ухаан')
                    st.markdown("""
                    <div style="background:linear-gradient(135deg,#F3E5F5,#EDE7F6);padding:16px 20px;border-radius:12px;border-left:4px solid #6A1B9A;box-shadow:0 2px 8px rgba(106,27,154,0.08);margin-bottom:15px;">
                        <b>🔍 Шинж чанарын ач холбогдол:</b> Доорх жагсаалт нь эрсдэлийг хамгийн их тайлбарлаж буй хүчин зүйлсийг харуулна.
                        Тус бүрд ОУАС (ISA) стандартын лавлагаа хавсаргасан.
                    </div>""", unsafe_allow_html=True)
                    xai_show = xai_importance_j.head(15).copy()
                    isa_map_full = {
                        'log_amount': 'ISA 240 — Гүйлгээний дүн', 'benford_dev': 'ISA 240 — Бенфордын зөрүү',
                        'is_round': 'ISA 240 — Тэгш дүнтэй гүйлгээ', 'amt_zscore': 'ISA 520 — Z-score хазайлт',
                        'cp_rare': 'ISA 550 — Ховор харилцагч', 'pair_rare': 'ISA 550 — Ховор данс×харилцагч',
                        'desc_empty': 'ISA 500 — Тайлбаргүй гүйлгээ', 'is_dup': 'ISA 240 — Давхардсан гүйлгээ',
                        'desc_mismatch': 'ISA 500 — Тайлбар зөрчил', 'dir_mismatch': 'ISA 240 — Чиглэлийн зөрчил',
                        'is_month_end': 'ISA 240 — Сарын эцэс', 'is_year_end': 'ISA 240 — Жилийн эцэс',
                        'acct_cat_num': 'ISA 315 — Дансны ангилал', 'is_debit': 'ISA 520 — Дебит/Кредит',
                        'name_no_overlap': 'ISA 500 — Нэр давхцахгүй',
                    }
                    xai_show['ISA лавлагаа'] = xai_show['feature'].map(isa_map_full).fillna('')
                    xai_show['importance'] = pd.to_numeric(xai_show['importance'], errors='coerce').fillna(0)
                    st.dataframe(xai_show[['feature','importance','ISA лавлагаа']], use_container_width=True, hide_index=True)
                    fig_xai = px.bar(xai_show.sort_values('importance',ascending=True), x='importance', y='feature',
                                     orientation='h', title='XAI: Шинж чанарын нөлөөллийн эрэмбэ',
                                     color='importance', color_continuous_scale='Blues')
                    fig_xai.update_layout(height=400, yaxis_title='', xaxis_title='Нөлөөллийн хэмжээ')
                    st.plotly_chart(fig_xai, use_container_width=True)

            with j_viz_tabs[4]:
                st.markdown("### 📥 ЕЖ аудитын тайлан татах")
                st.markdown("""
                <div style="background:linear-gradient(135deg,#E0F2F1,#E8F5E9);padding:16px 20px;border-radius:12px;border-left:4px solid #00897B;box-shadow:0 2px 8px rgba(0,137,123,0.08);margin-bottom:15px;">
                    <b>📋 Аудитын working paper:</b> Ерөнхий журналын бүх шинжилгээний үр дүн, аномали гүйлгээ,
                    Бенфордын шинжилгээ, XAI тайлбар нэг Excel файлд нэгтгэгдсэн.
                </div>""", unsafe_allow_html=True)
                j_report_buf = io.BytesIO()
                with pd.ExcelWriter(j_report_buf, engine='openpyxl') as jrw:
                    # Sheet 1: Бүх гүйлгээ
                    j_rpt_cols = [c for c in ['transaction_date','account_code','account_name','counterparty_name',
                                  'debit_mnt','credit_mnt','amount','ml_vote_count','ml_risk_level',
                                  'ml_anomaly_flag','xai_top_feature','transaction_description'] if c in ml_show_filtered.columns]
                    ml_show_filtered[j_rpt_cols].to_excel(jrw, sheet_name='Бүх_гүйлгээ', index=False)
                    # Sheet 2: Аномали гүйлгээ
                    if 'ml_anomaly_flag' in ml_show_filtered.columns:
                        anom_j_rpt = ml_show_filtered[ml_show_filtered['ml_anomaly_flag']==1][j_rpt_cols]
                        anom_j_rpt.to_excel(jrw, sheet_name='Аномали_гүйлгээ', index=False)
                    # Sheet 3: XAI
                    if not xai_importance_j.empty:
                        xai_importance_j.to_excel(jrw, sheet_name='XAI_feature_importance', index=False)
                    # Sheet 4: Алгоритмын тойм
                    if not model_summary_j.empty:
                        model_summary_j.to_excel(jrw, sheet_name='Алгоритмын_тойм', index=False)
                j_report_buf.seek(0)
                st.download_button('📥 ЕЖ Аудитын тайлан (Excel)', j_report_buf.getvalue(),
                                   file_name='audit_journal_report.xlsx',
                                   mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                   key='dl_journal_audit_report')
    else:
        st.info('👆 1️⃣ цэсэнд бэлтгэсэн ledger/ЕЖ файлаа ашиглах эсвэл энд шууд оруулж шинжилгээг эхлүүлнэ үү.')


def _pick_stats_source_df():
    """Статистик шинжилгээнд ашиглах хамгийн тохиромжтой эх өгөгдлийг сонгоно."""
    candidates = []

    tb_df = st.session_state.get('tb_all', pd.DataFrame())
    if tb_df is not None and not tb_df.empty:
        candidates.append(('Гүйлгээний балансын нэгтгэсэн өгөгдөл', tb_df))

    journal_df = st.session_state.get('journal_ml_result', pd.DataFrame())
    if journal_df is not None and not journal_df.empty:
        candidates.append(('Ерөнхий журналын машин сургалтын үр дүн', journal_df))

    prepared_ledger = _cache_files('prepared_ledger_cache')
    if prepared_ledger:
        try:
            _, ledger_sample = load_ledger_stats(prepared_ledger, sample_per_year=30000, chunksize=100000)
            if ledger_sample is not None and not ledger_sample.empty:
                candidates.append(('Ерөнхий журналын түүвэр өгөгдөл', ledger_sample))
        except Exception:
            pass

    if not candidates:
        return None, pd.DataFrame()

    labels = [name for name, _ in candidates]
    choice = st.selectbox('📚 Статистик шинжилгээний эх өгөгдөл', labels, key='stats_source_choice')
    for name, df in candidates:
        if name == choice:
            return name, df.copy()
    return candidates[0][0], candidates[0][1].copy()


def _prepare_numeric_stats_df(df):
    d = df.copy()
    if 'amount' not in d.columns and ('debit_mnt' in d.columns or 'credit_mnt' in d.columns):
        d['amount'] = pd.to_numeric(d.get('debit_mnt', 0), errors='coerce').fillna(0).abs() + pd.to_numeric(d.get('credit_mnt', 0), errors='coerce').fillna(0).abs()
    if 'month' not in d.columns and 'transaction_date' in d.columns:
        d['transaction_date'] = pd.to_datetime(d['transaction_date'], errors='coerce')
        d['month'] = d['transaction_date'].dt.to_period('M').astype(str)

    num = d.select_dtypes(include=[np.number]).copy()
    for c in d.columns:
        if c in num.columns:
            continue
        try:
            s = pd.to_numeric(d[c], errors='coerce')
            if s.notna().sum() > max(10, int(len(s) * 0.2)):
                num[c] = s
        except Exception:
            pass

    num = num.replace([np.inf, -np.inf], np.nan)
    num = num.dropna(axis=1, how='all')
    return d, num


def _safe_regression(y, X):
    try:
        X2 = X.replace([np.inf, -np.inf], np.nan).fillna(0)
        y2 = y.replace([np.inf, -np.inf], np.nan).fillna(0)
        coef = np.linalg.lstsq(X2.values, y2.values, rcond=None)[0]
        return pd.DataFrame({'Хувьсагч': X2.columns, 'Коэффициент': coef})
    except Exception:
        return pd.DataFrame(columns=['Хувьсагч', 'Коэффициент'])


def _run_stats_lab():
    st.header('Статистик шинжилгээ')  # unused
    st.markdown('Аудитад ашиглах статистикийн лаборатори — тодорхойлох статистик, тархалт, корреляци, регресс, хугацааны цуваа, PCA, outlier болон статистик тестүүдийг нэг дороос ажиллуулна.')

    source_name, source_df = _pick_stats_source_df()
    if source_df.empty:
        st.info('👆 Эхлээд 1️⃣ цэсэнд өгөгдөл бэлтгэх эсвэл 2️⃣/3️⃣ цэсэнд шинжилгээг ажиллуулж эх өгөгдлөө бүрдүүлнэ үү.')
        return

    raw_df, num = _prepare_numeric_stats_df(source_df)
    if num.empty:
        st.warning('Статистик шинжилгээ хийхэд хүрэлцэхүйц тоон өгөгдөл олдсонгүй.')
        return

    st.caption(f'Эх өгөгдөл: {source_name} • Нийт мөр: {len(raw_df):,} • Тоон багана: {len(num.columns)}')

    tabs = st.tabs([
        '📊 Тодорхойлох статистик',
        '📉 Тархалтын шинжилгээ',
        '🔗 Корреляци',
        '📈 Регрессийн загварууд',
        '⏱ Хугацааны цуваа',
        '🧠 Олон хувьсагчийн шинжилгээ',
        '🚨 Outlier илрүүлэлт',
        '🧪 Статистик тестүүд'
    ])

    with tabs[0]:
        st.markdown('### 📊 Тодорхойлох статистик')
        desc = pd.DataFrame({
            'Дундаж': num.mean(),
            'Медиан': num.median(),
            'Стандарт хазайлт': num.std(),
            'Дисперс': num.var(),
            'Хэлбийлт (Skewness)': num.skew(),
            'Оройлт (Kurtosis)': num.kurt(),
            'Хамгийн бага': num.min(),
            'Хамгийн их': num.max(),
            'NA бус мөрийн тоо': num.count()
        }).reset_index().rename(columns={'index': 'Хувьсагч'})
        st.dataframe(desc, use_container_width=True, hide_index=True)
        metric_col = st.selectbox('Тархалтын графикт харах хувьсагч', list(num.columns), key='stats_hist_col')
        fig = px.histogram(num, x=metric_col, nbins=50, title=f'{metric_col} тархалт')
        st.plotly_chart(fig, use_container_width=True)
        box = px.box(num, y=metric_col, points='outliers', title=f'{metric_col} boxplot')
        st.plotly_chart(box, use_container_width=True)
        _show_dataframe_download(desc, 'descriptive_statistics.csv', label='📥 Тодорхойлох статистикийг татах')

    with tabs[1]:
        st.markdown('### 📉 Тархалтын шинжилгээ')
        dist_col = st.selectbox('Тархалт шалгах хувьсагч', list(num.columns), key='dist_col')
        s = num[dist_col].dropna()
        fig = px.histogram(s.to_frame(name=dist_col), x=dist_col, marginal='violin', nbins=40, title=f'{dist_col} тархалтын дүрслэл')
        st.plotly_chart(fig, use_container_width=True)
        try:
            shapiro_stat, shapiro_p = stats.shapiro(s.sample(min(len(s), 5000), random_state=42))
        except Exception:
            shapiro_stat, shapiro_p = np.nan, np.nan
        try:
            denom = s.std() if s.std() else 1
            ks_stat, ks_p = stats.kstest((s - s.mean()) / denom, 'norm')
        except Exception:
            ks_stat, ks_p = np.nan, np.nan
        st.dataframe(pd.DataFrame([
            {'Тест': 'Shapiro-Wilk', 'Statistic': shapiro_stat, 'p-value': shapiro_p},
            {'Тест': 'Kolmogorov-Smirnov', 'Statistic': ks_stat, 'p-value': ks_p},
        ]), use_container_width=True, hide_index=True)

        if 'amount' in raw_df.columns:
            ben = raw_df['amount'].dropna()
            ben = ben[ben > 0]
            if not ben.empty:
                first_digits = ben.astype(int).astype(str).str[0]
                obs = first_digits.value_counts(normalize=True).sort_index()
                benford_exp = pd.Series({1: 0.301, 2: 0.176, 3: 0.125, 4: 0.097, 5: 0.079, 6: 0.067, 7: 0.058, 8: 0.051, 9: 0.046})
                ben_df = pd.DataFrame({
                    'digit': benford_exp.index,
                    'Ажиглагдсан хувь': [obs.get(str(i), 0) for i in benford_exp.index],
                    'Бенфордын хүлээгдэж буй хувь': benford_exp.values
                })
                figb = px.bar(ben_df.melt(id_vars='digit', var_name='Төрөл', value_name='Хувь'), x='digit', y='Хувь', color='Төрөл', barmode='group', title='Бенфордын хуулийн харьцуулалт')
                st.plotly_chart(figb, use_container_width=True)

    with tabs[2]:
        st.markdown('### 🔗 Корреляцийн шинжилгээ')
        method = st.selectbox('Корреляцийн арга', ['pearson', 'spearman', 'kendall'], key='corr_method')
        corr = num.corr(method=method)
        st.dataframe(corr, use_container_width=True)
        figc = px.imshow(corr, text_auto='.2f', title=f'{method.title()} корреляцийн матриц', aspect='auto')
        st.plotly_chart(figc, use_container_width=True)

    with tabs[3]:
        st.markdown('### 📈 Регрессийн загварууд')
        target = st.selectbox('Хамаарах хувьсагч (Y)', list(num.columns), key='reg_target')
        feature_options = [c for c in num.columns if c != target]
        default_features = feature_options[:min(5, len(feature_options))]
        features = st.multiselect('Тайлбарлагч хувьсагчид (X)', feature_options, default=default_features, key='reg_features')
        if features:
            y = num[target].fillna(0)
            X = num[features].fillna(0)
            reg_tabs = st.tabs(['Шугаман регресс', 'Ridge', 'Lasso', 'Elastic Net', 'Логистик регресс', 'Гүйцэтгэлийн самбар'])
            with reg_tabs[0]:
                coef_df = _safe_regression(y, X)
                st.dataframe(coef_df, use_container_width=True, hide_index=True)
            with reg_tabs[1]:
                mdl = Ridge(alpha=1.0).fit(X, y)
                st.dataframe(pd.DataFrame({'Хувьсагч': features, 'Коэффициент': mdl.coef_}), use_container_width=True, hide_index=True)
            with reg_tabs[2]:
                mdl = Lasso(alpha=max(float(np.std(y)) * 0.001, 0.001), max_iter=5000).fit(X, y)
                st.dataframe(pd.DataFrame({'Хувьсагч': features, 'Коэффициент': mdl.coef_}), use_container_width=True, hide_index=True)
            with reg_tabs[3]:
                mdl = ElasticNet(alpha=max(float(np.std(y)) * 0.001, 0.001), l1_ratio=0.5, max_iter=5000).fit(X, y)
                st.dataframe(pd.DataFrame({'Хувьсагч': features, 'Коэффициент': mdl.coef_}), use_container_width=True, hide_index=True)
            with reg_tabs[4]:
                y_bin = (y > y.median()).astype(int)
                try:
                    mdl = LogisticRegression(max_iter=2000).fit(X, y_bin)
                    st.dataframe(pd.DataFrame({'Хувьсагч': features, 'Коэффициент': mdl.coef_[0]}), use_container_width=True, hide_index=True)
                except Exception as e:
                    st.warning(f'Логистик регресс ажиллуулж чадсангүй: {e}')
            with reg_tabs[5]:
                lin = LinearRegression().fit(X, y)
                pred = lin.predict(X)
                r2 = lin.score(X, y)
                m1, m2, m3 = st.columns(3)
                m1.metric('R²', f'{r2:.4f}')
                m2.metric('Мөрийн тоо', f'{len(X):,}')
                m3.metric('Хувьсагчийн тоо', f'{len(features)}')
                plot_df = pd.DataFrame({'Бодит утга': y, 'Таамагласан утга': pred})
                figr = px.scatter(plot_df.sample(min(len(plot_df), 3000), random_state=42), x='Бодит утга', y='Таамагласан утга', title='Бодит ба таамагласан утгын харьцуулалт')
                st.plotly_chart(figr, use_container_width=True)
        else:
            st.info('Регресс ажиллуулахын тулд дор хаяж нэг тайлбарлагч хувьсагч сонгоно уу.')

    with tabs[4]:
        st.markdown('### ⏱ Хугацааны цувааны шинжилгээ')
        time_col = None
        if 'transaction_date' in raw_df.columns:
            time_col = 'transaction_date'
            raw_df[time_col] = pd.to_datetime(raw_df[time_col], errors='coerce')
        value_col = st.selectbox('Хугацааны цуваанд харах үзүүлэлт', list(num.columns), key='ts_value_col')
        if time_col and raw_df[time_col].notna().sum() > 10:
            freq = st.selectbox('Нэгтгэх давтамж', ['D', 'M', 'Q'], index=1, key='ts_freq')
            ts = raw_df[[time_col]].copy()
            ts[value_col] = pd.to_numeric(raw_df[value_col], errors='coerce').fillna(0)
            ts = ts.dropna(subset=[time_col]).set_index(time_col).resample(freq).sum(numeric_only=True)
            ts = ts[[value_col]].dropna()
            figt = px.line(ts.reset_index(), x=time_col, y=value_col, title=f'{value_col} хугацааны цуваа')
            st.plotly_chart(figt, use_container_width=True)
            ts['3-периодын хөдөлгөөнт дундаж'] = ts[value_col].rolling(3).mean()
            figma = px.line(ts.reset_index(), x=time_col, y=[value_col, '3-периодын хөдөлгөөнт дундаж'], title='Хөдөлгөөнт дундаж')
            st.plotly_chart(figma, use_container_width=True)
            if seasonal_decompose is not None and len(ts) >= 12:
                try:
                    decomp = seasonal_decompose(ts[value_col], model='additive', period=max(2, min(12, len(ts) // 2)))
                    decomp_df = pd.DataFrame({
                        time_col: ts.index,
                        'Тренд': decomp.trend,
                        'Улирлын нөлөө': decomp.seasonal,
                        'Үлдэгдэл': decomp.resid
                    }).dropna()
                    figd = px.line(decomp_df, x=time_col, y=['Тренд', 'Улирлын нөлөө', 'Үлдэгдэл'], title='Улирлын задлал')
                    st.plotly_chart(figd, use_container_width=True)
                except Exception as e:
                    st.info(f'Улирлын задлалыг ажиллуулж чадсангүй: {e}')
        else:
            st.info('Огнооны багана хангалтгүй тул хугацааны цувааны шинжилгээ хийгдсэнгүй.')

    with tabs[5]:
        st.markdown('### 🧠 Олон хувьсагчийн шинжилгээ')
        xcols = st.multiselect('PCA-д оруулах хувьсагчид', list(num.columns), default=list(num.columns[:min(8, len(num.columns))]), key='pca_cols')
        if len(xcols) >= 2:
            X = num[xcols].fillna(0)
            scaler = StandardScaler()
            Xs = scaler.fit_transform(X)
            pca = PCA(n_components=2)
            comps = pca.fit_transform(Xs)
            pca_df = pd.DataFrame({'PC1': comps[:, 0], 'PC2': comps[:, 1]})
            figp = px.scatter(pca_df.sample(min(len(pca_df), 5000), random_state=42), x='PC1', y='PC2', title='PCA 2D дүрслэл')
            st.plotly_chart(figp, use_container_width=True)
            st.dataframe(pd.DataFrame({
                'Үндсэн бүрэлдэхүүн': ['PC1', 'PC2'],
                'Тайлбарлах дисперсийн хувь': pca.explained_variance_ratio_
            }), use_container_width=True, hide_index=True)
            km_n = st.slider('KMeans кластерын тоо', 2, 12, 4, key='stats_kmeans_n')
            km = KMeans(n_clusters=km_n, n_init=10, random_state=42).fit(Xs)
            pca_df['Кластер'] = km.labels_.astype(str)
            figk = px.scatter(pca_df.sample(min(len(pca_df), 5000), random_state=42), x='PC1', y='PC2', color='Кластер', title='PCA + KMeans кластерчлал')
            st.plotly_chart(figk, use_container_width=True)
        else:
            st.info('PCA хийхэд дор хаяж хоёр хувьсагч сонгоно уу.')

    with tabs[6]:
        st.markdown('### 🚨 Outlier илрүүлэлт')
        out_col = st.selectbox('Outlier шалгах хувьсагч', list(num.columns), key='outlier_col')
        s = num[out_col].fillna(0)
        z = ((s - s.mean()) / (s.std() if s.std() else 1)).abs()
        q1, q3 = s.quantile([0.25, 0.75])
        iqr = q3 - q1
        iqr_flag = (s < q1 - 1.5 * iqr) | (s > q3 + 1.5 * iqr)
        out_df = pd.DataFrame({
            'Утга': s,
            'Z-score абсолют утга': z,
            'Z > 3': z > 3,
            'IQR outlier': iqr_flag
        })
        if len(num.columns) >= 2:
            sample = num.sample(min(len(num), 10000), random_state=42).fillna(0)
            try:
                iso = IsolationForest(contamination=0.05, random_state=42)
                out_df.loc[sample.index, 'Isolation Forest'] = (iso.fit_predict(sample) == -1)
            except Exception:
                pass
        st.dataframe(out_df.head(2000), use_container_width=True)
        figo = px.scatter(out_df.reset_index().head(5000), x='index', y='Утга', color='Z > 3', title=f'{out_col} хувьсагчийн outlier дүрслэл')
        st.plotly_chart(figo, use_container_width=True)

    with tabs[7]:
        st.markdown('### 🧪 Статистик тестүүд')
        test_col = st.selectbox('Тест хийх үндсэн хувьсагч', list(num.columns), key='test_col')
        s = num[test_col].dropna()
        group_col = None
        if 'year' in raw_df.columns:
            group_col = 'year'
        elif 'report_year' in raw_df.columns:
            group_col = 'report_year'
        results = []
        try:
            tstat, tp = stats.ttest_1samp(s.sample(min(len(s), 5000), random_state=42), popmean=s.median())
            results.append({'Тест': 'One-sample t-test', 'Statistic': tstat, 'p-value': tp})
        except Exception:
            pass
        if group_col and raw_df[group_col].nunique() >= 2:
            try:
                groups = []
                for g in raw_df[group_col].dropna().unique()[:3]:
                    vals = pd.to_numeric(raw_df.loc[raw_df[group_col] == g, test_col], errors='coerce').dropna()
                    if len(vals) > 3:
                        groups.append(vals.sample(min(len(vals), 5000), random_state=42))
                if len(groups) >= 2:
                    fstat, fp = stats.f_oneway(*groups)
                    results.append({'Тест': f'ANOVA by {group_col}', 'Statistic': fstat, 'p-value': fp})
            except Exception:
                pass
        try:
            chi_bins = pd.qcut(s.rank(method='first'), q=min(5, max(2, s.nunique())), duplicates='drop')
            freq = chi_bins.value_counts().sort_index()
            chi_stat, chi_p = stats.chisquare(freq.values)
            results.append({'Тест': 'Chi-square (binned uniformity)', 'Statistic': chi_stat, 'p-value': chi_p})
        except Exception:
            pass
        if results:
            st.dataframe(pd.DataFrame(results), use_container_width=True, hide_index=True)
        else:
            st.info('Статистик тест ажиллуулахад хүрэлцэхүйц нөхцөл бүрдээгүй байна.')


if page.startswith("4"):
    st.header("4️⃣ Материаллаг байдлын тооцоо")
    st.markdown("Гүйлгээ балансын файл дээр суурилсан материаллаг байдлын хуваарилалт.")

    mat_files = st.file_uploader("📎 Нэмэлт гүйлгээний баланс оруулах", type=['xlsx','xls','xlsm','xlsb','csv'], accept_multiple_files=True, key='mat_files_work')
    tb_inputs = _cache_files('prepared_tb_cache')
    for f in mat_files or []:
        ftype, year = detect_file_type(f)
        f.seek(0)
        if ftype == 'tb_std':
            tb_inputs.append(f)
        elif ftype == 'raw_tb':
            buf, tb_sum = process_raw_tb(f)
            if tb_sum is not None and not tb_sum.empty:
                bio = io.BytesIO(buf.getvalue()); bio.name = f'TB_standardized_{year}_{Path(f.name).stem}.xlsx'
                tb_inputs.append(bio)

    total_mat = st.number_input('Нийт материаллаг байдлын дүн', min_value=0.0, value=1000000.0, step=100000.0)
    if st.button('📐 Материаллаг байдлыг тооцоолох', type='primary', use_container_width=True):
        tb_all, _ = load_tb(tb_inputs) if tb_inputs else (pd.DataFrame(), {})
        if tb_all.empty:
            st.warning('Гүйлгээний балансын өгөгдөл олдсонгүй.')
        else:
            d = tb_all.copy()
            base_amt = d['closing_debit'].abs() + d['closing_credit'].abs() if 'closing_debit' in d.columns else d['turnover_debit'].abs() + d['turnover_credit'].abs()
            total_base = base_amt.sum()
            d['materiality_alloc'] = np.where(total_base > 0, total_mat * base_amt / total_base, 0)
            show_cols = [c for c in ['year','account_code','account_name','closing_debit','closing_credit','turnover_debit','turnover_credit','materiality_alloc'] if c in d.columns]
            st.session_state['materiality_result'] = d[show_cols].sort_values('materiality_alloc', ascending=False)
            st.session_state['materiality_total'] = total_mat

    # ── Кэшлэгдсэн үр дүн харуулах ──
    if st.session_state.get('materiality_result') is not None and not st.session_state.get('materiality_result', pd.DataFrame()).empty:
        mat_df = st.session_state['materiality_result']
        mat_total = st.session_state.get('materiality_total', 0)
        st.success(f"✅ Материаллаг байдал тооцоологдсон: ₮{mat_total:,.0f} → {len(mat_df):,} данс")
        st.dataframe(mat_df, use_container_width=True, hide_index=True)
        _show_dataframe_download(mat_df, 'materiality_by_account.csv')


# ═══════════════════════════════════════════════════════════
# 6️⃣ САЛБАРЫН ШИНЖИЛГЭЭ
# ═══════════════════════════════════════════════════════════
if page.startswith("5"):
    st.header("5️⃣ Салбар тус бүрийн шинжилгээ")
    st.markdown("Файлуудыг автоматаар салбараар ангилж, тус бүрд тусдаа эрсдэлийн шинжилгээ хийнэ.")

    branch_files = st.file_uploader("📎 Бүх ЕЖ файлуудаа энд оруулна уу", type=['xlsx','xls','xlsm','xlsb','csv','tsv','gz'], accept_multiple_files=True, key='branch_upload')

    # Мөн бэлтгэсэн ledger кэшээс авах
    prepared_led = _cache_files('prepared_ledger_cache')

    if not branch_files and not prepared_led:
        st.info("👆 ЕЖ/ledger файлуудаа оруулна уу, эсвэл 1️⃣ цэсэнд бэлтгэсэн файлууд автоматаар ачаалагдна.")

    branch_cont = st.slider("🎯 Аномалийн хувь (contamination)", 0.01, 0.20, 0.05, 0.01, key='branch_cont')
    MAX_BRANCH_ROWS = 15000

    if st.button("🏢 Салбарын шинжилгээ эхлүүлэх", type='primary', use_container_width=True, key='run_branch'):
        all_inputs = []
        # Бэлтгэсэн ledger
        for f in prepared_led:
            all_inputs.append(f)
        # Шинэ оруулсан
        for f in branch_files or []:
            all_inputs.append(f)

        if not all_inputs:
            st.warning("Файл оруулна уу.")
        else:
            MAX_BRANCH_ROWS = 8000  # Салбар бүрт хамгийн ихдээ 8,000 мөр
            progress = st.progress(0, text="Салбарын шинжилгээ эхэлж байна...")
            branch_data = {}
            detect_info = []
            total = len(all_inputs)

            for i, f in enumerate(all_inputs):
                fname = getattr(f, 'name', f'file_{i}')
                progress.progress((i+1)/total * 0.4, text=f"Файл уншиж байна: {fname}")

                # Салбар таних
                branch_id, branch_label = detect_branch(fname)

                # Файл уншиж ledger болгох
                ftype, year = detect_file_type(f)
                f.seek(0)
                detect_info.append({'Файл': fname, 'Салбар': branch_label, 'Төрөл': ftype, 'Он': year})

                try:
                    if ftype == 'edt':
                        f.seek(0)
                        edt_df, cnt = process_edt(f, year)
                        if cnt > 0 and not edt_df.empty:
                            if len(edt_df) > MAX_BRANCH_ROWS:
                                edt_df = edt_df.sample(n=MAX_BRANCH_ROWS, random_state=42).reset_index(drop=True)
                            edt_df['branch_id'] = branch_id
                            edt_df['branch_label'] = branch_label
                            if branch_id not in branch_data:
                                branch_data[branch_id] = []
                            branch_data[branch_id].append(edt_df)
                    elif ftype == 'ledger':
                        f.seek(0)
                        led_df = read_ledger(f)
                        if not led_df.empty:
                            if len(led_df) > MAX_BRANCH_ROWS:
                                led_df = led_df.sample(n=MAX_BRANCH_ROWS, random_state=42).reset_index(drop=True)
                            led_df['branch_id'] = branch_id
                            led_df['branch_label'] = branch_label
                            if branch_id not in branch_data:
                                branch_data[branch_id] = []
                            branch_data[branch_id].append(led_df)
                except Exception as e:
                    st.warning(f"⚠️ {fname}: {e}")

            # Салбар бүрийн өгөгдлийг нэгтгэх
            progress.progress(0.45, text="Салбаруудыг нэгтгэж байна...")
            branch_dfs = {}
            for bid, frames in branch_data.items():
                merged = pd.concat(frames, ignore_index=True)
                original_len = len(merged)
                if len(merged) > MAX_BRANCH_ROWS:
                    merged = merged.sample(n=MAX_BRANCH_ROWS, random_state=42).reset_index(drop=True)
                    st.caption(f"⚡ {bid}: {original_len:,} → {MAX_BRANCH_ROWS:,} мөр sample")
                branch_dfs[bid] = merged

            # Файл таних хүснэгт
            st.markdown("### 📋 Файлын салбарын таних үр дүн")
            st.dataframe(pd.DataFrame(detect_info), use_container_width=True, hide_index=True)

            if not branch_dfs:
                st.warning("Салбарын өгөгдөл олдсонгүй.")
            else:
                # Салбарын нэгтгэл
                st.markdown("### 🏢 Салбарын нэгтгэл")
                summary = get_branch_summary(branch_dfs)
                st.dataframe(summary, use_container_width=True, hide_index=True)

                # Салбарын хэмжээний бар чарт
                if not summary.empty and 'Нийт мөр' in summary.columns:
                    fig_bar = px.bar(summary, x='Салбар', y='Нийт мөр', color='Салбар',
                                     title='Салбар бүрийн гүйлгээний тоо')
                    st.plotly_chart(fig_bar, use_container_width=True)

                # Feature engineering + аномали (салбар тус бүрт progress)
                progress.progress(0.5, text="Салбар тус бүрийн шинж чанар бүтээж байна...")
                feat_cols = ['log_amount','acct_cat_num','benford_dev','is_round','amt_zscore',
                            'cp_rare','pair_rare','desc_empty','is_month_end','is_year_end',
                            'is_dup','is_debit','desc_mismatch','name_no_overlap','dir_mismatch']

                branch_engineered = {}
                total_branches = len(branch_dfs)
                for idx_b, (bid, df) in enumerate(branch_dfs.items()):
                    pct = 0.5 + (idx_b + 1) / total_branches * 0.4
                    short_name = BRANCH_REGISTRY.get(bid, {'short': bid}).get('short', bid)
                    progress.progress(min(pct, 0.9), text=f"Шинж чанар бүтээж байна: {short_name} ({idx_b+1}/{total_branches})")
                    df_clean = clean_for_risk(df)
                    if len(df_clean) > 5:
                        df_feat = engineer_txn_features(df_clean)
                        branch_engineered[bid] = df_feat

                # Салбар хоорондын харьцуулалт
                if branch_engineered:
                    st.markdown("### 📊 Салбар хоорондын эрсдэлийн харьцуулалт")
                    comparison_df, branch_anomalies = run_branch_comparison(
                        branch_engineered, feat_cols, contamination=branch_cont
                    )
                    if not comparison_df.empty:
                        st.dataframe(comparison_df.sort_values('Аномали %', ascending=False),
                                     use_container_width=True, hide_index=True)

                        # Аномали хувийн харьцуулалт
                        fig_comp = px.bar(comparison_df, x='Салбар', y='Аномали %', color='Салбар',
                                          title='Салбар бүрийн аномали хувь (%)',
                                          text='Аномали %')
                        fig_comp.update_traces(textposition='outside')
                        st.plotly_chart(fig_comp, use_container_width=True)

                        # Heatmap: салбар × vote
                        fig_heat = px.bar(comparison_df, x='Салбар', y=['Аномали тоо', 'Нийт мөр'],
                                          barmode='group', title='Аномали тоо vs Нийт мөр')
                        st.plotly_chart(fig_heat, use_container_width=True)

                    # Салбар тус бүрийн дэлгэрэнгүй
                    st.markdown("### 🔍 Салбар тус бүрийн дэлгэрэнгүй")
                    # Салбар тус бүрт expander (React tree тогтвортой)

                    for bid, df_feat in branch_engineered.items():
                        with st.expander(f"🏢 {BRANCH_REGISTRY.get(bid, {'short':bid}).get('short', bid)}", expanded=False):
                            info = BRANCH_REGISTRY.get(bid, {'label':bid, 'short':bid})
                            st.markdown(f"#### {info['label']}")
                            st.caption(f"Нийт: {len(df_feat):,} мөр • Данс: {df_feat['account_code'].nunique()} • ")

                            # Энэ салбарын аномали
                            labels, votes = create_pseudo_labels(df_feat, feat_cols, branch_cont)
                            df_feat = df_feat.copy()
                            df_feat['anomaly'] = labels
                            df_feat['votes'] = votes

                            c1, c2, c3 = st.columns(3)
                            c1.metric("Нийт гүйлгээ", f"{len(df_feat):,}")
                            c2.metric("Аномали тоо", f"{int(labels.sum()):,}")
                            c3.metric("Аномали %", f"{labels.mean()*100:.1f}%")

                            # Top аномали гүйлгээ
                            anom_df = df_feat[df_feat['anomaly']==1].sort_values('votes', ascending=False)
                            show_cols = [c for c in ['transaction_date','account_code','account_name',
                                         'counterparty_name','amount','votes','transaction_description'] if c in anom_df.columns]
                            st.dataframe(anom_df[show_cols].head(200), use_container_width=True, hide_index=True)
                            _show_dataframe_download(anom_df, f'branch_{bid}_anomalies.csv')

                progress.progress(1.0, text="Салбарын шинжилгээ дууслаа!")
                st.session_state['branch_engineered'] = branch_engineered
                st.session_state['branch_dfs'] = branch_dfs
                st.session_state['branch_comparison'] = comparison_df if not comparison_df.empty else pd.DataFrame()
                st.session_state['branch_summary'] = summary if not summary.empty else pd.DataFrame()
                st.session_state['branch_detect_info'] = detect_info
                st.session_state['branch_done'] = True
                st.success("✅ Салбарын шинжилгээ дууслаа.")

    # ═══ КЭШЛЭГДСЭН ҮР ДҮН (цэс хооронд шилжихэд) ═══
    elif st.session_state.get('branch_done'):
        st.info("📦 Өмнөх салбарын шинжилгээний үр дүн:")
        _det = st.session_state.get('branch_detect_info', [])
        if _det:
            st.markdown("### 📋 Файлын салбарын таних үр дүн")
            st.dataframe(pd.DataFrame(_det), use_container_width=True, hide_index=True)
        _bsum = st.session_state.get('branch_summary', pd.DataFrame())
        if not _bsum.empty:
            st.markdown("### 🏢 Салбарын нэгтгэл")
            st.dataframe(_bsum, use_container_width=True, hide_index=True)
            if 'Нийт мөр' in _bsum.columns:
                fig_bar = px.bar(_bsum, x='Салбар', y='Нийт мөр', color='Салбар', title='Салбар бүрийн гүйлгээний тоо')
                st.plotly_chart(fig_bar, use_container_width=True)
        _bcomp = st.session_state.get('branch_comparison', pd.DataFrame())
        if not _bcomp.empty:
            st.markdown("### 📊 Салбар хоорондын эрсдэлийн харьцуулалт")
            st.dataframe(_bcomp.sort_values('Аномали %', ascending=False), use_container_width=True, hide_index=True)
            fig_comp = px.bar(_bcomp, x='Салбар', y='Аномали %', color='Салбар', title='Салбар бүрийн аномали хувь (%)', text='Аномали %')
            fig_comp.update_traces(textposition='outside')
            st.plotly_chart(fig_comp, use_container_width=True)
        _beng = st.session_state.get('branch_engineered', {})
        if _beng:
            st.markdown("### 🔍 Салбар тус бүрийн дэлгэрэнгүй")
            for bid, df_feat in _beng.items():
                with st.expander(f"🏢 {BRANCH_REGISTRY.get(bid, {'short':bid}).get('short', bid)}", expanded=False):
                    info = BRANCH_REGISTRY.get(bid, {'label':bid, 'short':bid})
                    st.markdown(f"#### {info['label']}")
                    st.caption(f"Нийт: {len(df_feat):,} мөр • Данс: {df_feat['account_code'].nunique()}")
                    if 'anomaly' in df_feat.columns:
                        c1, c2, c3 = st.columns(3)
                        c1.metric("Нийт гүйлгээ", f"{len(df_feat):,}")
                        c2.metric("Аномали тоо", f"{int(df_feat['anomaly'].sum()):,}")
                        c3.metric("Аномали %", f"{df_feat['anomaly'].mean()*100:.1f}%")
                        anom_df = df_feat[df_feat['anomaly']==1].sort_values('votes', ascending=False) if 'votes' in df_feat.columns else df_feat[df_feat['anomaly']==1]
                        show_cols = [c for c in ['transaction_date','account_code','account_name','counterparty_name','amount','votes','transaction_description'] if c in anom_df.columns]
                        st.dataframe(anom_df[show_cols].head(200), use_container_width=True, hide_index=True)
                        _show_dataframe_download(anom_df, f'branch_{bid}_anomalies.csv')
                    else:
                        st.caption("Аномали шинжилгээ хийгдээгүй. Дахин ажиллуулна уу.")


# ═══════════════════════════════════════════════════════════
# 7️⃣ TRAIN/TEST ML PIPELINE
# ═══════════════════════════════════════════════════════════
if page.startswith("6"):
    st.header("6️⃣ Сургалт/Шалгалт — Машин сургалтын дамжуулалт")
    st.markdown("""
    Диссертацийн шаардлагад нийцсэн бүрэн машин сургалтын дамжуулалт:
    **Pseudo-labeling → Stratified Split → Supervised Models → Cross-validation → Гүйцэтгэлийн үнэлгээ**
    """)

    # Эх өгөгдөл сонгох
    source_options = []
    if st.session_state.get('branch_engineered'):
        source_options.append('Салбарын шинжилгээний өгөгдөл (5️⃣ цэсээс)')
    if st.session_state.get('journal_ml_result') is not None and not st.session_state.get('journal_ml_result', pd.DataFrame()).empty:
        source_options.append('Ерөнхий журналын машин сургалтын өгөгдөл (3️⃣ цэсээс)')
    source_options.append('Шинэ файл оруулах')

    source = st.selectbox("📚 Эх өгөгдөл сонгох", source_options, key='ml7_source')

    # Параметрүүд
    st.markdown("### ⚙️ Машин сургалтын параметрүүд")
    cp1, cp2, cp3 = st.columns(3)
    with cp1:
        ml_cont = st.slider("Аномалийн хувь", 0.01, 0.20, 0.05, 0.01, key='ml7_cont')
    with cp2:
        ml_test_size = st.slider("Тестийн хувь", 0.10, 0.40, 0.20, 0.05, key='ml7_test')
    with cp3:
        ml_cv_folds = st.slider("CV fold тоо", 2, 10, 5, 1, key='ml7_cv')

    feat_cols = ['log_amount','acct_cat_num','benford_dev','is_round','amt_zscore',
                'cp_rare','pair_rare','desc_empty','is_month_end','is_year_end',
                'is_dup','is_debit','desc_mismatch','name_no_overlap','dir_mismatch']

    # Салбар сонголт (салбарын өгөгдөл ашиглах бол)
    selected_branch = None
    ml_df = pd.DataFrame()

    if source.startswith('Салбарын'):
        branch_eng = st.session_state.get('branch_engineered', {})
        if branch_eng:
            branch_choices = ['🌐 Бүх салбар нэгтгэсэн'] + [
                BRANCH_REGISTRY.get(bid, {'label':bid})['label'] for bid in branch_eng.keys()
            ]
            branch_choice = st.selectbox("🏢 Салбар сонгох", branch_choices, key='ml7_branch')

            if branch_choice == '🌐 Бүх салбар нэгтгэсэн':
                ml_df = pd.concat(list(branch_eng.values()), ignore_index=True)
            else:
                for bid, df in branch_eng.items():
                    info = BRANCH_REGISTRY.get(bid, {'label': bid})
                    if info['label'] == branch_choice:
                        ml_df = df.copy()
                        selected_branch = bid
                        break
    elif source.startswith('Ерөнхий'):
        ml_df = st.session_state.get('journal_ml_result', pd.DataFrame()).copy()
        if not ml_df.empty:
            # Feature-үүд байгаа эсэхийг шалгах
            missing_feats = [c for c in feat_cols if c not in ml_df.columns]
            if missing_feats:
                ml_df = engineer_txn_features(clean_for_risk(ml_df))
    else:
        ml_files = st.file_uploader("📎 ЕЖ/ledger файл оруулна уу", type=['xlsx','xls','xlsm','xlsb','csv','tsv','gz'],
                                     accept_multiple_files=True, key='ml7_files')
        if ml_files:
            frames = []
            for f in ml_files:
                ftype, year = detect_file_type(f)
                f.seek(0)
                if ftype == 'edt':
                    edt_df, cnt = process_edt(f, year)
                    if cnt > 0: frames.append(edt_df)
                elif ftype == 'ledger':
                    frames.append(read_ledger(f))
            if frames:
                ml_df = pd.concat(frames, ignore_index=True)
                ml_df = engineer_txn_features(clean_for_risk(ml_df))

    if not ml_df.empty:
        MAX_ML_ROWS = 80000
        if len(ml_df) > MAX_ML_ROWS:
            st.info(f"📊 Өгөгдөл бэлэн: **{len(ml_df):,}** мөр → хурдасгахын тулд **{MAX_ML_ROWS:,}** мөр sample авна")
        else:
            st.success(f"📊 Өгөгдөл бэлэн: **{len(ml_df):,}** мөр, **{ml_df['account_code'].nunique() if 'account_code' in ml_df.columns else 0}** данс")

    if st.button("🚀 Сургалт/Шалгалт машин сургалт ажиллуулах", type='primary', use_container_width=True, key='run_ml7'):
        if ml_df.empty or len(ml_df) < 30:
            st.warning("Хангалтгүй өгөгдөл. Дор хаяж 30 мөр шаардлагатай.")
        else:
            with st.spinner("Машин сургалтын дамжуулалт ажиллаж байна..."):
                progress7 = st.progress(0, text="Өгөгдөл бэлтгэж байна...")

                # ── Том өгөгдлийг sample авч хурдасгах ──
                ml_work = ml_df.copy()
                if len(ml_work) > MAX_ML_ROWS:
                    progress7.progress(0.05, text=f"Том өгөгдөл: {len(ml_work):,} → {MAX_ML_ROWS:,} мөр sample авч байна...")
                    ml_work = ml_work.sample(n=MAX_ML_ROWS, random_state=42).reset_index(drop=True)
                    st.info(f"⚡ {len(ml_df):,} мөрөөс {MAX_ML_ROWS:,} мөр sample авлаа (хурдасгах зорилгоор)")

                progress7.progress(0.1, text="Pseudo-label үүсгэж байна...")

                # Feature шалгах
                for c in feat_cols:
                    if c not in ml_work.columns:
                        ml_work[c] = 0

                progress7.progress(0.2, text="Сургалт/Шалгалт хуваалт хийж байна...")
                results = run_train_test_ml(ml_work, feat_cols, test_size=ml_test_size,
                                            contamination=ml_cont, random_state=42)

                if not results['success']:
                    st.error(f"❌ {results['error']}")
                else:
                    progress7.progress(0.6, text="Unsupervised vs Supervised харьцуулж байна...")

                    # Unsupervised харьцуулалт (sample-аас)
                    unsup_results = compare_unsupervised_supervised(ml_work, feat_cols, ml_cont)

                    progress7.progress(1.0, text="Машин сургалтын дамжуулалт дууслаа!")
                    st.success("✅ Сургалт/Шалгалт амжилттай дууслаа!")

                    # ═══ ҮР ДҮН ХАРУУЛАХ ═══

                    # Pseudo-label статистик
                    st.markdown("### 📊 Pseudo-label статистик")
                    ps = results['pseudo_label_stats']
                    cc1, cc2, cc3, cc4 = st.columns(4)
                    cc1.metric("Нийт мөр", f"{ps['total']:,}")
                    cc2.metric("Аномали", f"{ps['anomaly']:,}")
                    cc3.metric("Хэвийн", f"{ps['normal']:,}")
                    cc4.metric("Аномали %", f"{ps['anomaly_pct']:.1f}%")

                    trc, tec = st.columns(2)
                    trc.metric("Сургалтын мөр", f"{len(results['train_df']):,}")
                    tec.metric("Тестийн мөр", f"{len(results['test_df']):,}")

                    # Загварын гүйцэтгэл (TEST)
                    st.markdown("### 🎯 Загварын гүйцэтгэл (Тестийн өгөгдөл)")
                    model_m = results['model_metrics']
                    if not model_m.empty:
                        st.dataframe(model_m, use_container_width=True, hide_index=True)

                        # Метрик харьцуулалтын бар чарт
                        melt_cols = ['Precision','Recall','F1-Score','AUC-ROC']
                        existing_cols = [c for c in melt_cols if c in model_m.columns]
                        if existing_cols:
                            melt_df = model_m.melt(id_vars='Загвар', value_vars=existing_cols,
                                                    var_name='Метрик', value_name='Утга')
                            fig_met = px.bar(melt_df, x='Загвар', y='Утга', color='Метрик',
                                             barmode='group', title='Загварын гүйцэтгэлийн харьцуулалт',
                                             text='Утга')
                            fig_met.update_traces(texttemplate='%{text:.3f}', textposition='outside')
                            fig_met.update_layout(yaxis_range=[0, 1.1])
                            st.plotly_chart(fig_met, use_container_width=True)

                    # Cross-validation
                    cv_m = results.get('cv_metrics', pd.DataFrame())
                    if cv_m is not None and not cv_m.empty:
                        st.markdown("### 🔄 Cross-validation үр дүн (Сургалтын өгөгдөл)")
                        st.dataframe(cv_m, use_container_width=True, hide_index=True)

                    # ROC Curve
                    roc = results.get('roc_data', {})
                    if roc:
                        st.markdown("### 📈 ROC муруй")
                        fig_roc = go.Figure()
                        colors = {'Random Forest': '#2196F3', 'Gradient Boosting': '#4CAF50', 'Logistic Regression': '#FF9800'}
                        for name, data in roc.items():
                            fig_roc.add_trace(go.Scatter(
                                x=data['fpr'], y=data['tpr'], mode='lines',
                                name=f"{name} (AUC={data['auc']:.3f})",
                                line=dict(color=colors.get(name, '#999'), width=2)
                            ))
                        fig_roc.add_trace(go.Scatter(
                            x=[0,1], y=[0,1], mode='lines',
                            name='Random (AUC=0.500)', line=dict(dash='dash', color='gray')
                        ))
                        fig_roc.update_layout(
                            title='ROC муруй — Загварын харьцуулалт',
                            xaxis_title='False Positive Rate (1 - Specificity)',
                            yaxis_title='True Positive Rate (Sensitivity)',
                            legend=dict(x=0.55, y=0.05),
                            width=800, height=500
                        )
                        st.plotly_chart(fig_roc, use_container_width=True)

                    # ═══ DR = 1 - Recall ═══
                    st.markdown("### 🎯 Илрүүлэлтийн эрсдэл (DR = 1 − Recall) — AI vs MUS")
                    dr_df = compute_detection_risk(results, mus_coverage=0.20)
                    if not dr_df.empty:
                        st.dataframe(dr_df, use_container_width=True, hide_index=True)
                        # DR харьцуулалтын бар чарт
                        fig_dr = go.Figure()
                        fig_dr.add_trace(go.Bar(name='AI DR (%)', x=dr_df['Загвар'], y=dr_df['AI DR (%)'], marker_color='#2196F3', text=dr_df['AI DR (%)'], textposition='outside'))
                        fig_dr.add_trace(go.Bar(name='MUS DR (%)', x=dr_df['Загвар'], y=dr_df['MUS DR (%)'], marker_color='#FF5722', text=dr_df['MUS DR (%)'], textposition='outside'))
                        fig_dr.add_hline(y=5, line_dash="dash", line_color="green", annotation_text="ISA 200 шаардлага (<5%)")
                        fig_dr.update_layout(title='Илрүүлэлтийн эрсдэл: AI загвар vs MUS (20%)', barmode='group', yaxis_title='DR (%)')
                        st.plotly_chart(fig_dr, use_container_width=True)
                    
                    # ═══ ISA Standard Mapping ═══
                    fi_df = results.get('feature_importance', pd.DataFrame())
                    if not fi_df.empty:
                        st.markdown("### 📋 ISA Стандарт ↔ Шинж чанарын уялдаа")
                        isa_report = get_isa_feature_report(fi_df)
                        if not isa_report.empty:
                            st.dataframe(isa_report, use_container_width=True, hide_index=True)

                    # ═══ Бенфордын шинжилгээ ═══
                    test_df = results.get('test_df', pd.DataFrame())
                    if not test_df.empty and ('amount' in test_df.columns or 'debit_mnt' in test_df.columns):
                        st.markdown("### 📊 Бенфордын хуулийн шинжилгээ")
                        amt_col = 'amount' if 'amount' in test_df.columns else 'debit_mnt'
                        benford = run_benford_analysis(test_df[amt_col])
                        if benford:
                            bc1, bc2 = st.columns(2)
                            with bc1:
                                st.metric("χ² статистик", f"{benford['chi2']:.2f}")
                                st.metric("p-утга", f"{benford['p_value']:.6f}")
                            with bc2:
                                st.metric("Нийт гүйлгээ", f"{benford['n']:,}")
                                verdict = "✅ Нийцэж байна" if benford['conform'] else "⚠️ Нийцэхгүй (p<0.05)"
                                st.metric("Бенфордын хууль", verdict)
                            
                            # Benford chart
                            ben_df = pd.DataFrame({
                                'Цифр': list(range(1, 10)),
                                'Ажиглагдсан': [benford['observed'][d] for d in range(1, 10)],
                                'Хүлээгдэж буй': [benford['expected'][d] for d in range(1, 10)],
                            })
                            fig_ben = go.Figure()
                            fig_ben.add_trace(go.Bar(name='Ажиглагдсан', x=ben_df['Цифр'], y=ben_df['Ажиглагдсан'], marker_color='#2196F3'))
                            fig_ben.add_trace(go.Scatter(name='Бенфордын хууль', x=ben_df['Цифр'], y=ben_df['Хүлээгдэж буй'], mode='lines+markers', line=dict(color='red', width=2)))
                            fig_ben.update_layout(title='Эхний цифрийн тархалт vs Бенфордын хууль', xaxis_title='Эхний цифр', yaxis_title='Давтамж', barmode='overlay')
                            st.plotly_chart(fig_ben, use_container_width=True)

                    # ═══ Эрсдэлийн оноо ═══
                    if not test_df.empty:
                        risk_scores = compute_risk_score(test_df, feat_cols)
                        test_df['risk_score'] = risk_scores
                        st.markdown("### 🔥 Эрсдэлийн оноо (0-100)")
                        rc1, rc2, rc3, rc4 = st.columns(4)
                        rc1.metric("🔴 Өндөр (>70)", f"{(risk_scores > 70).sum():,}")
                        rc2.metric("🟡 Дунд (40-70)", f"{((risk_scores >= 40) & (risk_scores <= 70)).sum():,}")
                        rc3.metric("🟢 Бага (<40)", f"{(risk_scores < 40).sum():,}")
                        rc4.metric("Дундаж оноо", f"{risk_scores.mean():.1f}")
                        
                        fig_risk = px.histogram(x=risk_scores, nbins=50, title='Эрсдэлийн оноон тархалт', 
                                               labels={'x': 'Эрсдэлийн оноо', 'y': 'Тоо'}, color_discrete_sequence=['#0F4C81'])
                        fig_risk.add_vline(x=70, line_dash="dash", line_color="red", annotation_text="Өндөр эрсдэл")
                        fig_risk.add_vline(x=40, line_dash="dash", line_color="orange", annotation_text="Дунд")
                        st.plotly_chart(fig_risk, use_container_width=True)

                    # Confusion Matrix
                    cms = results.get('confusion_matrices', {})
                    if cms:
                        st.markdown("### 🧮 Confusion Matrix")
                        for name, cm in cms.items():
                            with st.expander(f"📊 {name}", expanded=False):
                                labels_cm = ['Хэвийн (0)', 'Аномали (1)']
                                fig_cm = go.Figure(data=go.Heatmap(
                                    z=cm, x=labels_cm, y=labels_cm,
                                    text=cm, texttemplate="%{text:,}",
                                    colorscale='Blues', showscale=True
                                ))
                                fig_cm.update_layout(
                                    title=f'{name} — Confusion Matrix',
                                    xaxis_title='Таамаглал', yaxis_title='Бодит утга',
                                    width=450, height=400
                                )
                                st.plotly_chart(fig_cm, use_container_width=True)

                                tn, fp, fn, tp = cm.ravel()
                                m1, m2, m3, m4 = st.columns(4)
                                m1.metric("True Negative", f"{tn:,}")
                                m2.metric("False Positive", f"{fp:,}")
                                m3.metric("False Negative", f"{fn:,}")
                                m4.metric("True Positive", f"{tp:,}")

                    # Feature Importance
                    fi = results.get('feature_importance', pd.DataFrame())
                    if fi is not None and not fi.empty:
                        st.markdown("### 🔎 Feature Importance (загвар тус бүрээр)")
                        fi_tabs = st.tabs(fi['Загвар'].unique().tolist())
                        for tab, model_name in zip(fi_tabs, fi['Загвар'].unique()):
                            with tab:
                                fi_sub = fi[fi['Загвар'] == model_name].sort_values('Ач холбогдол', ascending=True)
                                fig_fi = px.bar(fi_sub, x='Ач холбогдол', y='Шинж чанар',
                                                orientation='h', title=f'{model_name} — Feature Importance',
                                                color='Ач холбогдол', color_continuous_scale='Blues')
                                st.plotly_chart(fig_fi, use_container_width=True)

                    # SHAP
                    shap_fi = results.get('shap_importance', pd.DataFrame())
                    if shap_fi is not None and not shap_fi.empty:
                        st.markdown("### 🧠 SHAP тайлбарлагч")
                        fig_shap = px.bar(shap_fi.sort_values('SHAP ач холбогдол', ascending=True),
                                          x='SHAP ач холбогдол', y='Шинж чанар', orientation='h',
                                          title='SHAP Feature Importance', color='SHAP ач холбогдол',
                                          color_continuous_scale='Reds')
                        st.plotly_chart(fig_shap, use_container_width=True)

                    # Unsupervised vs Supervised
                    st.markdown("### ⚖️ Unsupervised vs Supervised харьцуулалт")
                    unsup_sum = unsup_results.get('unsupervised_summary', pd.DataFrame())
                    if not unsup_sum.empty and not model_m.empty:
                        sup_sum = model_m[['Загвар','Precision','Recall','F1-Score','AUC-ROC']].copy()
                        sup_sum['Төрөл'] = 'Supervised'
                        sup_sum.columns = ['Загвар','Precision','Recall','F1-Score','AUC-ROC','Төрөл']

                        unsup_display = unsup_sum[['Загвар','Аномали %']].copy()
                        unsup_display['Төрөл'] = 'Unsupervised'
                        st.markdown("**Unsupervised загварууд:**")
                        st.dataframe(unsup_sum, use_container_width=True, hide_index=True)
                        st.markdown("**Supervised загварууд (тестийн өгөгдөл):**")
                        st.dataframe(sup_sum, use_container_width=True, hide_index=True)

                        # Нэгдсэн харьцуулалт чарт
                        fig_vs = make_subplots(rows=1, cols=2,
                                               subplot_titles=["Unsupervised: Аномали %", "Supervised: F1 & AUC"])
                        fig_vs.add_trace(
                            go.Bar(x=unsup_sum['Загвар'], y=unsup_sum['Аномали %'], name='Аномали %',
                                   marker_color='#FF7043'), row=1, col=1
                        )
                        if 'F1-Score' in sup_sum.columns:
                            fig_vs.add_trace(
                                go.Bar(x=sup_sum['Загвар'], y=sup_sum['F1-Score'], name='F1-Score',
                                       marker_color='#42A5F5'), row=1, col=2
                            )
                        if 'AUC-ROC' in sup_sum.columns:
                            fig_vs.add_trace(
                                go.Bar(x=sup_sum['Загвар'], y=sup_sum['AUC-ROC'], name='AUC-ROC',
                                       marker_color='#66BB6A'), row=1, col=2
                            )
                        fig_vs.update_layout(title='Unsupervised vs Supervised загварын харьцуулалт', height=450)
                        st.plotly_chart(fig_vs, use_container_width=True)

                    # Хамгийн сайн загвар
                    best = results.get('best_model_name', '')
                    if best:
                        best_row = model_m[model_m['Загвар'] == best].iloc[0] if not model_m[model_m['Загвар'] == best].empty else None
                        if best_row is not None:
                            st.markdown(f"""
                            ### 🏆 Хамгийн сайн загвар: **{best}**
                            | Метрик | Утга |
                            |--------|------|
                            | Precision | {best_row.get('Precision', 0):.4f} |
                            | Recall | {best_row.get('Recall', 0):.4f} |
                            | F1-Score | {best_row.get('F1-Score', 0):.4f} |
                            | AUC-ROC | {best_row.get('AUC-ROC', 0):.4f} |
                            """)

                    # Тестийн аномали жагсаалт
                    test_df = results.get('test_df', pd.DataFrame())
                    if not test_df.empty:
                        st.markdown("### 📋 Тестийн аномали гүйлгээнүүд")
                        # Он-оор шүүх
                        yr_col = None
                        for yc in ['report_year', 'year']:
                            if yc in test_df.columns:
                                yr_col = yc
                                break
                        if yr_col:
                            yrs = ['Бүгд'] + sorted(test_df[yr_col].dropna().unique().astype(str).tolist())
                            sel_yr = st.selectbox("📅 Он сонгох", yrs, key='ml7_anom_year')
                            if sel_yr != 'Бүгд':
                                test_filtered = test_df[test_df[yr_col].astype(str) == sel_yr]
                            else:
                                test_filtered = test_df
                        else:
                            test_filtered = test_df
                        test_anom = test_filtered[test_filtered['pseudo_label']==1]
                        show_cols = [c for c in ['transaction_date','account_code','account_name',
                                     'counterparty_name','amount','vote_count','transaction_description'] if c in test_anom.columns]
                        st.dataframe(test_anom[show_cols].head(500), use_container_width=True, hide_index=True)

                    # CSV татах
                    _show_dataframe_download(model_m, 'ml_model_metrics.csv')
                    _show_dataframe_download(test_df, 'ml_test_predictions.csv')

                    st.session_state['ml7_results'] = results
                    st.session_state['ml7_unsup'] = unsup_results

    # ═══ КЭШЛЭГДСЭН ҮР ДҮН ХАРУУЛАХ (цэс хооронд шилжихэд) ═══
    elif st.session_state.get('ml7_results', {}).get('success'):
        results = st.session_state['ml7_results']
        st.info("📦 Өмнөх шинжилгээний үр дүн (дахин ажиллуулахгүйгээр харж байна)")
        
        ps = results.get('pseudo_label_stats', {})
        if ps:
            cc1, cc2, cc3, cc4 = st.columns(4)
            cc1.metric("Нийт мөр", f"{ps.get('total',0):,}")
            cc2.metric("Аномали", f"{ps.get('anomaly',0):,}")
            cc3.metric("Хэвийн", f"{ps.get('normal',0):,}")
            cc4.metric("Аномали %", f"{ps.get('anomaly_pct',0):.1f}%")

        model_m = results.get('model_metrics', pd.DataFrame())
        if not model_m.empty:
            st.markdown("### 🎯 Загварын гүйцэтгэл")
            st.dataframe(model_m, use_container_width=True, hide_index=True)

        # DR тооцоо
        dr_df = compute_detection_risk(results, mus_coverage=0.20)
        if not dr_df.empty:
            st.markdown("### 🎯 Илрүүлэлтийн эрсдэл (DR = 1 − Recall)")
            st.dataframe(dr_df, use_container_width=True, hide_index=True)

        # Тест өгөгдлийн аномали
        test_df = results.get('test_df', pd.DataFrame())
        if not test_df.empty:
            # Он-оор шүүх
            year_col_ml = None
            for yc in ['report_year', 'year', 'fiscal_year']:
                if yc in test_df.columns:
                    year_col_ml = yc
                    break
            if year_col_ml:
                years_ml = ['Бүгд'] + sorted(test_df[year_col_ml].dropna().unique().astype(str).tolist())
                sel_year_ml = st.selectbox("📅 Он сонгох", years_ml, key='ml7_cache_year')
                if sel_year_ml != 'Бүгд':
                    test_show = test_df[test_df[year_col_ml].astype(str) == sel_year_ml]
                else:
                    test_show = test_df
            else:
                test_show = test_df
            
            test_anom = test_show[test_show['pseudo_label']==1] if 'pseudo_label' in test_show.columns else pd.DataFrame()
            if not test_anom.empty:
                st.markdown(f"### 📋 Тестийн аномали: {len(test_anom):,} мөр")
                show_cols = [c for c in ['account_code','account_name','counterparty_name','amount','vote_count','risk_score','transaction_date'] if c in test_anom.columns]
                st.dataframe(test_anom[show_cols].head(500), use_container_width=True, hide_index=True)
                _show_dataframe_download(test_anom, 'anomaly_test.csv')


# ═══════════════════════════════════════════════════════════
# 8️⃣ НАРИЙВЧИЛСАН ML ШИНЖИЛГЭЭ
# ═══════════════════════════════════════════════════════════
if page.startswith("7"):
    st.header("7️⃣ Нарийвчилсан машин сургалт")
    st.markdown("""
    **Үе шат 3-4:** Салбар хоорондын pattern detection • Learning curve • Hyperparameter tuning • 
    Тогтвортой байдал • McNemar тест • Contamination мэдрэмж
    """)

    feat_cols_8 = ['log_amount','acct_cat_num','benford_dev','is_round','amt_zscore',
                   'cp_rare','pair_rare','desc_empty','is_month_end','is_year_end',
                   'is_dup','is_debit','desc_mismatch','name_no_overlap','dir_mismatch']

    # Эх өгөгдөл
    has_branch = bool(st.session_state.get('branch_engineered'))
    has_ml7 = bool(st.session_state.get('ml7_results', {}).get('success'))

    if not has_branch and not has_ml7:
        st.warning("⚠️ Эхлээд 5️⃣ Салбарын шинжилгээ эсвэл 6️⃣ Сургалт/Шалгалт машин сургалтыг ажиллуулна уу.")
    else:
        adv_tabs = st.tabs(["🔀 Салбар хоорондын хэв маяг", "📈 Learning Curve",
                             "🔧 Hyperparameter Tuning", "🔒 Тогтвортой байдал",
                             "📊 McNemar тест", "🎚️ Contamination мэдрэмж"])

        # ─── TAB 1: Cross-branch patterns ───
        with adv_tabs[0]:
            st.markdown("### 🔀 Салбар хоорондын хэв маяг")
            branch_eng = st.session_state.get('branch_engineered', {})

            if len(branch_eng) < 2:
                st.info("Дор хаяж 2 салбарын өгөгдөл шаардлагатай. 5️⃣ цэсэнд бэлтгэнэ үү.")
            else:
                if st.button("🔀 Cross-branch шинжилгээ", key='run_cross_branch'):
                    with st.spinner("Салбар хоорондын хэв маяг хайж байна..."):
                        cb_cont = st.session_state.get('branch_cont', 0.05)
                        cb_results = run_cross_branch_patterns(branch_eng, feat_cols_8, cb_cont)

                        # Аномали төрлийн матриц
                        atm = cb_results.get('anomaly_type_matrix', pd.DataFrame())
                        if not atm.empty:
                            st.markdown("#### 📋 Салбар бүрийн аномали профайл")
                            st.dataframe(atm, use_container_width=True, hide_index=True)

                            fig_atm = px.bar(atm, x='branch', y='anomaly_pct', color='top_feature_1',
                                             title='Салбар бүрийн аномали % + давамгайлах feature',
                                             text='anomaly_pct')
                            fig_atm.update_traces(textposition='outside')
                            st.plotly_chart(fig_atm, use_container_width=True)

                        # Нийтлэг хэв маяг
                        common = cb_results.get('common_patterns', [])
                        if common:
                            st.markdown("#### 🌐 Бүх салбарт нийтлэг хэв маяг")
                            st.dataframe(pd.DataFrame(common), use_container_width=True, hide_index=True)
                        else:
                            st.info("Бүх салбарт нийтлэг хэв маяг олдсонгүй — салбарууд өвөрмөц.")

                        # Өвөрмөц хэв маяг
                        unique = cb_results.get('unique_patterns', [])
                        if unique:
                            st.markdown("#### 🔍 Зөвхөн нэг салбарт илэрсэн өвөрмөц хэв маяг")
                            st.dataframe(pd.DataFrame(unique), use_container_width=True, hide_index=True)

                        # Корреляцийн матриц
                        corr = cb_results.get('branch_correlation', pd.DataFrame())
                        if not corr.empty:
                            st.markdown("#### 🔗 Салбар хоорондын корреляци")
                            fig_corr = go.Figure(data=go.Heatmap(
                                z=corr.values, x=corr.columns, y=corr.index,
                                text=corr.round(2).values, texttemplate="%{text}",
                                colorscale='RdBu_r', zmin=-1, zmax=1
                            ))
                            fig_corr.update_layout(title='Салбар хоорондын feature корреляци',
                                                    width=600, height=500)
                            st.plotly_chart(fig_corr, use_container_width=True)

                        st.session_state['cross_branch_results'] = cb_results

        # ─── TAB 2: Learning Curve ───
        with adv_tabs[1]:
            st.markdown("### 📈 Learning Curve (Сургалтын процессийн муруй)")
            st.caption("Сургалтын өгөгдлийн хэмжээ нэмэгдэхэд гүйцэтгэл хэрхэн өөрчлөгдөхийг харуулна.")

            ml7_res = st.session_state.get('ml7_results', {})
            if not ml7_res.get('success'):
                st.info("Эхлээд 6️⃣ Сургалт/Шалгалт машин сургалтыг ажиллуулна уу.")
            else:
                lc_model = st.selectbox("Загвар сонгох", ['Random Forest', 'Gradient Boosting', 'Logistic Regression'], key='lc_model')
                lc_points = st.slider("Цэгийн тоо", 5, 15, 8, key='lc_points')

                if st.button("📈 Learning curve тооцоолох", key='run_lc'):
                    train_df = ml7_res['train_df']
                    test_df = ml7_res['test_df']
                    all_df = pd.concat([train_df, test_df], ignore_index=True)

                    for c in feat_cols_8:
                        if c not in all_df.columns:
                            all_df[c] = 0
                    X_all = all_df[feat_cols_8].fillna(0).replace([np.inf,-np.inf],0).astype(float)
                    y_all = all_df['pseudo_label'].values if 'pseudo_label' in all_df.columns else np.zeros(len(all_df))

                    model_map = {
                        'Random Forest': (RandomForestClassifier, {'n_estimators':200, 'class_weight':'balanced', 'max_depth':10, 'random_state':42}),
                        'Gradient Boosting': (GradientBoostingClassifier, {'n_estimators':150, 'max_depth':5, 'random_state':42}),
                        'Logistic Regression': (LogisticRegression, {'max_iter':1000, 'class_weight':'balanced', 'random_state':42}),
                    }

                    with st.spinner("Learning curve тооцоолж байна..."):
                        mc, mp = model_map[lc_model]
                        lc_df = run_learning_curve(X_all, y_all, mc, mp, n_points=lc_points)

                    if lc_df is not None:
                        fig_lc = go.Figure()
                        fig_lc.add_trace(go.Scatter(
                            x=lc_df['train_size'], y=lc_df['train_f1_mean'],
                            mode='lines+markers', name='Сургалт (Train)',
                            line=dict(color='#2196F3'), fill='tonexty' if False else None
                        ))
                        fig_lc.add_trace(go.Scatter(
                            x=lc_df['train_size'], y=lc_df['test_f1_mean'],
                            mode='lines+markers', name='Шалгалт (Validation)',
                            line=dict(color='#FF5722')
                        ))
                        # Error bands
                        fig_lc.add_trace(go.Scatter(
                            x=np.concatenate([lc_df['train_size'], lc_df['train_size'][::-1]]),
                            y=np.concatenate([lc_df['train_f1_mean']+lc_df['train_f1_std'],
                                              (lc_df['train_f1_mean']-lc_df['train_f1_std'])[::-1]]),
                            fill='toself', fillcolor='rgba(33,150,243,0.15)',
                            line=dict(color='rgba(255,255,255,0)'), showlegend=False
                        ))
                        fig_lc.add_trace(go.Scatter(
                            x=np.concatenate([lc_df['train_size'], lc_df['train_size'][::-1]]),
                            y=np.concatenate([lc_df['test_f1_mean']+lc_df['test_f1_std'],
                                              (lc_df['test_f1_mean']-lc_df['test_f1_std'])[::-1]]),
                            fill='toself', fillcolor='rgba(255,87,34,0.15)',
                            line=dict(color='rgba(255,255,255,0)'), showlegend=False
                        ))
                        fig_lc.update_layout(
                            title=f'{lc_model} — Learning Curve',
                            xaxis_title='Сургалтын мөрийн тоо',
                            yaxis_title='F1-Score',
                            yaxis_range=[0, 1.05]
                        )
                        st.plotly_chart(fig_lc, use_container_width=True)
                        st.dataframe(lc_df.round(4), use_container_width=True, hide_index=True)
                    else:
                        st.warning("Learning curve тооцоолж чадсангүй.")

        # ─── TAB 3: Hyperparameter Tuning ───
        with adv_tabs[2]:
            st.markdown("### 🔧 Hyperparameter Tuning")
            ml7_res = st.session_state.get('ml7_results', {})
            if not ml7_res.get('success'):
                st.info("Эхлээд 6️⃣ Сургалт/Шалгалт машин сургалтыг ажиллуулна уу.")
            else:
                hp_model = st.selectbox("Загвар", ['Random Forest','Gradient Boosting','Logistic Regression'], key='hp_model')

                if st.button("🔧 Hyperparameter хайлт эхлүүлэх", key='run_hp'):
                    train_df = ml7_res['train_df']
                    test_df = ml7_res['test_df']
                    for c in feat_cols_8:
                        if c not in train_df.columns: train_df[c] = 0
                        if c not in test_df.columns: test_df[c] = 0
                    X_tr = train_df[feat_cols_8].fillna(0).replace([np.inf,-np.inf],0).astype(float)
                    X_te = test_df[feat_cols_8].fillna(0).replace([np.inf,-np.inf],0).astype(float)
                    y_tr = train_df['pseudo_label'].values
                    y_te = test_df['pseudo_label'].values

                    with st.spinner(f"{hp_model} hyperparameter хайж байна..."):
                        hp_df = run_hyperparameter_search(X_tr, y_tr, X_te, y_te, hp_model)

                    if not hp_df.empty:
                        st.dataframe(hp_df.sort_values('F1', ascending=False), use_container_width=True, hide_index=True)

                        # Хамгийн сайн тохиргоо
                        best = hp_df.loc[hp_df['F1'].idxmax()]
                        st.success(f"🏆 Хамгийн сайн тохиргоо: F1={best['F1']:.4f}, AUC={best['AUC']:.4f}")

                        # F1 vs параметр график
                        if 'n_estimators' in hp_df.columns:
                            fig_hp = px.scatter(hp_df, x='n_estimators', y='F1',
                                                color='max_depth' if 'max_depth' in hp_df.columns else None,
                                                size='AUC', title=f'{hp_model} — Hyperparameter vs F1')
                            st.plotly_chart(fig_hp, use_container_width=True)
                        elif 'C' in hp_df.columns:
                            fig_hp = px.line(hp_df, x='C', y=['F1','AUC'], title='Logistic Regression — C vs Performance', log_x=True)
                            st.plotly_chart(fig_hp, use_container_width=True)

                        st.session_state['hp_results'] = hp_df

        # ─── TAB 4: Stability ───
        with adv_tabs[3]:
            st.markdown("### 🔒 Загварын тогтвортой байдал")
            st.caption("Олон удаа (10 run) train/test хуваалт хийж, метрикүүдийн тархалтыг шалгана.")
            ml7_res = st.session_state.get('ml7_results', {})

            if not ml7_res.get('success'):
                st.info("Эхлээд 7️⃣ pipeline-г ажиллуулна уу.")
            else:
                n_runs = st.slider("Run тоо", 5, 30, 10, key='stab_runs')
                if st.button("🔒 Тогтвортой байдлын шинжилгээ", key='run_stab'):
                    all_df = pd.concat([ml7_res['train_df'], ml7_res['test_df']], ignore_index=True)
                    for c in feat_cols_8:
                        if c not in all_df.columns: all_df[c] = 0
                    X = all_df[feat_cols_8].fillna(0).replace([np.inf,-np.inf],0).astype(float)
                    y = all_df['pseudo_label'].values

                    with st.spinner(f"Тогтвортой байдал шалгаж байна ({n_runs} run)..."):
                        stab_raw, stab_summary = run_stability_analysis(X, y, feat_cols_8, n_runs=n_runs)

                    if not stab_summary.empty:
                        st.markdown("#### 📊 Нэгтгэсэн үр дүн")
                        st.dataframe(stab_summary, use_container_width=True, hide_index=True)

                        # Box plot
                        fig_box = px.box(stab_raw, x='Загвар', y='F1', color='Загвар',
                                         title=f'F1-Score тархалт ({n_runs} run)', points='all')
                        st.plotly_chart(fig_box, use_container_width=True)

                        fig_box2 = px.box(stab_raw, x='Загвар', y='AUC', color='Загвар',
                                          title=f'AUC-ROC тархалт ({n_runs} run)', points='all')
                        st.plotly_chart(fig_box2, use_container_width=True)

                        # CV% тайлбар
                        st.markdown("#### 📖 Тогтвортой байдлын тайлбар")
                        for _, row in stab_summary.iterrows():
                            cv_f1 = row.get('F1_CV%', 0)
                            if cv_f1 < 5:
                                verdict = "🟢 Маш тогтвортой"
                            elif cv_f1 < 10:
                                verdict = "🟡 Хүлээн зөвшөөрөгдөхүйц"
                            else:
                                verdict = "🔴 Тогтворгүй"
                            st.write(f"**{row['Загвар']}**: F1 CV={cv_f1:.1f}% → {verdict}")

                        st.session_state['stability_summary'] = stab_summary
                        st.session_state['stability_raw'] = stab_raw

        # ─── TAB 5: McNemar test ───
        with adv_tabs[4]:
            st.markdown("### 📊 McNemar тест — Загвар хоорондын статистик харьцуулалт")
            ml7_res = st.session_state.get('ml7_results', {})

            if not ml7_res.get('success'):
                st.info("Эхлээд 7️⃣ pipeline-г ажиллуулна уу.")
            else:
                if st.button("📊 McNemar тест ажиллуулах", key='run_mcnemar'):
                    test_df = ml7_res['test_df']
                    y_true = test_df['pseudo_label'].values

                    model_names = ['Random Forest', 'Gradient Boosting', 'Logistic Regression']
                    preds = {}
                    for name in model_names:
                        col = f'{name}_pred'
                        if col in test_df.columns:
                            preds[name] = test_df[col].values

                    mcnemar_rows = []
                    pairs = [(model_names[i], model_names[j])
                             for i in range(len(model_names)) for j in range(i+1, len(model_names))]

                    for a, b in pairs:
                        if a in preds and b in preds:
                            result = run_mcnemar_test(y_true, preds[a], preds[b], a, b)
                            mcnemar_rows.append(result)

                    if mcnemar_rows:
                        mcn_df = pd.DataFrame(mcnemar_rows)
                        st.dataframe(mcn_df, use_container_width=True, hide_index=True)

                        # Тайлбар
                        for row in mcnemar_rows:
                            p = row.get('p-value', 1)
                            icon = "🟢" if p < 0.05 else "⚪"
                            st.write(f"{icon} **{row.get('Model A','')} vs {row.get('Model B','')}**: "
                                     f"χ²={row.get('McNemar χ²',0):.3f}, p={p:.4f} — {row.get('Дүгнэлт','')}")

                        st.session_state['mcnemar_results'] = mcnemar_rows
                    else:
                        st.warning("Таамаглалын өгөгдөл олдсонгүй.")

        # ─── TAB 6: Contamination sensitivity ───
        with adv_tabs[5]:
            st.markdown("### 🎚️ Contamination параметрийн мэдрэмжийн шинжилгээ")
            st.caption("Аномалийн хувийг (contamination) өөрчлөхөд загварын гүйцэтгэл хэрхэн өөрчлөгдөхийг харуулна.")

            ml7_res = st.session_state.get('ml7_results', {})
            if not ml7_res.get('success'):
                st.info("Эхлээд 7️⃣ pipeline-г ажиллуулна уу.")
            else:
                if st.button("🎚️ Contamination мэдрэмж тооцоолох", key='run_cont_sens'):
                    all_df = pd.concat([ml7_res['train_df'], ml7_res['test_df']], ignore_index=True)
                    for c in feat_cols_8:
                        if c not in all_df.columns: all_df[c] = 0
                    X = all_df[feat_cols_8].fillna(0).replace([np.inf,-np.inf],0).astype(float)

                    def pseudo_fn(cont):
                        return create_pseudo_labels(all_df, feat_cols_8, cont)

                    with st.spinner("Contamination мэдрэмж тооцоолж байна..."):
                        sens_df = run_contamination_sensitivity(X, pseudo_fn, feat_cols_8)

                    if not sens_df.empty:
                        st.dataframe(sens_df, use_container_width=True, hide_index=True)

                        fig_sens = go.Figure()
                        fig_sens.add_trace(go.Scatter(
                            x=sens_df['Contamination'], y=sens_df['F1'],
                            mode='lines+markers', name='F1-Score', line=dict(color='#2196F3', width=3)
                        ))
                        fig_sens.add_trace(go.Scatter(
                            x=sens_df['Contamination'], y=sens_df['AUC'],
                            mode='lines+markers', name='AUC-ROC', line=dict(color='#4CAF50', width=3)
                        ))
                        fig_sens.add_trace(go.Scatter(
                            x=sens_df['Contamination'], y=sens_df['Precision'],
                            mode='lines+markers', name='Precision', line=dict(color='#FF9800', width=2, dash='dash')
                        ))
                        fig_sens.add_trace(go.Scatter(
                            x=sens_df['Contamination'], y=sens_df['Recall'],
                            mode='lines+markers', name='Recall', line=dict(color='#F44336', width=2, dash='dash')
                        ))
                        fig_sens.update_layout(
                            title='Contamination параметрийн мэдрэмж',
                            xaxis_title='Contamination',
                            yaxis_title='Метрикийн утга',
                            yaxis_range=[0, 1.05]
                        )
                        st.plotly_chart(fig_sens, use_container_width=True)

                        # Зохистой contamination олох
                        best_row = sens_df.loc[sens_df['F1'].idxmax()]
                        st.success(f"🏆 Зохистой contamination = **{best_row['Contamination']}** "
                                   f"(F1={best_row['F1']:.4f}, AUC={best_row['AUC']:.4f})")

                        st.session_state['contamination_sens'] = sens_df


# ═══════════════════════════════════════════════════════════
# 8️⃣ ДИССЕРТАЦИЙН ГАРАЛТ
# ═══════════════════════════════════════════════════════════
if page.startswith("8"):
    st.header("8️⃣ Диссертацийн чанартай гаралт")
    st.markdown("""
    Бүх шинжилгээний үр дүнг диссертацид бэлэн формат руу нэгтгэнэ:
    **Хүснэгтүүд • Графикууд • Excel нэгтгэл**
    """)

    ml7_res = st.session_state.get('ml7_results', {})
    branch_eng = st.session_state.get('branch_engineered', {})
    stab_summary = st.session_state.get('stability_summary', pd.DataFrame())
    mcnemar_res = st.session_state.get('mcnemar_results', [])
    hp_res = st.session_state.get('hp_results', pd.DataFrame())
    cb_results = st.session_state.get('cross_branch_results', {})
    cont_sens = st.session_state.get('contamination_sens', pd.DataFrame())

    # Бэлэн болсон мэдээлэл
    ready_items = []
    if ml7_res.get('success'):
        ready_items.append("✅ Сургалт/Шалгалт машин сургалт")
    else:
        ready_items.append("❌ Сургалт/Шалгалт машин сургалт (6️⃣)")
    if branch_eng:
        ready_items.append(f"✅ Салбарын шинжилгээ ({len(branch_eng)} салбар)")
    else:
        ready_items.append("❌ Салбарын шинжилгээ (6️⃣)")
    if not stab_summary.empty:
        ready_items.append("✅ Тогтвортой байдлын шинжилгээ")
    if mcnemar_res:
        ready_items.append("✅ McNemar тест")
    if not hp_res.empty:
        ready_items.append("✅ Hyperparameter tuning")
    if cb_results:
        ready_items.append("✅ Салбар хоорондын хэв маяг")
    if not cont_sens.empty:
        ready_items.append("✅ Contamination мэдрэмж")

    st.markdown("### 📋 Бэлэн болсон мэдээллүүд:")
    for item in ready_items:
        st.write(item)

    if st.button("📄 Диссертацийн хүснэгтүүд үүсгэх", type='primary', use_container_width=True, key='gen_diss'):
        with st.spinner("Хүснэгтүүд үүсгэж байна..."):
            # Салбарын харьцуулалт
            feat_cols_9 = ['log_amount','acct_cat_num','benford_dev','is_round','amt_zscore',
                           'cp_rare','pair_rare','desc_empty','is_month_end','is_year_end',
                           'is_dup','is_debit','desc_mismatch','name_no_overlap','dir_mismatch']
            branch_comp = pd.DataFrame()
            if branch_eng:
                branch_comp, _ = run_branch_comparison(branch_eng, feat_cols_9,
                    contamination=st.session_state.get('branch_cont', 0.05))

            tables = generate_dissertation_tables(
                ml_results=ml7_res if ml7_res.get('success') else None,
                branch_comparison=branch_comp if not branch_comp.empty else None,
                stability=stab_summary if not stab_summary.empty else None,
                mcnemar_results=mcnemar_res if mcnemar_res else None,
                hp_results=hp_res if not hp_res.empty else None,
            )

            if not tables:
                st.warning("Хүснэгт үүсгэхэд хангалттай мэдээлэл алга.")
            else:
                st.success(f"✅ {len(tables)} хүснэгт үүсгэлээ!")

                # Хүснэгт тус бүрийг харуулах
                for tname, tdf in tables.items():
                    display_name = tname.replace('_', ' ')
                    st.markdown(f"#### 📊 {display_name}")
                    st.dataframe(tdf, use_container_width=True, hide_index=True)

                # Excel нэгтгэл
                try:
                    buf = io.BytesIO()
                    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                        for tname, tdf in tables.items():
                            sheet = tname[:31]
                            tdf.to_excel(writer, sheet_name=sheet, index=False)

                        # Нэмэлт sheet-үүд
                        if not cont_sens.empty:
                            cont_sens.to_excel(writer, sheet_name='Contamination_мэдрэмж', index=False)
                        if cb_results:
                            atm = cb_results.get('anomaly_type_matrix', pd.DataFrame())
                            if not atm.empty:
                                atm.to_excel(writer, sheet_name='Cross_branch_profile', index=False)
                            common = cb_results.get('common_patterns', [])
                            if common:
                                pd.DataFrame(common).to_excel(writer, sheet_name='Нийтлэг_хэв_маяг', index=False)
                            unique = cb_results.get('unique_patterns', [])
                            if unique:
                                pd.DataFrame(unique).to_excel(writer, sheet_name='Өвөрмөц_хэв_маяг', index=False)

                    buf.seek(0)
                    st.download_button(
                        "📥 Диссертацийн хүснэгтүүд (Excel)",
                        buf.getvalue(),
                        file_name='dissertation_tables.xlsx',
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        key='dl_diss_xlsx'
                    )
                except Exception as e:
                    st.error(f"Excel үүсгэх үед алдаа: {e}")

    # Диссертацийн текст загвар
    st.markdown("---")
    st.markdown("### 📝 Диссертацийн текст загвар")
    if ml7_res.get('success'):
        mm = ml7_res.get('model_metrics', pd.DataFrame())
        ps = ml7_res.get('pseudo_label_stats', {})
        best_name = ml7_res.get('best_model_name', '')

        best_row = mm[mm['Загвар']==best_name].iloc[0] if not mm.empty and best_name else None

        text = f"""## Судалгааны үр дүн

### 3.1 Өгөгдлийн тодорхойлолт

Судалгаанд нийт {ps.get('total', 0):,} ерөнхий журналын гүйлгээ хамрагдсан.
Unsupervised ensemble арга (Isolation Forest, KMeans, Z-score)-аар
нийт гүйлгээний {ps.get('anomaly_pct', 0):.1f}% буюу {ps.get('anomaly', 0):,} гүйлгээ
хэвийн бус (аномали) гэж тэмдэглэгдсэн.

Өгөгдлийг 80/20 харьцаагаар сургалтын ({ps.get('total',0) - len(ml7_res.get('test_df',[])):,} мөр)
болон тестийн ({len(ml7_res.get('test_df',[])):,} мөр) хэсэгт хуваасан.

### 3.2 Загварын гүйцэтгэлийн харьцуулалт

"""
        if not mm.empty:
            for _, row in mm.iterrows():
                text += f"- **{row['Загвар']}**: Precision={row.get('Precision',0):.4f}, Recall={row.get('Recall',0):.4f}, F1={row.get('F1-Score',0):.4f}, AUC={row.get('AUC-ROC',0):.4f}\n"

        if best_row is not None:
            text += f"""
### 3.3 Хамгийн сайн загвар

Тестийн өгөгдөл дээр F1-Score метрикээр хамгийн өндөр гүйцэтгэл үзүүлсэн загвар нь
**{best_name}** (F1={best_row.get('F1-Score',0):.4f}, AUC-ROC={best_row.get('AUC-ROC',0):.4f}) юм.
"""
        if branch_eng:
            text += f"""
### 3.4 Салбарын шинжилгээ

Нийт {len(branch_eng)} салбарын өгөгдлийг тусд нь шинжилж, салбар хоорондын
эрсдэлийн ялгааг тодорхойлсон.
"""
        if not stab_summary.empty:
            text += """
### 3.5 Загварын тогтвортой байдал

Загвар бүрийн тогтвортой байдлыг олон удаа давтан сургалт хийж шалгасан.
"""
            for _, row in stab_summary.iterrows():
                text += f"- **{row['Загвар']}**: F1 дундаж={row.get('F1_дундаж',0):.4f}±{row.get('F1_std',0):.4f} (CV={row.get('F1_CV%',0):.1f}%)\n"

        st.text_area("📝 Диссертацийн текст (засварлаж болно)", text, height=500, key='diss_text')
        st.download_button(
            "📥 Текст татах (.md)",
            text.encode('utf-8'),
            file_name='dissertation_results.md',
            mime='text/markdown',
            key='dl_diss_md'
        )
    else:
        st.info("6️⃣ Сургалт/Шалгалт машин сургалтыг ажиллуулсны дараа текст загвар автоматаар үүснэ.")



# ═══════════════════════════════════════════════════════════
# ТӨРИЙН АУДИТЫН ТУСГАЙ ХЭСГҮҮД (v10.0-д gov_audit_app-аас нэмсэн)
# ═══════════════════════════════════════════════════════════

# ═══ 2️⃣ МӨНГӨН ГҮЙЛГЭЭНИЙ ЖУРНАЛ ═══
if page.startswith("🏛️"):
    st.header("🏛️ Мөнгөн гүйлгээний журнал")
    st.markdown("*Харилцах дансны мөнгөн гүйлгээний бүртгэл, орлого/зарлага шинжилгээ*")

    # ── Шууд файл оруулах (хэдэн ч он, хэдэн ч файл) ──
    mj_files = st.file_uploader(
        "📎 МЖ файл оруулах (хэдэн ч он, хэдэн ч файл)",
        type=['xlsx','xls','xlsm','xlsb','csv','tsv','gz'],
        accept_multiple_files=True, key='mj_direct_upload'
    )

    # Файл оруулсан бол нэгтгэх
    if mj_files:
        all_dfs = []
        for mf in mj_files:
            try:
                mf_name = mf.name.lower()
                if mf_name.endswith('.csv') or mf_name.endswith('.tsv'):
                    sep = '\t' if mf_name.endswith('.tsv') else ','
                    _df = pd.read_csv(mf, sep=sep)
                elif mf_name.endswith('.gz'):
                    _df = pd.read_csv(mf, compression='gzip')
                else:
                    _df = pd.read_excel(mf)
                _df['_file'] = mf.name
                all_dfs.append(_df)
                st.success(f"✅ {mf.name}: {len(_df):,} мөр")
            except Exception as e:
                st.error(f"❌ {mf.name}: {e}")
        if all_dfs:
            combined = pd.concat(all_dfs, ignore_index=True)
            st.session_state['g_cash'] = combined

    # 1-р цэснээс оруулсан эсвэл шууд оруулсан
    mj = st.session_state.get('g_cash')
    _has_mj = mj is not None and isinstance(mj, pd.DataFrame) and not mj.empty

    if _has_mj:
        col_map = {}
        for c in mj.columns:
            cl = str(c).lower()
            if 'огноо' in cl or 'date' in cl: col_map[c] = 'Огноо'
            elif 'орлого' in cl or 'income' in cl or 'receipt' in cl: col_map[c] = 'Орлого'
            elif 'зарлага' in cl or 'expense' in cl or 'payment' in cl or 'disbursement' in cl: col_map[c] = 'Зарлага'
            elif any(k in cl for k in ['харилцах','банк','bank','үндсэн данс']): col_map[c] = 'Харилцах данс'
            elif any(k in cl for k in ['байгууллага','харилцагч','partner','counterparty','харьцсан данс','харьцсан']): col_map[c] = 'Харилцагч'
            elif any(k in cl for k in ['утга','тайлбар','description','гүйлгээний утга','memo','narration']): col_map[c] = 'Тайлбар'
            elif any(k in cl for k in ['дебит','дебет','debit','dt']): col_map[c] = 'Дебит'
            elif any(k in cl for k in ['кредит','кредет','credit','ct']): col_map[c] = 'Кредит'
            elif 'он' in cl or 'year' in cl: col_map[c] = 'Он'
        mj2 = mj.rename(columns=col_map)

        # Тоон баганыг цэвэрлэх
        for c in ['Орлого','Зарлага','Дебит','Кредит']:
            if c in mj2.columns:
                mj2[c] = pd.to_numeric(mj2[c].astype(str).str.replace(',','').str.replace(' ',''), errors='coerce').fillna(0)

        # Орлого/Зарлага байхгүй бол Дебит/Кредит-ээс тооцох
        if 'Орлого' not in mj2.columns and 'Дебит' in mj2.columns:
            mj2['Орлого'] = mj2['Дебит']
        if 'Зарлага' not in mj2.columns and 'Кредит' in mj2.columns:
            mj2['Зарлага'] = mj2['Кредит']

        # Огнооноос он гаргах
        if 'Огноо' in mj2.columns:
            mj2['Огноо'] = pd.to_datetime(mj2['Огноо'], errors='coerce')
            mj2['Он'] = mj2['Огноо'].dt.year

        # ── Хураангуй тоонууд ──
        _v1 = f"{len(mj2):,}"
        _v2 = f"₮{mj2['Орлого'].sum():,.0f}" if 'Орлого' in mj2.columns else "—"
        _v3 = f"₮{mj2['Зарлага'].sum():,.0f}" if 'Зарлага' in mj2.columns else "—"
        _v4 = f"₮{(mj2.get('Орлого',pd.Series([0])).sum() - mj2.get('Зарлага',pd.Series([0])).sum()):,.0f}"

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Нийт гүйлгээ", _v1)
        m2.metric("Орлого", _v2)
        m3.metric("Зарлага", _v3)
        m4.metric("Цэвэр дүн", _v4)

        # ── Оны шүүлтүүр ──
        if 'Он' in mj2.columns:
            years_available = sorted(mj2['Он'].dropna().unique().astype(int))
            if years_available:
                selected_years = st.multiselect("📅 Он сонгох", years_available, default=years_available, key='mj_year_filter')
                if selected_years:
                    mj2 = mj2[mj2['Он'].isin(selected_years)]

        # ── Харилцах дансны шүүлтүүр ──
        if 'Харилцах данс' in mj2.columns:
            accounts = ['Бүгд'] + sorted(mj2['Харилцах данс'].dropna().astype(str).unique().tolist())
            sel_acc = st.selectbox("🏦 Харилцах данс сонгох", accounts, key='mj_acc_filter')
            if sel_acc != 'Бүгд':
                mj2 = mj2[mj2['Харилцах данс'].astype(str) == sel_acc]

        # ── Таб-ууд ──
        tab_data, tab_chart, tab_partner = st.tabs(["📋 Гүйлгээний жагсаалт", "📊 Графикууд", "👥 Харилцагчийн шинжилгээ"])

        with tab_data:
            st.dataframe(mj2.head(1000), use_container_width=True, hide_index=True)

        with tab_chart:
            _has_orl = 'Орлого' in mj2.columns
            _has_zarl = 'Зарлага' in mj2.columns
            if _has_orl or _has_zarl:
                _ycols = [c for c in ['Орлого','Зарлага'] if c in mj2.columns]
                # Жилээрх харьцуулалт
                if 'Он' in mj2.columns and len(mj2['Он'].dropna().unique()) > 1:
                    yr_sum = mj2.groupby('Он')[_ycols].sum().reset_index()
                    fig2 = px.bar(yr_sum, x='Он', y=_ycols, barmode='group',
                                  title='Жилээрх орлого/зарлага', color_discrete_sequence=['#2ecc71','#e74c3c'])
                    st.plotly_chart(fig2, use_container_width=True)

                # Топ 15 их дүнтэй гүйлгээ
                if _has_orl:
                    top_income = mj2.nlargest(15, 'Орлого')[['Огноо','Орлого','Тайлбар','Харилцагч'] if 'Тайлбар' in mj2.columns and 'Харилцагч' in mj2.columns else [c for c in ['Огноо','Орлого'] if c in mj2.columns]]
                    st.markdown("**🔝 Топ 15 их орлого**")
                    st.dataframe(top_income, use_container_width=True, hide_index=True)
                if _has_zarl:
                    top_expense = mj2.nlargest(15, 'Зарлага')[['Огноо','Зарлага','Тайлбар','Харилцагч'] if 'Тайлбар' in mj2.columns and 'Харилцагч' in mj2.columns else [c for c in ['Огноо','Зарлага'] if c in mj2.columns]]
                    st.markdown("**🔝 Топ 15 их зарлага**")
                    st.dataframe(top_expense, use_container_width=True, hide_index=True)
            else:
                st.info("Орлого/Зарлага баганыг тодорхойлж чадсангүй.")

        with tab_partner:
            if 'Харилцагч' in mj2.columns:
                _cp = mj2.groupby('Харилцагч').agg(
                    Тоо=('Харилцагч','count'),
                    Орлого_нийт=('Орлого','sum') if 'Орлого' in mj2.columns else ('Харилцагч','count'),
                    Зарлага_нийт=('Зарлага','sum') if 'Зарлага' in mj2.columns else ('Харилцагч','count'),
                ).reset_index().sort_values('Тоо', ascending=False)
                st.dataframe(_cp.head(50), use_container_width=True, hide_index=True)

                fig3 = px.bar(_cp.head(15), x='Харилцагч', y='Тоо', title='Топ 15 харилцагч (гүйлгээний тоо)',
                              color_discrete_sequence=['#0F4C81'])
                st.plotly_chart(fig3, use_container_width=True)
            else:
                st.info("Харилцагч баганыг тодорхойлж чадсангүй.")

        # ── Татах ──
        buf = io.BytesIO()
        mj2.to_excel(buf, index=False, engine='openpyxl')
        st.download_button("📥 МЖ тайлан татах", buf.getvalue(), file_name='МЖ_тайлан.xlsx', key='dl_mj')
    else:
        st.info("📎 Дээрх хэсэгт МЖ файлаа оруулна уу, эсвэл 1️⃣ цэсэнд оруулсан бол автоматаар ачаалагдана.")
        st.metric("Нийт гүйлгээ", "—")



# ═══ 4️⃣ ЭДИЙН ЗАСГИЙН АНГИЛАЛ ═══
if page.startswith("📊 ЭЗ"):
    st.header("4️⃣ Эдийн засгийн ангилал — Орлого, зарлагын нийцэл")
    st.markdown("*Сангийн сайдын ЭЗ ангилалтай таарч байна уу? Бодлоготой нийцэж байна уу?*")
    ej = st.session_state.get('g_ej')
    ez = st.session_state.get('g_ez')
    policy = st.session_state.get('g_policy')
    budget = st.session_state.get('g_budget')

    if ej is None:
        st.info("1️⃣ цэсэнд ЕЖ + ЭЗ ангилалын файл оруулна уу.")
    else:
        if ez is not None and not ez.empty:
            st.markdown("### 📊 ЭЗ ангилалын нийцэл")
            # ЕЖ дотор ЭЗ код байгаа эсэхийг шалгах
            if 'ez_code' in ej.columns:
                ej_ez = ej[ej['ez_code'].astype(str).str.strip() != ''].copy()
                no_ez = ej[ej['ez_code'].astype(str).str.strip() == ''].copy()
                _ezc = st.container()
                m1, m2 = _ezc.columns(2)
                m1.metric("ЭЗ кодтой гүйлгээ", f"{len(ej_ez):,}")
                m2.metric("❌ ЭЗ кодгүй гүйлгээ", f"{len(no_ez):,}")
                if not no_ez.empty:
                    st.warning(f"⚠️ {len(no_ez):,} гүйлгээнд ЭЗ ангилалын код байхгүй!")
                    show = [c for c in ['transaction_date','account_code','counterparty_name','amount','transaction_description'] if c in no_ez.columns]
                    st.dataframe(no_ez[show].head(200), use_container_width=True, hide_index=True)
            st.dataframe(ez, use_container_width=True, hide_index=True)

        # Төсвийн хэтрэлт
        if budget is not None and not budget.empty and ej is not None:
            st.markdown("### 📐 Төсвийн гүйцэтгэл vs Батлагдсан")
            # ЕЖ-ээс зардлын дансны эргэлтийг тооцох
            ej_expense = ej[ej['account_code'].astype(str).str.startswith('6')].copy()
            if not ej_expense.empty:
                actual = ej_expense.groupby('ez_code')['debit_mnt'].sum().reset_index().rename(columns={'debit_mnt':'Гүйцэтгэсэн'})
                bud_col = [c for c in budget.columns if 'батлагдсан' in str(c).lower() or 'дүн' in str(c).lower()]
                ez_col = [c for c in budget.columns if 'код' in str(c).lower()]
                if bud_col and ez_col:
                    bud2 = budget[[ez_col[0], bud_col[0]]].rename(columns={ez_col[0]:'ez_code', bud_col[0]:'Батлагдсан'})
                    bud2['ez_code'] = bud2['ez_code'].astype(str)
                    actual['ez_code'] = actual['ez_code'].astype(str)
                    comp = actual.merge(bud2, on='ez_code', how='outer').fillna(0)
                    comp['Зөрүү'] = comp['Гүйцэтгэсэн'] - comp['Батлагдсан']
                    comp['Хэтрэлт %'] = (comp['Зөрүү'] / comp['Батлагдсан'].replace(0, np.nan) * 100).round(1)
                    comp['Үр дүн'] = comp['Зөрүү'].apply(lambda x: '🔴 ХЭТЭРСЭН' if x > 0 else '✅ Хүрээнд')
                    st.dataframe(comp, use_container_width=True, hide_index=True)
                    exceeded = comp[comp['Зөрүү'] > 0]
                    if not exceeded.empty:
                        st.error(f"🔴 {len(exceeded)} ЭЗ ангилалд төсвийн хэтрэлт илэрсэн!")
                        buf = io.BytesIO()
                        exceeded.to_excel(buf, index=False)
                        st.download_button("📥 Хэтрэлтийн жагсаалт", buf.getvalue(), file_name='Төсвийн_хэтрэлт.xlsx', key='dl_budget')

        # Бодлогын нийцэл
        if policy is not None:
            st.markdown("### 📜 Бодлогын бичиг баримт")
            st.dataframe(policy, use_container_width=True, hide_index=True)



# ═══ 5️⃣ ХУУЛИЙН ЗӨРЧИЛ ШАЛГАЛТ ═══
if page.startswith("⚖️"):
    st.header("5️⃣ Хуулийн заалтуудын зөрчил шалгалт")
    st.markdown("*Төрийн хэмнэлт • Худалдан авах ажиллагаа • Шилэн данс • Төсвийн тухай*")
    laws = st.session_state.get('g_laws')
    ej = st.session_state.get('g_ej')
    _has5 = laws is not None and ej is not None
    if not _has5:
        st.info("1️⃣ цэсэнд хуулийн заалтууд + ЕЖ файл оруулна уу.")
    if st.button("🔍 Хуулийн зөрчил шалгах", type='primary', use_container_width=True, key='g_law_run', disabled=not _has5):
        violations = []
        st.markdown("### ⚖️ Хуулийн заалтууд")
        st.dataframe(laws, use_container_width=True, hide_index=True)
        for _, law in laws.iterrows():
            law_code = str(law.get('Хууль','')) + ' ' + str(law.get('Зүйл',''))
            check_field = str(law.get('Шүүх шалгуур',''))
            acct_codes = str(law.get('Дансны код','')).split(',')
            acct_codes = [c.strip() for c in acct_codes if c.strip()]

            # Дансны кодоор шүүх
            if acct_codes:
                matched = ej[ej['account_code'].astype(str).isin(acct_codes)]
                if not matched.empty:
                    total = matched['debit_mnt'].sum() + matched['credit_mnt'].sum()
                    # Тодорхой зөрчлүүд
                    if 'ногдол' in check_field.lower() or 'ногдол' in str(law.get('Заалтын агуулга','')).lower():
                        nogdol = ej[ej['transaction_description'].str.contains('ногдол|ноогдол', case=False, na=False)]
                        if not nogdol.empty:
                            violations.append({'Хууль':law_code, 'Заалт':str(law.get('Заалтын агуулга',''))[:60],
                                              'Зөрчлийн тоо':len(nogdol), 'Дүн':nogdol['amount'].sum(),
                                              'Тайлбар':f'Ногдол ашиг олгосон {len(nogdol)} гүйлгээ'})
                    elif 'урамшуулал' in check_field.lower():
                        reward = ej[ej['transaction_description'].str.contains('урамшуулал|шагнал|bonus', case=False, na=False)]
                        if not reward.empty:
                            violations.append({'Хууль':law_code, 'Заалт':str(law.get('Заалтын агуулга',''))[:60],
                                              'Зөрчлийн тоо':len(reward), 'Дүн':reward['amount'].sum(),
                                              'Тайлбар':f'Урамшуулал олгосон {len(reward)} гүйлгээ'})
                    elif '5M' in check_field or '5 сая' in check_field.lower():
                        big = matched[matched['amount'] >= 5e6]
                        if not big.empty:
                            violations.append({'Хууль':law_code, 'Заалт':str(law.get('Заалтын агуулга',''))[:60],
                                              'Зөрчлийн тоо':len(big), 'Дүн':big['amount'].sum(),
                                              'Тайлбар':f'5 сая+ гүйлгээ {len(big)} ширхэг → тендер/мэдээлэх шаардлагатай'})

        if violations:
            vio_df = pd.DataFrame(violations)
            st.error(f"🔴 {len(violations)} зөрчил илэрсэн!")
            st.dataframe(vio_df, use_container_width=True, hide_index=True)
            buf = io.BytesIO()
            vio_df.to_excel(buf, index=False)
            st.download_button("📥 Зөрчлийн жагсаалт", buf.getvalue(), file_name='Хуулийн_зөрчил.xlsx', key='dl_law')
        else:
            st.success("✅ Хуулийн зөрчил илрээгүй.")



# ═══ 6️⃣ ЗАРДЛЫН АНГИЛАЛ ═══
if page.startswith("💰"):
    st.header("6️⃣ Зардлын ангилал — Сангийн сайдын тушаалтай уях")
    ec = st.session_state.get('g_expense_class')
    ej = st.session_state.get('g_ej')
    _has6 = ec is not None and ej is not None
    if not _has6:
        st.info("1️⃣ цэсэнд зардлын ангилал + ЕЖ файл оруулна уу.")
    if st.button("🔍 Зардлын ангилал шалгах", type='primary', use_container_width=True, key='g_ec_run', disabled=not _has6):
        st.dataframe(ec, use_container_width=True, hide_index=True)
        ej_expense = ej[ej['account_code'].astype(str).str.startswith('6')].copy()
        if ej_expense.empty:
            st.warning("Зардлын данс (6xxxx) олдсонгүй.")
        else:
            ec_map = {}
            acct_col = [c for c in ec.columns if any(k in str(c).lower() for k in ['дансны код', 'зардлын код', 'code', 'код'])]
            name_col = [c for c in ec.columns if any(k in str(c).lower() for k in ['зардлын нэр', 'ангилал', 'description', 'тайлбар', 'нэр'])]
            if acct_col and name_col:
                for _, r in ec.iterrows():
                    ec_map[str(r[acct_col[0]])] = str(r[name_col[0]])
            ej_expense['Зардлын ангилал'] = ej_expense['account_code'].astype(str).map(ec_map).fillna('❌ ТААРУУЛЖ ЧАДСАНГҮЙ')
            mismatch = ej_expense[ej_expense['Зардлын ангилал'].str.contains('❌')]
            _c6 = st.container()
            m1, m2 = _c6.columns(2)
            m1.metric("✅ Тааруулсан", f"{len(ej_expense)-len(mismatch):,}")
            m2.metric("❌ Таарахгүй", f"{len(mismatch):,}")
            if not mismatch.empty:
                st.warning(f"⚠️ {len(mismatch):,} зардлын гүйлгээ ангилалтай таарахгүй!")
                show = [c for c in ['transaction_date','account_code','account_name','amount','transaction_description','Зардлын ангилал'] if c in mismatch.columns]
                st.dataframe(mismatch[show].head(200), use_container_width=True, hide_index=True)
                buf = io.BytesIO()
                mismatch.to_excel(buf, index=False)
                st.download_button("📥 Таарахгүй зардал", buf.getvalue(), file_name='Зардал_зөрүү.xlsx', key='dl_ec')



# ═══ 8️⃣ ХАРИЛЦАГЧИЙН ШИНЖИЛГЭЭ ═══
if page.startswith("🔍 Харилцагч"):
    st.header("8️⃣ Харилцагчийн шинжилгээ — ISA 550")
    ej = st.session_state.get('g_ej')
    if ej is None:
        st.info("1️⃣ цэсэнд ЕЖ файл оруулна уу.")
    else:
        cp_col = 'counterparty_name' if 'counterparty_name' in ej.columns else None
        if cp_col:
            cp_sum = ej.groupby(cp_col).agg(
                Гүйлгээний_тоо=(cp_col,'count'),
                Дүн=('amount','sum'),
                Данс_тоо=('account_code','nunique'),
            ).reset_index().sort_values('Дүн', ascending=False)
            cp_sum['Ховор'] = cp_sum['Гүйлгээний_тоо'].apply(lambda x: '⚠️ Ховор' if x <= 3 else '')

            m1, m2, m3 = st.columns(3)
            m1.metric("Нийт харилцагч", f"{len(cp_sum):,}")
            m2.metric("⚠️ Ховор (≤3 удаа)", f"{len(cp_sum[cp_sum['Ховор']!=''])}")
            m3.metric("Нийт дүн", f"₮{cp_sum['Дүн'].sum():,.0f}")

            st.dataframe(cp_sum.head(100), use_container_width=True, hide_index=True)
            fig = px.bar(cp_sum.head(20), x=cp_col, y='Дүн', color='Ховор', title='Топ 20 харилцагч (дүнгээр)')
            fig.update_layout(xaxis_tickangle=-30)
            st.plotly_chart(fig, use_container_width=True)

            buf = io.BytesIO()
            cp_sum.to_excel(buf, index=False)
            st.download_button("📥 Харилцагчийн тайлан", buf.getvalue(), file_name='Харилцагч.xlsx', key='dl_cp')



# ═══ 📊 ЭРСДЭЛИЙН НЭГТГЭЛ ═══
if page.startswith("📋 Эрсдэлийн"):
    st.header("📊 Эрсдэлийн нэгтгэл")
    risks = []
    if st.session_state.get('g_ej_done'):
        ml = st.session_state.get('g_ej_ml')
        if ml is not None and 'ml_anomaly_flag' in ml.columns:
            n = int(ml['ml_anomaly_flag'].sum())
            risks.append({'Шинжилгээ':'📘 ЕЖ аномали','ISA':'ISA 240,500,550','Тоо':n,'Тайлбар':f'{n} аномали гүйлгээ'})
    if risks:
        st.dataframe(pd.DataFrame(risks), use_container_width=True, hide_index=True)
        total = sum(r['Тоо'] for r in risks)
        st.metric("Нийт эрсдэл", f"{total:,}")
    else:
        st.info("Шинжилгээ хийгдээгүй. 2-8 цэсүүдэд шинжилгээ хийнэ үү.")

st.markdown("---")
st.caption("Төрийн аудит СТА v1.0 © 2026 — МС/ХОУ суурьтай нийцлийн шалгалтын систем")


